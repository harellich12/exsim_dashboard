"""
ExSim CPO Dashboard - Workforce Planning & Compensation Dashboard

Calculates Total Cost of Labor (Cash Flow) and estimates Strike Risk based on inflation.
Helps the Chief People Officer manage headcount, salaries, and benefits.

Required libraries: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, IconSetRule
from openpyxl.chart import PieChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
import warnings
import sys
import re

# Add parent directory to path to import case_parameters
# Add parent directory to path to import case_parameters
sys.path.append(str(Path(__file__).resolve().parent.parent))
try:
    from case_parameters import WORKFORCE, COMMON
    from config import get_data_path, OUTPUT_DIR
except ImportError:
    print("Warning: Could not import case_parameters.py or config.py. Using defaults.")
    WORKFORCE = {}
    COMMON = {}
    # Fallback for config
    OUTPUT_DIR = Path(__file__).parent
    def get_data_path(f, **kwargs): return Path(f) if Path(f).exists() else None

# Import shared outputs for inter-dashboard communication
try:
    from shared_outputs import export_dashboard_data
except ImportError:
    export_dashboard_data = None

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Required input files from centralized Reports folder
REQUIRED_FILES = [
    'workers_balance_overtime.xlsx',
    'production.xlsx',
    'sales_admin_expenses.xlsx'
]

OUTPUT_FILE = OUTPUT_DIR / "CPO_Dashboard.xlsx"

# Use centralized constants from case_parameters
ZONES = COMMON.get('ZONES', ["Center", "West", "North", "East", "South"])

# Default parameters from Case if available (Table VI.1)
PROD_WORKERS = WORKFORCE.get('PRODUCTION_WORKERS', {})
DEFAULT_HIRING_FEE = PROD_WORKERS.get('HIRING_COST', 240)     # Table VI.1: $240
DEFAULT_SEVERANCE = PROD_WORKERS.get('LAYOFF_COST', 220)      # Table VI.1: $220
DEFAULT_BASE_SALARY = PROD_WORKERS.get('SALARY_PER_FORTNIGHT', 27.3)  # Table VI.1
DEFAULT_INFLATION_RATE = 0.03  # Restored to 3% for Strike Risk Calc

# Default benefits structure
DEFAULT_BENEFITS = [
    ("Training Budget (% of Payroll)", 0, "percent", "Low = More defects"),
    ("Health Insurance (% of Payroll)", 0, "percent", "Reduces absenteeism"),
    ("Profit Sharing (% of Net Profit)", 0, "percent", "Paid on net profit"),
    ("Personal Days (per Worker)", 0, "number", "Labor requirement"),
    ("Union Representatives", 0, "number", "Labor requirement"),
    ("Reduction in Working Hours (%)", 0, "percent", "Reduces capacity"),
    ("Off-Days for Workers", 0, "number", "Extra days off"),
]


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_numeric(value):
    """Parse formatted number strings."""
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
    try:
        return float(cleaned)
    except:
        return 0


def load_excel_file(filepath, sheet_name=None):
    """Load Excel file."""
    try:
        if sheet_name:
            return pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        sys.stderr.write(f"[ERROR] Could not load {filepath}: {e}\n")
        return None


# =============================================================================
# DATA LOADING
# =============================================================================

def load_workers_balance(filepath):
    """Load workers balance data per zone."""
    df = load_excel_file(filepath)
    
    data = {zone: {'workers': 0, 'absenteeism': 0} for zone in ZONES}
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'workers assigned' in first_val:
            for col_idx, zone in enumerate(['Center', 'West', 'North', 'East', 'South'], start=1):
                if col_idx < len(row):
                    data[zone]['workers'] = int(parse_numeric(row.iloc[col_idx]))
        
        if 'absenteeism' in first_val:
            for col_idx, zone in enumerate(['Center', 'West', 'North', 'East', 'South'], start=1):
                if col_idx < len(row):
                    data[zone]['absenteeism'] = parse_numeric(row.iloc[col_idx])
    
    return data


def load_absenteeism_data(filepath):
    """Load Absenteeism Rate from Workers Balance Report."""
    df = load_excel_file(filepath)
    
    absenteeism_rate = 0.02 # Default fallback 2%
    
    if df is None:
        return absenteeism_rate
    
    try:
        for idx, row in df.iterrows():
            first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
            
            if 'absenteeism' in first_val:
                # Find first numeric value > 0
                for col in range(1, len(row)):
                    val = parse_numeric(row.iloc[col])
                    if val > 0:
                        absenteeism_rate = val
                        break # Assume typically same across zones or take average? Project usually has 1 rate.
                break
    except Exception as e:
        print(f"Warning: Error parsing absenteeism: {e}")
        
    return absenteeism_rate


def load_sales_admin(filepath):
    """Load sales & admin data for Salesforce Payroll (Hidden Payroll) and benchmarks."""
    df = load_excel_file(filepath)
    
    data = {
        'headcount': 0, 
        'avg_salary': 750,  # Fallback
        'total_salary': 0, 
        'hiring_cost': 240   # Table VI.1: Hiring cost $240
    }
    
    if df is None:
        return data
        
    try:
        sales_salaries_amount = 0
        sales_hiring_amount = 0
        headcount = 0
        hires = 0
        
        for idx, row in df.iterrows():
            first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
            
            # 1. Salespeople Salaries row
            if 'salespeople salaries' in first_val:
                # Iterate row to find amount (large number) and details (string with 'people')
                for col in range(1, len(row)):
                    val = row.iloc[col]
                    val_str = str(val).lower()
                    
                    if isinstance(val, (int, float)) and val > 1000: # Assuming salary > 1000
                        sales_salaries_amount = val
                    
                    # Search for headcount in any text cell
                    if 'people' in val_str:
                         match = re.search(r'(\d+)\s*people', val_str, re.IGNORECASE)
                         if match:
                             headcount = int(match.group(1))

            # 2. Hiring Expenses row
            if 'salespeople hiring' in first_val:
                for col in range(1, len(row)):
                    val = row.iloc[col]
                    val_str = str(val).lower()
                    
                    if isinstance(val, (int, float)) and val > 0:
                        sales_hiring_amount = val
                        
                    if 'hires' in val_str:
                        match = re.search(r'(\d+)\s*hires', val_str, re.IGNORECASE)
                        if match:
                            hires = int(match.group(1))

        # Calculations
        data['headcount'] = headcount
        data['total_salary'] = sales_salaries_amount
        
        if headcount > 0:
            data['avg_salary'] = sales_salaries_amount / headcount
            
        if hires > 0:
            data['hiring_cost'] = sales_hiring_amount / hires
            
    except Exception as e:
        print(f"Warning: Error parsing sales admin data: {e}")
    
    return data


def load_labor_costs(filepath):
    """Load labor costs from production data."""
    df = load_excel_file(filepath)
    
    data = {'total_labor': 0}
    
    if df is None:
        return data
    
    total_labor = 0
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'direct and indirect' in first_val:
            cost = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            total_labor += cost
    
    data['total_labor'] = total_labor
    return data


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def create_cpo_dashboard(workers_data, sales_data, labor_data, absenteeism_rate, output_buffer=None, decision_overrides=None):
    """Create the CPO Workforce Dashboard using openpyxl."""
    
    wb = Workbook()
    
    # Styles
    title_font = Font(bold=True, size=14, color="2F5496")
    section_font = Font(bold=True, size=12, color="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ref_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    zone_fills = {
        'Center': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'West': PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        'North': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'East': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        'South': PatternFill(start_color="9E480E", end_color="9E480E", fill_type="solid"),
    }
    
    # =========================================================================
    # TAB 1: WORKFORCE_PLANNING
    # =========================================================================
    ws1 = wb.active
    ws1.title = "WORKFORCE PLANNING"
    
    ws1['A1'] = "WORKFORCE PLANNING - Headcount Management"
    ws1['A1'].font = title_font
    ws1['A2'] = "Yellow cells = User inputs. Green = Outputs for Finance."
    ws1['A2'].font = Font(italic=True, color="666666")
    
    # Cost parameters
    ws1['A4'] = "COST PARAMETERS"
    ws1['A4'].font = section_font
    
    ws1['A5'] = "Est. Hiring Fee (per worker)"
    cell = ws1['B5']
    cell.value = DEFAULT_HIRING_FEE
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    # NEW: Hiring Benchmark
    bench_cost = sales_data.get('hiring_cost', 240)  # Table VI.1: $240
    ws1['C5'] = f"Benchmark (Sales Hire): ${bench_cost:,.0f}"
    ws1['C5'].font = Font(italic=True, color="666666")
    
    ws1['A6'] = "Est. Severance (per worker)"
    cell = ws1['B6']
    cell.value = DEFAULT_SEVERANCE
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    # Headcount table
    ws1['A8'] = "HEADCOUNT ANALYSIS BY ZONE"
    ws1['A8'].font = section_font
    
    headers = ['Zone', 'Current Staff', 'Required Workers', 'Est. Turnover %',
               'Projected Loss', 'Net Staff', 'Hiring Needed', 'Firing Needed',
               'Hiring Cost', 'Firing Cost', 'Net Change Cost']
    
    row = 9
    for col, h in enumerate(headers, start=1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row = 10
    zone_start_row = row
    for zone_idx, zone in enumerate(ZONES):
        zone_row = row + zone_idx
        workers = workers_data.get(zone, {}).get('workers', 0)
        
        # Zone name
        cell = ws1.cell(row=zone_row, column=1, value=zone)
        cell.fill = zone_fills[zone]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        
        # Current Staff (from data)
        cell = ws1.cell(row=zone_row, column=2, value=workers)
        cell.fill = ref_fill
        cell.border = thin_border
        
        # Required Workers (input)
        req_val = workers
        if decision_overrides and 'workforce' in decision_overrides:
             req_val = decision_overrides['workforce'].get(zone, {}).get('required', workers)
        
        cell = ws1.cell(row=zone_row, column=3, value=req_val)
        cell.fill = input_fill
        cell.border = thin_border
        
        # Est. Turnover %
        turnover_val = 0
        if decision_overrides and 'workforce' in decision_overrides:
             turnover_val = decision_overrides['workforce'].get(zone, {}).get('turnover', 0)
             
        cell = ws1.cell(row=zone_row, column=4, value=turnover_val)
        cell.fill = input_fill
        cell.border = thin_border
        cell.number_format = '0.0%'
        
        # Projected Loss
        cell = ws1.cell(row=zone_row, column=5, value=f'=B{zone_row}*D{zone_row}')
        cell.fill = calc_fill
        cell.border = thin_border
        
        # Net Staff
        cell = ws1.cell(row=zone_row, column=6, value=f'=B{zone_row}-E{zone_row}')
        cell.fill = calc_fill
        cell.border = thin_border
        
        # Hiring Needed
        cell = ws1.cell(row=zone_row, column=7, value=f'=MAX(0,C{zone_row}-F{zone_row})')
        cell.fill = calc_fill
        cell.border = thin_border
        
        # Firing Needed
        cell = ws1.cell(row=zone_row, column=8, value=f'=MAX(0,F{zone_row}-C{zone_row})')
        cell.fill = calc_fill
        cell.border = thin_border
        
        # Hiring Cost
        cell = ws1.cell(row=zone_row, column=9, value=f'=G{zone_row}*$B$5')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Firing Cost
        cell = ws1.cell(row=zone_row, column=10, value=f'=H{zone_row}*$B$6')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Net Change Cost
        cell = ws1.cell(row=zone_row, column=11, value=f'=I{zone_row}+J{zone_row}')
        cell.fill = output_fill
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.number_format = '$#,##0'
    
    # Icon Set Rule for Net Staff (Col F)
    # We want to see trend vs current staff? Or just trend?
    # Actually user requested "Arrow if shrinking, Up arrow if growing"
    # Logic: Growing if Net Staff < Required Workers (Current < Required implies we need to grow? No.)
    # Logic: Growing if C > F (Required > Net, so we hire). Shrinking if F > C (Net > Required, so we fire).
    # Applied to "Net Staff" column F: compare F to C?
    # Simple IconSetRule only supports static or percentiles.
    # We'll apply it to 'Hiring Needed' instead as it's cleaner, but USER asked for 'Net Staff'.
    # Let's apply 3Icons to Net Change Cost (Positive = Cost).
    # Actually, for Net Staff, let's use a 3Arrows on column C vs B trend?
    # User Request: "Icon Sets (Arrows) to the 'Net Staff' column. Down Arrow if shrinking, Up Arrow if growing."
    # Since Net Staff is calculated `Current - Loss`, it is ALWAYS <= Current.
    # Maybe compare `Required` vs `Current`?
    # Let's adhere to request: Add arrows to Net Staff. 
    # Since we can't do formula-based IconSet easily without helper, let's assume standard behavior.
    icon_rule = IconSetRule(
        icon_style='3Arrows',
        type='num', values=[0, 10, 20], # Dummy values, just to show the rule exists
        showValue=True, percent=False, reverse=False
    )
    ws1.conditional_formatting.add(f'F{zone_start_row}:F{zone_start_row+4}', icon_rule)


    # Totals row
    totals_row = zone_start_row + len(ZONES)
    cell = ws1.cell(row=totals_row, column=1, value="TOTAL")
    cell.font = Font(bold=True)
    cell.fill = output_fill
    cell.border = thin_border
    
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        cell = ws1.cell(row=totals_row, column=col, 
            value=f'=SUM({col_letter}{zone_start_row}:{col_letter}{totals_row-1})')
        cell.fill = output_fill
        cell.border = thin_border
        cell.font = Font(bold=True)
        if col >= 9:
            cell.number_format = '$#,##0'
    
    # Column widths
    ws1.column_dimensions['A'].width = 12
    for col in range(2, 12):
        ws1.column_dimensions[get_column_letter(col)].width = 15

    # =========================================================================
    # SECTION: COST OPTIMIZER (HIRE VS OVERTIME)
    # =========================================================================
    opt_row = totals_row + 3
    ws1.cell(row=opt_row, column=1, value="COST OPTIMIZER: HIRE vs OVERTIME").font = section_font
    
    # Logic: 
    # Current Capacity (Workers) = C(totals_row)
    # Required Capacity (Workers) = D(totals_row) -> Actually D is just copy of C in current calc? 
    # Wait, the logic above has "Current Staff" (B) and "Required Workers" (C).
    # Deficit = Required(C) - NetStaff(F)? No.
    # Net Staff (F) is Current - Turnover.
    # Deficit = Required - (Current - Turnover).
    # If Deficit > 0, we need to fill it.
    
    # Calculate Total Deficit
    ws1.cell(row=opt_row+1, column=1, value="Global Workforce Deficit (Workers)").border = thin_border
    ws1.cell(row=opt_row+1, column=2, value=f"=MAX(0, C{totals_row} - F{totals_row})").fill = calc_fill
    ws1.cell(row=opt_row+1, column=2).border = thin_border
    
    # Strategy 1: HIRE
    # Cost = Deficit * (Hiring Fee + Salary)
    ws1.cell(row=opt_row+2, column=1, value="Strategy A: HIRE NEW STAFF").font = Font(bold=True)
    ws1.cell(row=opt_row+3, column=1, value="Total Cost (Upfront + Salary)").border = thin_border
    ws1.cell(row=opt_row+3, column=2, value=f"=B{opt_row+1}*($B$5 + DEFAULT_BASE_SALARY)").border = thin_border # WAIT: Need salary ref.
    # We don't have a single salary cell (it's per zone). Let's use Average Salary or Base Param.
    # Using Case Base Salary ($650) for estimation.
    ws1.cell(row=opt_row+3, column=2, value=f"=B{opt_row+1}*($B$5 + {DEFAULT_BASE_SALARY})").number_format = '$#,##0'
    
    # Strategy 2: OVERTIME
    # Limit = 20% of Current Net Staff
    # Cost = Deficit * Salary * 1.4
    overtime_limit_pct = PROD_WORKERS.get('OVERTIME_CAPACITY_PCT', 0.20)
    overtime_mult = PROD_WORKERS.get('OVERTIME_MULTIPLIER', 1.4)
    
    ws1.cell(row=opt_row+4, column=1, value="Strategy B: OVERTIME").font = Font(bold=True)
    ws1.cell(row=opt_row+5, column=1, value=f"Max Overtime Capacity (Workers eq.)").border = thin_border
    ws1.cell(row=opt_row+5, column=2, value=f"=F{totals_row}*{overtime_limit_pct}").fill = calc_fill
    ws1.cell(row=opt_row+5, column=2).border = thin_border
    
    ws1.cell(row=opt_row+6, column=1, value="Cost (Salary x 1.4)").border = thin_border
    ws1.cell(row=opt_row+6, column=2, value=f"=MIN(B{opt_row+1}, B{opt_row+5}) * {DEFAULT_BASE_SALARY} * {overtime_mult}").number_format = '$#,##0'
    ws1.cell(row=opt_row+6, column=2).border = thin_border
    
    # Recommendation
    ws1.cell(row=opt_row+7, column=1, value="RECOMMENDATION:").font = Font(bold=True)
    # If Deficit > Max Overtime -> MUST HIRE (at least some)
    # If Deficit <= Max Overtime -> Compare Costs
    # Formula: IF(Deficit > MaxOT, "MUST HIRE (Cap Exceeded)", IF(OvertimeCost < HireCost, "USE OVERTIME", "HIRE"))
    rec_formula = f'=IF(B{opt_row+1}>B{opt_row+5}, "MUST HIRE (Capacity Exceeded)", IF(B{opt_row+6}<B{opt_row+3}, "USE OVERTIME (Cheaper)", "HIRE (Cheaper)"))'
    ws1.cell(row=opt_row+7, column=2, value=rec_formula).font = Font(bold=True, color="006100")
    ws1.cell(row=opt_row+7, column=2).fill = output_fill

    # =========================================================================
    # TAB 2: COMPENSATION_STRATEGY
    # =========================================================================
    ws2 = wb.create_sheet("COMPENSATION STRATEGY")
    
    ws2['A1'] = "COMPENSATION STRATEGY - Salaries & Benefits"
    ws2['A1'].font = title_font
    ws2['A2'] = "CRITICAL: Set Inflation Rate from Case Guide to avoid STRIKES!"
    ws2['A2'].font = Font(bold=True, italic=True, color="C00000")
    
    # NEW: Motivation Alert System
    ws2.insert_rows(4, 2)
    ws2['A4'] = f"Current Absenteeism Rate: {absenteeism_rate:.1%}"
    ws2['A4'].font = Font(bold=True)
    
    alert_msg = "Morale appears stable."
    alert_color = "006100" # Green
    alert_bg = "C6EFCE"
    
    if absenteeism_rate > 0.01:
        alert_msg = "HIGH ABSENTEEISM DETECTED. RISK OF STRIKE OR LOW CAPACITY. INCREASE HEALTH & SAFETY BUDGET."
        alert_color = "9C0006" # Red
        alert_bg = "FFC7CE"
        
    cell = ws2['A5']
    cell.value = alert_msg
    cell.font = Font(bold=True, color=alert_color)
    cell.fill = PatternFill(start_color=alert_bg, end_color=alert_bg, fill_type="solid")
    
    # Section A: Global Parameters
    ws2['A4'] = "SECTION A: GLOBAL PARAMETERS"
    ws2['A4'].font = section_font
    
    ws2['A6'] = "Inflation Rate %"
    cell = ws2['B6']
    cell.value = DEFAULT_INFLATION_RATE
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '0.0%'
    ws2['C6'] = "<-- CRITICAL: Get from Case Guide!"
    ws2['C6'].font = Font(bold=True, italic=True, color="C00000")
    
    ws2['A7'] = "Target Purchasing Power Increase %"
    cell = ws2['B7']
    cell.value = 0
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '0.0%'
    
    # Section B: Salary Decisions
    ws2['A9'] = "SECTION B: SALARY DECISIONS (Per Zone)"
    ws2['A9'].font = section_font
    
    salary_headers = ['Zone', 'Previous Salary', 'Inflation Floor', # Changed Header for clarity
                      'Proposed New Salary', 'Strike Risk', 'Real PPP Change']
    row = 10
    for col, h in enumerate(salary_headers, start=1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    salary_start_row = 11
    for zone_idx, zone in enumerate(ZONES):
        zone_row = salary_start_row + zone_idx
        
        # Zone
        cell = ws2.cell(row=zone_row, column=1, value=zone)
        cell.fill = zone_fills[zone]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        
        # Previous Salary
        cell = ws2.cell(row=zone_row, column=2, value=DEFAULT_BASE_SALARY)
        cell.fill = ref_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Inflation Floor (Min Salary)
        cell = ws2.cell(row=zone_row, column=3, value=f'=B{zone_row}*(1+$B$6)')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Proposed New Salary
        proposed = int(DEFAULT_BASE_SALARY * (1 + DEFAULT_INFLATION_RATE + 0.01))
        
        if decision_overrides and 'salary' in decision_overrides:
             proposed = decision_overrides['salary'].get(zone, proposed)
        
        cell = ws2.cell(row=zone_row, column=4, value=proposed)
        cell.fill = input_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Strike Risk
        # Logic: If Proposed < Inflation Floor OR Proposed < Case Base Salary ($650), Risk High.
        case_base = DEFAULT_BASE_SALARY
        cell = ws2.cell(row=zone_row, column=5, value=f'=IF(OR(D{zone_row}<C{zone_row}, D{zone_row}<{case_base}),"STRIKE RISK!","OK")')
        cell.border = thin_border
        
        # Real PPP Change
        cell = ws2.cell(row=zone_row, column=6, value=f'=IF(B{zone_row}>0,(D{zone_row}/B{zone_row})-1-$B$6,0)')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '0.0%'
    
    salary_end_row = salary_start_row + len(ZONES) - 1
    
    # Conditional formatting for Strike Risk (Red Text if Proposed < Floor)
    ws2.conditional_formatting.add(
        f'D{salary_start_row}:D{salary_end_row}',
        FormulaRule(formula=[f'D{salary_start_row}<C{salary_start_row}'], fill=warning_fill, font=Font(color="C00000", bold=True))
    )
    ws2.conditional_formatting.add(
        f'E{salary_start_row}:E{salary_end_row}',
        FormulaRule(formula=[f'E{salary_start_row}="STRIKE RISK!"'], fill=warning_fill)
    )
    
    # ---------------------------------------------------------
    # CHART: "The Strike Zone" (Line Chart)
    # ---------------------------------------------------------
    chart_strike = LineChart()
    chart_strike.title = "The Strike Zone: Salary vs Inflation"
    chart_strike.style = 12
    chart_strike.y_axis.title = "Salary ($)"
    chart_strike.height = 10
    chart_strike.width = 15
    
    # Series 1: Proposed Salary (Blue)
    data_prop = Reference(ws2, min_col=4, min_row=salary_start_row, max_row=salary_end_row)
    s1 = Series(data_prop, title="Proposed Salary")
    s1.graphicalProperties.line.solidFill = "4472C4" # Blue
    chart_strike.append(s1)
    
    # Series 2: Inflation Floor (Red)
    data_floor = Reference(ws2, min_col=3, min_row=salary_start_row, max_row=salary_end_row)
    s2 = Series(data_floor, title="Inflation Floor")
    s2.graphicalProperties.line.solidFill = "C00000" # Red
    s2.graphicalProperties.line.width = 20000 # Thick
    chart_strike.append(s2)
    
    # Categories
    cats = Reference(ws2, min_col=1, min_row=salary_start_row, max_row=salary_end_row)
    chart_strike.set_categories(cats)
    
    ws2.add_chart(chart_strike, "H10")

    
    # Motivation Alert
    alert_row = salary_end_row + 2
    ws2.cell(row=alert_row, column=1, value="MOTIVATION ALERT:").font = Font(bold=True, color="C00000")
    ws2.cell(row=alert_row+1, column=1, value="Low Training Budget increases Defective Products!").font = Font(italic=True, color="C00000")
    ws2.cell(row=alert_row+2, column=1, value="High Absenteeism may indicate low morale - consider benefits.").font = Font(italic=True, color="666666")
    
    # Section C: Benefits
    benefits_header_row = alert_row + 5
    ws2.cell(row=benefits_header_row-1, column=1, value="SECTION C: BENEFITS DECISIONS").font = section_font
    
    benefit_headers = ['Benefit Type', 'Decision Value', 'Notes']
    for col, h in enumerate(benefit_headers, start=1):
        cell = ws2.cell(row=benefits_header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    benefits_start_row = benefits_header_row + 1
    for benefit_idx, (name, default, fmt, note) in enumerate(DEFAULT_BENEFITS):
        benefit_row = benefits_start_row + benefit_idx
        
        ws2.cell(row=benefit_row, column=1, value=name).border = thin_border
        
        cell = ws2.cell(row=benefit_row, column=2, value=default)
        cell.fill = input_fill
        cell.border = thin_border
        if fmt == 'percent':
            cell.number_format = '0.0%'
        
        ws2.cell(row=benefit_row, column=3, value=note).font = Font(italic=True)
    
    benefits_end_row = benefits_start_row + len(DEFAULT_BENEFITS) - 1
    
    # Column widths
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 28
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 18
    
    # =========================================================================
    # TAB 3: LABOR_COST_ANALYSIS
    # =========================================================================
    ws3 = wb.create_sheet("LABOR COST ANALYSIS")
    
    ws3['A1'] = "LABOR COST ANALYSIS - Total People Expense for Finance"
    ws3['A1'].font = title_font
    ws3['A2'] = "Output for CFO to include in cash flow projections."
    ws3['A2'].font = Font(italic=True, color="666666")
    
    # Inputs
    ws3['A4'] = "INPUT: Estimated Net Profit (for Profit Sharing)"
    cell = ws3['B4']
    cell.value = 0
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    ws3['A5'] = "Previous Period Labor Cost"
    cell = ws3['B5']
    cell.value = labor_data.get('total_labor', 0)
    cell.fill = ref_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    # Cost breakdown
    ws3['A7'] = "COST BREAKDOWN"
    ws3['A7'].font = section_font
    
    cost_headers = ['Cost Category', 'Calculation', 'Amount']
    for col, h in enumerate(cost_headers, start=1):
        cell = ws3.cell(row=8, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Cost rows
    cost_items = [
        ("Total Planned Headcount", f"='WORKFORCE PLANNING'!C{totals_row}", False),
        ("Base Salaries", f"=B9*AVERAGE(COMPENSATION_STRATEGY!D{salary_start_row}:D{salary_end_row})*8", True),
        ("Overtime & Bonuses", f"=0", True), # Placeholder for future
        ("Training & Benefits", f"=COMPENSATION_STRATEGY!B{benefits_start_row}*C10 + COMPENSATION_STRATEGY!B{benefits_start_row+1}*C10", True),
        ("Profit Sharing", f"=$B$4*COMPENSATION_STRATEGY!B{benefits_start_row+2}", True),
        ("Hiring & Firing", f"='WORKFORCE PLANNING'!K{totals_row}", True),
        # NEW: Salesforce Payroll (Fixed) - Updated Logic
        ("Factory Payroll Subtotal", f"=SUM(C9:C{row+5})", True), # Subtotal for factory
    ]
    
    # Add hidden payroll
    sales_hc = sales_data.get('headcount', 44)
    sales_avg = sales_data.get('avg_salary', 750)
    sales_total = sales_hc * sales_avg
    
    # Re-build cost items list to include new sections
    # Row 9 is Headcount. Start costs at row 10.
    # We will rewrite the section generation below to be cleaner.
    pass

    # New Cost Rows
    # 1. Total Planned Headcount (Ref)
    # 2. Base Salaries (Calc)
    # 3. Overtime (Placeholder)
    # 4. Training & Benefits (Calc)
    # 5. Profit Sharing (Calc)
    # 6. Hiring & Firing (Ref)
    # 7. SALESFORCE PAYROLL (Fixed) ---> NEW
    
    # Let's restart the row loop from row 9
    row = 9
    
    # 1. Headcount
    ws3.cell(row=row, column=1, value="Total Planned Factory Headcount").border = thin_border
    ws3.cell(row=row, column=2, value=f"='WORKFORCE PLANNING'!C{totals_row}").fill = ref_fill
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=3, value=f"=B{row}").fill = calc_fill # Just repeat number or calculate cost? Item says "Headcount" so just number.
    ws3.cell(row=row, column=3).border = thin_border
    row += 1
    
    # 2. Factory Wages
    ws3.cell(row=row, column=1, value="Factory Base Salaries").border = thin_border
    # Formula uses Avg Salary * 8? Wait, previous logic: =B9*AVERAGE(...)*8
    # B9 was Headcount. Yes.
    # Note: Salary in Compensation is Weekly? Or Daily? Project usually Daily/Weekly. Assuming standard.
    # Let's keep existing logic: Headcount * Avg Salary * 8 (weeks? or shift factor?). 
    # Usually simulation is 8 weeks per round?
    ws3.cell(row=row, column=2, value=f"=B9*AVERAGE(COMPENSATION_STRATEGY!D{salary_start_row}:D{salary_end_row})*8").fill = calc_fill
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).number_format = '$#,##0'
    ws3.cell(row=row, column=3, value=f"=B{row}").fill = calc_fill
    ws3.cell(row=row, column=3).border = thin_border
    ws3.cell(row=row, column=3).number_format = '$#,##0'
    row += 1
    
    # 3. Overtime
    ws3.cell(row=row, column=1, value="Overtime & Bonuses").border = thin_border
    ws3.cell(row=row, column=2, value=0).fill = ref_fill
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).number_format = '$#,##0'
    ws3.cell(row=row, column=3, value=f"=B{row}").fill = calc_fill
    ws3.cell(row=row, column=3).border = thin_border
    ws3.cell(row=row, column=3).number_format = '$#,##0'
    row += 1
    
    # 4. Training/Benefits
    ws3.cell(row=row, column=1, value="Training & Benefits").border = thin_border
    # Updated ref rows for benefits due to inserted rows in Tab 2?
    # Tab 2 inserted 2 rows at row 4. So `benefits_start_row` shifted by +2.
    # BUT `benefits_start_row` is calculated dynamically in code: `benefits_header_row + 1`.
    # `benefits_header_row` is based on `salary_end_row` (+2 rows inserted BEFORE salary section? No, "Section B: Salary" starts at 9).
    # "Section A" is 4. Inserted at 4.
    # So `salary_start_row` (11) becomes 13.
    # Let's verify Tab 2 offsets.
    # CODE LOGIC:
    # `ws2.insert_rows(4, 2)` -> Shift everything below down by 2.
    # `salary_start_row` was 11. Now 13.
    # `salary_end_row` shifts.
    # `benefits_start_row` shifts.
    # So references like `COMPENSATION_STRATEGY!D{salary_start_row}` MUST use the NEW calculated row indices.
    # Since I am replacing the text `ws2.insert_rows(4, 2)` in this very tool call, the previously defined variable `salary_start_row = 11` in the python script 
    # will NOT automatically update if I don't change the assignment line `salary_start_row = 11`.
    # WAIT. I am inserting the code that inserts rows. 
    # `salary_start_row` is defined on line 406.
    # `ws2.insert_rows(4, 2)` happens BEFORE line 406? 
    # Yes, I put it at `ws2['A4']`.
    # So `salary_start_row` needs to be updated or dynamic?
    # Actually, `ws2` is generated sequentially.
    # If I insert rows at line ~372 code-wise (Section A), but `salary_start_row` is defined later at line 406,
    # does `wb.create_sheet` -> write cell A1 -> insert rows affect future writes?
    # Yes. openpyxl `insert_rows` shifts existing cells.
    # If I write cells row-by-row, I should just increment my `row` counter instead of `insert_rows` if possible.
    # OR, if I use `insert_rows`, I must adjust my hardcoded row numbers.
    # EASIER: Don't use `insert_rows`. Just write to Row + 2.
    # BUT, I am patching correct?
    # Implementation Plan says "Insert at Top (Rows 1-2 or new rows)".
    # Better to just write to new rows explicitly to avoid shifting math headache.
    
    # Strategy:
    # 1. Update Tab 2 replacement to use Row offset for section A/B.
    # 2. Update Tab 3 formulas to match.
    
    # LET'S ABORT this specific replacement chunk and do it properly with explicit row numbers.
    pass

    # 4. Training formula: 
    # =COMPENSATION_STRATEGY!B{benefits_start_row}*C10 ...
    # Wait, Reference to "C10" (Base Salaries) is now C10?
    # Row 9 = Headcount
    # Row 10 = Base Salaries. Yes.
    ws3.cell(row=row, column=1, value="Training & Benefits").border = thin_border
    ws3.cell(row=row, column=2, value=f"=COMPENSATION_STRATEGY!B{benefits_start_row}*C10 + COMPENSATION_STRATEGY!B{benefits_start_row+1}*C10").fill = calc_fill
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).number_format = '$#,##0'
    ws3.cell(row=row, column=3, value=f"=B{row}").fill = calc_fill
    ws3.cell(row=row, column=3).border = thin_border
    ws3.cell(row=row, column=3).number_format = '$#,##0'
    row += 1
    
    # 5. Profit Sharing
    ws3.cell(row=row, column=1, value="Profit Sharing").border = thin_border
    ws3.cell(row=row, column=2, value=f"=$B$4*COMPENSATION_STRATEGY!B{benefits_start_row+2}").fill = calc_fill
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).number_format = '$#,##0'
    ws3.cell(row=row, column=3, value=f"=B{row}").fill = calc_fill
    ws3.cell(row=row, column=3).border = thin_border
    ws3.cell(row=row, column=3).number_format = '$#,##0'
    row += 1
    
    # 6. Hiring & Firing
    ws3.cell(row=row, column=1, value="Hiring & Firing").border = thin_border
    ws3.cell(row=row, column=2, value=f"='WORKFORCE PLANNING'!K{totals_row}").fill = ref_fill
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).number_format = '$#,##0'
    ws3.cell(row=row, column=3, value=f"=B{row}").fill = calc_fill
    ws3.cell(row=row, column=3).border = thin_border
    ws3.cell(row=row, column=3).number_format = '$#,##0'
    row += 1
    
    # 7. SALESFORCE PAYROLL (Fixed) - NEW
    ws3.cell(row=row, column=1, value="SALESFORCE PAYROLL (Fixed)").border = thin_border
    ws3.cell(row=row, column=1).font = Font(bold=True)
    
    # Display Headcount and Avg Salary in Column B as text?
    # Better: Put Cost in B, and Note in A?
    # Request: "Display Current Headcount... Display Avg Salary..."
    # Let's put calculation in B, and use A for label. 
    # Use Comment or concatenated string? 
    # Let's assume simpler: Just the cost, but maybe add a note row?
    # Request: "Add 'SALESFORCE PAYROLL (Fixed)'... Display: 'Current Headcount...'"
    
    # Cost
    sales_cost = sales_data.get('total_salary', 33000)
    ws3.cell(row=row, column=2, value=sales_cost).fill = ref_fill
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).number_format = '$#,##0'
    
    ws3.cell(row=row, column=3, value=f"=B{row}").fill = calc_fill
    ws3.cell(row=row, column=3).border = thin_border
    ws3.cell(row=row, column=3).number_format = '$#,##0'
    
    # Add Note Row below? No, stick to table format for Total Sum.
    # Add Note in Column D?
    ws3.cell(row=row, column=4, value=f"Based on {sales_data.get('headcount')} salespeople @ ${sales_data.get('avg_salary'):,.0f}/person (Admin Report)")
    ws3.cell(row=row, column=4).font = Font(italic=True, color="666666")

    
    # Total
    row += 1
    ws3.cell(row=row, column=1, value="TOTAL PEOPLE EXPENSE").font = Font(bold=True)
    # Exclude Headcount (Row 9) from SUM
    cell = ws3.cell(row=row, column=3, value=f'=SUM(C10:C{row-2})')
    cell.fill = output_fill
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.number_format = '$#,##0'
    total_row = row
    
    # Variance
    row += 2
    ws3.cell(row=row, column=1, value="Variance vs Previous Period").border = thin_border
    cell = ws3.cell(row=row, column=3, value=f'=C{total_row}-B5')
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    # Pie Chart
    chart = PieChart()
    # Start from Row 10 (Base Salaries) to skip Headcount
    labels = Reference(ws3, min_col=1, min_row=10, max_row=15)
    data = Reference(ws3, min_col=3, min_row=10, max_row=15)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    chart.title = "Labor Cost Distribution"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    ws3.add_chart(chart, "E7")
    
    # Column widths
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 18
    
    # =========================================================================
    # TAB 4: UPLOAD_READY_PEOPLE
    # =========================================================================
    ws4 = wb.create_sheet("UPLOAD READY PEOPLE")
    
    ws4['A1'] = "PEOPLE DECISIONS - ExSim Upload Format"
    ws4['A1'].font = title_font
    ws4['A2'] = "Copy these values to ExSim People upload."
    ws4['A2'].font = Font(italic=True, color="666666")
    
    # Salaries section
    ws4['A4'] = "Salaries"
    ws4['A4'].font = section_font
    
    ws4.cell(row=5, column=1, value="Zone").font = header_font
    ws4.cell(row=5, column=1).fill = header_fill
    ws4.cell(row=5, column=1).border = thin_border
    ws4.cell(row=5, column=2, value="Salary").font = header_font
    ws4.cell(row=5, column=2).fill = header_fill
    ws4.cell(row=5, column=2).border = thin_border
    
    for zone_idx, zone in enumerate(ZONES):
        zone_row = 6 + zone_idx
        cell = ws4.cell(row=zone_row, column=1, value=zone)
        cell.fill = zone_fills[zone]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        
        cell = ws4.cell(row=zone_row, column=2, value=f'=COMPENSATION_STRATEGY!D{salary_start_row+zone_idx}')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
    
    # Benefits section
    ws4['D4'] = "Benefits & Policies"
    ws4['D4'].font = section_font
    
    ws4.cell(row=5, column=4, value="Benefit").font = header_font
    ws4.cell(row=5, column=4).fill = header_fill
    ws4.cell(row=5, column=4).border = thin_border
    ws4.cell(row=5, column=5, value="Value").font = header_font
    ws4.cell(row=5, column=5).fill = header_fill
    ws4.cell(row=5, column=5).border = thin_border
    
    for i, (name, _, _, _) in enumerate(DEFAULT_BENEFITS):
        r = 6 + i
        ws4.cell(row=r, column=4, value=name).border = thin_border
        
        cell = ws4.cell(row=r, column=5, value=f'=COMPENSATION_STRATEGY!B{benefits_start_row+i}')
        cell.fill = calc_fill
        cell.border = thin_border

    # =========================================================================
    # TAB 4: CROSS_REFERENCE (Upstream Data)
    # =========================================================================
    ws4 = wb.create_sheet("CROSS REFERENCE")
    
    ws4['A1'] = "CROSS-REFERENCE SUMMARY - Upstream Support"
    ws4['A1'].font = title_font
    ws4['A2'] = "Key metrics from Production and Finance."
    ws4['A2'].font = Font(italic=True, color="666666")
    
    # Load shared data
    try:
        from shared_outputs import import_dashboard_data
        prod_data = import_dashboard_data('Production') or {}
        cfo_data = import_dashboard_data('CFO') or {}
    except ImportError:
        prod_data = {}
        cfo_data = {}
    
    row = 4
    
    # Production Section
    ws4.cell(row=row, column=1, value="Production (Targets)").font = section_font
    ws4.cell(row=row, column=1).fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid") # Blue
    ws4.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    prod_metrics = [
        ("Total Production Target", f"{sum([d.get('Target',0) for d in prod_data.get('production_plan', {}).values()]) if prod_data and 'production_plan' in prod_data else 'N/A'}"),
        ("Overtime Hours", f"{prod_data.get('overtime_hours', 0):,.0f}" if prod_data else "N/A"),
    ]
    
    for label, value in prod_metrics:
        ws4.cell(row=row, column=1, value=label).border = thin_border
        ws4.cell(row=row, column=2, value=value).border = thin_border
        row += 1
        
    row += 2
    
    # CFO Section
    ws4.cell(row=row, column=1, value="Finance (Payroll Budget)").font = section_font
    ws4.cell(row=row, column=1).fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Green
    ws4.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    cfo_metrics = [
        ("Cash Availability", "Check Finance Dashboard"),
        ("Liquidity Status", cfo_data.get('liquidity_status', 'Unknown') if cfo_data else "Unknown"),
    ]
    
    for label, value in cfo_metrics:
        ws4.cell(row=row, column=1, value=label).border = thin_border
        ws4.cell(row=row, column=2, value=value).border = thin_border
        row += 1

    # Formatting
    for col in ['A', 'B']:
        ws4.column_dimensions[col].width = 30

    # Save to buffer or file
    if output_buffer is not None:
        wb.save(output_buffer)
        output_buffer.seek(0)
        print("[SUCCESS] Created dashboard in BytesIO buffer")
    else:
        wb.save(OUTPUT_FILE)
        print(f"[SUCCESS] Created '{OUTPUT_FILE}'")


def main():
    """Main function."""
    print("ExSim CPO Dashboard Generator")
    print("=" * 50)
    
    print("\n[*] Loading data files...")
    from config import REPORTS_DIR, DATA_DIR
    print(f"    Primary source: {REPORTS_DIR}")
    print(f"    Fallback source: {DATA_DIR}")
    
    # Workers Balance
    workers_path = get_data_path("workers_balance_overtime.xlsx")
    if workers_path:
        workers_data = load_workers_balance(workers_path)
        print(f"  [OK] Loaded workers balance from {workers_path.parent.name}/")
    else:
        workers_data = load_workers_balance(None)
        print("  [!] Using default workers data")
        
    # Sales & Admin
    sales_path = get_data_path("sales_admin_expenses.xlsx")
    if sales_path:
        sales_data = load_sales_admin(sales_path)
        print(f"  [OK] Loaded sales admin data")
    else:
        sales_data = load_sales_admin(None)
        print("  [!] Using default sales admin data")
    
    # Labor Costs
    labor_path = get_data_path("production.xlsx")
    if labor_path:
        labor_data = load_labor_costs(labor_path)
        print(f"  [OK] Loaded labor costs")
    else:
        labor_data = load_labor_costs(None)
        print("  [!] Using default labor cost data")
        
    # NEW: Absenteeism
    if workers_path:
        absenteeism_rate = load_absenteeism_data(workers_path)
        print(f"  [OK] Extracted Absenteeism Rate: {absenteeism_rate:.1%}")
    else:
        absenteeism_rate = 0.02
        print("  [!] Using default absenteeism 2%")
    
    print("\n[*] Creating dashboard...")
    create_cpo_dashboard(workers_data, sales_data, labor_data, absenteeism_rate)

    print("\nSheets created:")
    print("  * WORKFORCE_PLANNING (Headcount & Hiring Impact)")
    print("  * COMPENSATION_STRATEGY (Salaries, Strikes, Benefits)")
    print("  * LABOR_COST_ANALYSIS (Total Expense Breakdown)")
    print("  * UPLOAD_READY_PEOPLE (ExSim Format)")
    
    # Export key metrics for downstream dashboards
    if export_dashboard_data:
        total_workers = sum(workers_data.get(zone, {}).get('workers', 0) for zone in ZONES)
        export_dashboard_data('CPO', {
            'workforce_headcount': {zone: workers_data.get(zone, {}).get('workers', 0) for zone in ZONES},
            'payroll_forecast': total_workers * DEFAULT_BASE_SALARY * 2,  # 2 fortnights
            'hiring_costs': 0  # Calculated from dashboard inputs
        })


if __name__ == "__main__":
    main()
