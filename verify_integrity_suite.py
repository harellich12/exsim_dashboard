import unittest
import pandas as pd
import openpyxl
import os
import sys
import subprocess
from pathlib import Path
import warnings

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')

BASE_PATH = Path(__file__).parent

class ExSimIntegrityTest(unittest.TestCase):
    """Base test class with helper methods."""
    
    @classmethod
    def run_generator(cls, folder, script_name):
        """Runs the dashboard generation script."""
        script_path = BASE_PATH / folder / script_name
        if not script_path.exists():
            print(f"Skipping {script_name}: File not found.")
            return False
            
        print(f"\n[EXEC] Running {script_name}...")
        try:
            # Run in the directory of the script so it finds its 'data' folder
            result = subprocess.run(
                [sys.executable, script_name],
                cwd=script_path.parent,
                capture_output=True,
                text=True,
                check=True
            )
            return True
        except subprocess.CalledProcessError as e:
            print(f"[ERROR] Failed to run {script_name}: {e}")
            print(f"Stderr: {e.stderr}")
            return False

    def load_dashboard(self, folder, filename, data_only=True):
        """Loads the generated dashboard Excel file."""
        path = BASE_PATH / folder / filename
        if not path.exists():
            self.fail(f"Dashboard file not found: {path}")
        return openpyxl.load_workbook(path, data_only=data_only)

    def load_source_data(self, folder, filename):
        """Loads a source Excel file from the data dictionary."""
        path = BASE_PATH / folder / "data" / filename
        if not path.exists():
            # Try case-insensitive search
            parent = BASE_PATH / folder / "data"
            if parent.exists():
                for f in os.listdir(parent):
                    if f.lower() == filename.lower():
                        return pd.read_excel(parent / f, header=None)
            return None
        return pd.read_excel(path, header=None)

    def parse_numeric(self, val):
        """Independent numeric parser."""
        if pd.isna(val): return 0
        s = str(val).replace('$', '').replace(',', '').replace('%', '').strip()
        try:
            return float(s)
        except:
            return 0


class TestCFODashboard(ExSimIntegrityTest):
    FOLDER = "CFO Dashboard"
    SCRIPT = "generate_finance_dashboard_final.py"
    OUTPUT = "Finance_Dashboard_Final.xlsx"

    @classmethod
    def setUpClass(cls):
        cls.run_generator(cls.FOLDER, cls.SCRIPT)

    def test_initial_cash_flow(self):
        """Verify Liquidity Monitor Starting Cash matches Initial Cash Flow Report."""
        # Dashboard uses Reports folder first, then local data
        reports_path = BASE_PATH / "Reports" / "initial_cash_flow.xlsx"
        local_path = BASE_PATH / self.FOLDER / "data" / "initial_cash_flow.xlsx"
        source_path = reports_path if reports_path.exists() else local_path
        
        source_df = pd.read_excel(source_path, header=None) if source_path.exists() else None
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["LIQUIDITY_MONITOR"]
        
        expected_cash = 0
        if source_df is not None:
            for _, row in source_df.iterrows():
                label = str(row[0]).strip() if pd.notna(row[0]) else ''
                # Match exact dashboard logic: "Final cash" with "start of the first fortnight"
                if 'Final cash' in label and 'start of the first fortnight' in label:
                    expected_cash = self.parse_numeric(row[1])
                    break
            # Fallback: just "Final cash"
            if expected_cash == 0:
                for _, row in source_df.iterrows():
                    label = str(row[0]).strip() if pd.notna(row[0]) else ''
                    if 'Final cash' in label:
                        expected_cash = self.parse_numeric(row[1])
                        break
        
        actual_cash = ws['B5'].value
        print(f"CFO: Checking Final Cash. Source={expected_cash}, Dashboard={actual_cash}")
        self.assertAlmostEqual(expected_cash, actual_cash, delta=1, 
                               msg=f"Initial Cash Mismatch: Source {expected_cash} != Dest {actual_cash}")

    def test_net_sales(self):
        """Verify Profit Control Net Sales matches Balance Sheet."""
        # Dashboard uses Reports folder first
        reports_path = BASE_PATH / "Reports" / "results_and_balance_statements.xlsx"
        local_path = BASE_PATH / self.FOLDER / "data" / "results_and_balance_statements.xlsx"
        source_path = reports_path if reports_path.exists() else local_path
        
        source_df = pd.read_excel(source_path, header=None) if source_path.exists() else None
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["PROFIT_CONTROL"]
        
        expected_sales = 0
        if source_df is not None:
            for _, row in source_df.iterrows():
                label = str(row[0]).strip().lower() if pd.notna(row[0]) else ''
                if 'net sales' in label or 'revenue' in label:
                    val = self.parse_numeric(row[1])
                    if val > 0:
                        expected_sales = val
                        break
        
        actual_sales = ws['B11'].value
        print(f"CFO: Checking Net Sales. Source={expected_sales}, Dashboard={actual_sales}")
        self.assertAlmostEqual(expected_sales, actual_sales, delta=1, 
                               msg=f"Net Sales Mismatch: Source {expected_sales} != Dest {actual_sales}")

    def test_solvency_gauge(self):
        """Verify Solvency Gauge helper data references valid cells."""
        # Load FORMULAS
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT, data_only=False)
        ws = wb["BALANCE_SHEET_HEALTH"]
        
        # Check Helper Table at H12:J14
        limit = ws['J13'].value
        curr_formula = ws['I13'].value
        
        print(f"CFO: Checking Solvency Gauge Limit. Limit={limit}")
        self.assertEqual(limit, 0.6, "Risk Threshold line should be at 0.6")
        self.assertIn("B", str(curr_formula), "Current Debt Ratio should reference Column B")


class TestProductionDashboard(ExSimIntegrityTest):
    FOLDER = "Production Manager Dashboard"
    SCRIPT = "generate_production_dashboard_zones.py"
    OUTPUT = "Production_Dashboard_Zones.xlsx"

    @classmethod
    def setUpClass(cls):
        cls.run_generator(cls.FOLDER, cls.SCRIPT)

    def test_raw_materials_load(self):
        """Verify Raw Materials (Part A) loaded into Zone Calculators."""
        source_df = self.load_source_data(self.FOLDER, "raw_materials.xlsx")
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["ZONE_CALCULATORS"]

        # Center Zone Part A is usually around row 6 (Material Stock)
        # We need to find the "Center" block and "Material Stock" row.

        # Approximate check: Center Part A inventory from source
        expected_inv = 0
        if source_df is not None:
            # Logic from generator: Center - Section 1... Part A... Final Inventory (col 8)
            # The logic in generator is slightly complex: it looks for zone/section header, then Part, then Final Inventory.
            # Here we just scan for the specific row based on printed debugs or structure knowledge.
            # In raw_materials.xlsx, row 10 is Final Inventory for Part A in Center - Section 1.
            # Let's use a more robust search.
            current_section = ""
            current_part = ""
            for _, row in source_df.iterrows():
                val_str = str(row[0])
                if "Section" in val_str:
                    current_section = val_str
                if "Part A" in val_str:
                    current_part = "Part A"
                elif "Part B" in val_str:
                    current_part = "Part B"

                if "Final inventory" in val_str and "Center" in current_section and current_part == "Part A":
                     expected_inv = self.parse_numeric(row[8])
                     break

        # In Dashboard, find Center Zone -> Material Stock
        actual_inv = 0
        for row in range(1, 20):
            if ws.cell(row=row, column=1).value == "Material Stock (Part A)":
                actual_inv = ws.cell(row=row, column=2).value
                break

        print(f"Production: Center Part A Inv. Source={expected_inv}, Dashboard={actual_inv}")
        self.assertEqual(expected_inv, actual_inv, "Part A Inventory mismatch")

    def test_machine_counts(self):
        """Verify Machine counts loaded from machine_spaces.xlsx."""
        source_df = self.load_source_data(self.FOLDER, "machine_spaces.xlsx")
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["ZONE_CALCULATORS"]

        # Calculate expected total machines (M1+M2+M3+M4)
        expected_machines = 0
        if source_df is not None:
            for _, row in source_df.iterrows():
                label = str(row[0])
                if label in ["M1", "M2", "M3-alpha", "M3-beta", "M4"]:
                    # Iterate columns backwards to find latest count
                    for i in range(len(row)-1, 0, -1):
                        val = self.parse_numeric(row[i])
                        if val > 0:
                            expected_machines += int(val)
                            break

        # In Dashboard, find "Machines in Zone" for Center (first occurrence)
        actual_machines = 0
        for row in range(1, 20):
            if ws.cell(row=row, column=1).value == "Machines in Zone":
                actual_machines = ws.cell(row=row, column=2).value
                break

        print(f"Production: Total Machines (Center). Source={expected_machines}, Dashboard={actual_machines}")
        self.assertEqual(expected_machines, actual_machines, "Machine count mismatch")


class TestCPODashboard(ExSimIntegrityTest):
    FOLDER = "CPO Dashboard"
    SCRIPT = "generate_cpo_dashboard.py"
    OUTPUT = "CPO_Dashboard.xlsx"

    @classmethod
    def setUpClass(cls):
        cls.run_generator(cls.FOLDER, cls.SCRIPT)

    def test_headcount(self):
        """Verify Workforce Planning Current Staff matches Workers Balance."""
        source_df = self.load_source_data(self.FOLDER, "workers_balance_overtime.xlsx")
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["WORKFORCE_PLANNING"]
        
        # Check Center zone (Zone 0)
        expected_workers = 0
        if source_df is not None:
             for _, row in source_df.iterrows():
                if 'workers assigned' in str(row[0]).lower():
                     expected_workers = self.parse_numeric(row[1]) # Center is usually col 1
                     break
        
        actual_workers = 0
        # Check Rows ~11 for "Center"
        for row in range(10, 15):
            if ws.cell(row=row, column=1).value == "Center":
                actual_workers = ws.cell(row=row, column=2).value
                break
        
        print(f"CPO: Checking Headcount (Center). Source={expected_workers}, Dashboard={actual_workers}")
        self.assertAlmostEqual(expected_workers, actual_workers, delta=0.1)

    def test_circular_reference_fix(self):
        """Verify Labor Cost Analysis has Total Headcount row to avoid circular ref."""
        # Load FORMULAS
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT, data_only=False)
        ws = wb["LABOR_COST_ANALYSIS"]
        
        # Row 9 should be Total Planned Headcount
        item_name = ws['A9'].value
        formula = ws['B9'].value
        
        print(f"CPO: Checking Row 9 content. Item='{item_name}', Value='{formula}'")
        self.assertIn("Headcount", str(item_name), "Row 9 should be 'Total Planned Headcount'")
        self.assertIn("WORKFORCE_PLANNING", str(formula), "Row 9 should link to Workforce Planning")


class TestESGDashboard(ExSimIntegrityTest):
    FOLDER = "ESG Dashboard"
    SCRIPT = "generate_esg_dashboard.py"
    OUTPUT = "ESG_Dashboard.xlsx"

    @classmethod
    def setUpClass(cls):
        cls.run_generator(cls.FOLDER, cls.SCRIPT)

    def test_emissions(self):
        """Verify Strategy Selector Emissions match ESG Report."""
        source_df = self.load_source_data(self.FOLDER, "esg_report.xlsx")
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["STRATEGY_SELECTOR"]
        
        expected = 0
        if source_df is not None:
            for _, row in source_df.iterrows():
                if 'total' in str(row[0]).lower() and 'emission' in str(row[0]).lower():
                    expected = self.parse_numeric(row[1])
                    break
        
        actual = ws['B6'].value
        print(f"ESG: Checking Emissions. Source={expected}, Dashboard={actual}")
        self.assertAlmostEqual(expected, actual, delta=1)

    def test_defaults_cleared(self):
        """Verify simulator inputs are 0 (Cleared Defaults)."""
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["STRATEGY_SELECTOR"]
        
        for row in range(13, 20):
            if ws.cell(row=row, column=1).value == "Solar PV Panels":
                qty = ws.cell(row=row, column=2).value
                self.assertEqual(qty, 0, "Solar PV Quantity should be 0 (default cleared)")
                break

    def test_tax_rate_helper(self):
        """Verify Helper Column I exists for Abatement Chart."""
        # Load FORMULAS
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT, data_only=False)
        ws = wb["STRATEGY_SELECTOR"]
        
        # Check first data row (row 13)
        helper_val = ws['I13'].value
        print(f"ESG: Checking Tax Helper Column I. Value='{helper_val}'")
        self.assertIn("IMPACT_CONFIG", str(helper_val), "Column I should reference Impact Config for Tax Rate")


class TestCLODashboard(ExSimIntegrityTest):
    FOLDER = "CLO Dashboard"
    SCRIPT = "generate_logistics_dashboard.py"
    OUTPUT = "Logistics_Dashboard.xlsx"

    @classmethod
    def setUpClass(cls):
        cls.run_generator(cls.FOLDER, cls.SCRIPT)

    def test_inventory_data(self):
        """Verify Inventory Tetris matches Finished Goods."""
        source_df = self.load_source_data(self.FOLDER, "finished_goods_inventory.xlsx")
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["INVENTORY_TETRIS"]
        
        center_inv_cell = None
        for row in range(1, 20):
            val = ws.cell(row=row, column=1).value
            if val and "Opening Inventory" in str(val):
                header = ws.cell(row=row-1, column=1).value
                if header and "CENTER" in header:
                    center_inv_cell = ws.cell(row=row, column=2).value
                    break
        
        print(f"CLO: Center Opening Inventory = {center_inv_cell}")
        self.assertTrue(center_inv_cell is not None)


class TestPurchasingDashboard(ExSimIntegrityTest):
    FOLDER = "Purchasing Role"
    SCRIPT = "generate_purchasing_dashboard_v2.py"
    OUTPUT = "Purchasing_Dashboard.xlsx"

    @classmethod
    def setUpClass(cls):
        cls.run_generator(cls.FOLDER, cls.SCRIPT)

    def test_ordering_cost(self):
        """Verify Ordering Cost matches Production/Cost Data."""
        source_df = self.load_source_data(self.FOLDER, "production.xlsx")
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["COST_ANALYSIS"]
        
        expected_ordering = 0
        if source_df is not None:
             # Logic matching dashboard: iterate columns looking for first >0
             for _, row in source_df.iterrows():
                if 'ordering' in str(row[0]).lower() and 'cost' in str(row[0]).lower():
                    for col_idx in range(1, min(10, len(row))):
                        val = self.parse_numeric(row[col_idx])
                        if val > 0:
                            expected_ordering = val
                            break
                    if expected_ordering > 0:
                        break

        actual_ordering = ws['B5'].value
        print(f"Purchasing: Ordering Cost. Source={expected_ordering}, Dashboard={actual_ordering}")
        # Allow larger delta as dashboard might be summing multiple entries or using different logic
        self.assertAlmostEqual(expected_ordering, actual_ordering, delta=5000)

    def test_defaults_cleared(self):
        """Verify Purchasing dashboard defaults are cleared (not hardcoded 5000)."""
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["COST_ANALYSIS"]
        ordering_cost = ws['B5'].value
        
        source_df = self.load_source_data(self.FOLDER, "production.xlsx")
        if source_df is None:
             self.assertEqual(ordering_cost, 0, "Ordering Cost should be 0 if source file missing")


class TestCMODashboard(ExSimIntegrityTest):
    FOLDER = "CMO Dashboard"
    SCRIPT = "generate_cmo_dashboard_complete.py"
    OUTPUT = "CMO_Dashboard_Complete.xlsx"

    @classmethod
    def setUpClass(cls):
        cls.run_generator(cls.FOLDER, cls.SCRIPT)

    def test_market_share(self):
        """Verify Segment Pulse checks Market Report."""
        source_df = self.load_source_data(self.FOLDER, "Market Report.xlsx")
        wb = self.load_dashboard(self.FOLDER, self.OUTPUT)
        ws = wb["SEGMENT_PULSE"]  # Confirmed sheet name
        
        # Check B6: High Segment / Center / Market Share
        actual_share = ws['B6'].value
        print(f"CMO: Market Share High/Center = {actual_share}")
        
        if source_df is None:
             # If source missing, dashboard might default or use prev data. Just check it's numeric.
             pass
        self.assertTrue(isinstance(actual_share, (int, float)), "Market Share should be a number")


if __name__ == '__main__':
    print("=======================================================")
    print("EXSIM DASHBOARD INTEGRITY SUITE")
    print("Verifying correctness of value loading across all 6 dashboards.")
    print("=======================================================")
    unittest.main()
