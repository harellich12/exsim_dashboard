"""
ExSim Purchasing Dashboard - MRP & Cost-Optimized Sourcing

Builds an MRP calculator that maps Production Needs to Supplier Orders,
handles Lead Time shifts, and analyzes Batch Size efficiency.

Required libraries: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.chart import LineChart, BarChart, Reference, Series
import warnings

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Required input files from centralized Reports folder
REQUIRED_FILES = [
    'raw_materials.xlsx',
    'production.xlsx',
    'Procurement Decisions.xlsx'
]

# Data source: Primary = Reports folder at project root, Fallback = local /data
REPORTS_FOLDER = Path(__file__).parent.parent / "Reports"
LOCAL_DATA_FOLDER = Path(__file__).parent / "data"

def get_data_path(filename):
    """Get data file path, checking Reports folder first, then local fallback."""
    primary = REPORTS_FOLDER / filename
    fallback = LOCAL_DATA_FOLDER / filename
    if primary.exists():
        return primary
    elif fallback.exists():
        return fallback
    return None

OUTPUT_FILE = "Purchasing_Dashboard.xlsx"

FORTNIGHTS = list(range(1, 9))  # 1-8
ZONES = ["Center", "West", "North", "East", "South"]
PARTS = ["Part A", "Part B"]
PIECES = ["Piece 1", "Piece 2", "Piece 3", "Piece 4", "Piece 5", "Piece 6"]
SUPPLIERS = ["Supplier A", "Supplier B", "Supplier C"]

# Default supplier configuration
DEFAULT_SUPPLIERS = {
    "Part A": [
        {"name": "Supplier A", "lead_time": 0, "cost": 0, "payment_terms": 0, "batch_size": 0},
        {"name": "Supplier B", "lead_time": 0, "cost": 0, "payment_terms": 0, "batch_size": 0},
        {"name": "Supplier C", "lead_time": 0, "cost": 0, "payment_terms": 0, "batch_size": 0},
    ],
    "Part B": [
        {"name": "Supplier A", "lead_time": 0, "cost": 0, "payment_terms": 0, "batch_size": 0},
        {"name": "Supplier B", "lead_time": 0, "cost": 0, "payment_terms": 0, "batch_size": 0},
        {"name": "Supplier C", "lead_time": 0, "cost": 0, "payment_terms": 0, "batch_size": 0},
    ]
}

DEFAULT_PIECES_CONFIG = {
    "Piece 1": {"cost": 0, "batch_size": 0},
    "Piece 2": {"cost": 0, "batch_size": 0},
    "Piece 3": {"cost": 0, "batch_size": 0},
    "Piece 4": {"cost": 0, "batch_size": 0},
    "Piece 5": {"cost": 0, "batch_size": 0},
    "Piece 6": {"cost": 0, "batch_size": 0},
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_numeric(value):
    """Parse formatted number strings."""
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
    try:
        return float(cleaned)
    except:
        return 0


def load_excel_file(filepath, sheet_name=None):
    """Load Excel file."""
    try:
        if sheet_name:
            return pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


# =============================================================================
# DATA LOADING
# =============================================================================

def load_raw_materials(filepath):
    """Load raw materials inventory data."""
    df = load_excel_file(filepath)
    
    data = {
        'parts': {part: {'final_inventory': 0} for part in PARTS},
        'pieces': {piece: {'final_inventory': 0} for piece in PIECES}
    }
    
    if df is None:
        return data
    
    current_item = None
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        
        # Detect Part/Piece names
        for part in PARTS:
            if part.lower() in first_val.lower():
                current_item = part
                break
        for piece in PIECES:
            if piece.lower() in first_val.lower():
                current_item = piece
                break
        
        # Get Final Inventory
        if 'final' in first_val.lower() and 'inventory' in first_val.lower():
            # Get fortnight 8 value (column 8)
            final_val = parse_numeric(row.iloc[8]) if len(row) > 8 else 0
            if current_item:
                if current_item in PARTS:
                    data['parts'][current_item]['final_inventory'] = final_val
                elif current_item in PIECES:
                    data['pieces'][current_item]['final_inventory'] = final_val
    
    return data


def load_production_costs(filepath):
    """Load production cost data for batch size analysis."""
    df = load_excel_file(filepath)
    
    data = {
        'ordering_cost': 0,
        'holding_cost': 0,
        'consumption_cost': 0
    }
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'ordering' in first_val and 'cost' in first_val:
            for col_idx in range(1, min(10, len(row))):
                val = parse_numeric(row.iloc[col_idx])
                if val > 0:
                    data['ordering_cost'] = val
                    break
        
        if 'holding' in first_val and 'cost' in first_val:
            for col_idx in range(1, min(10, len(row))):
                val = parse_numeric(row.iloc[col_idx])
                if val > 0:
                    data['holding_cost'] = val
                    break
        
        if 'consumed' in first_val or 'consumption' in first_val:
            for col_idx in range(1, min(10, len(row))):
                val = parse_numeric(row.iloc[col_idx])
                if val > 0:
                    data['consumption_cost'] = val
                    break
    
    return data


def load_procurement_template(filepath):
    """Load procurement template structure."""
    df = load_excel_file(filepath, sheet_name='Procurement')
    return {'df': df, 'exists': df is not None}


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def create_purchasing_dashboard(materials_data, cost_data, template_data):
    """Create the comprehensive Purchasing Dashboard."""
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5496")
    title_font = Font(bold=True, size=14, color="2F5496")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ref_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # =========================================================================
    # TAB 1: SUPPLIER_CONFIG
    # =========================================================================
    ws1 = wb.active
    ws1.title = "SUPPLIER_CONFIG"
    
    ws1['A1'] = "SUPPLIER CONFIGURATION"
    ws1['A1'].font = title_font
    ws1['A2'] = "Enter your case study supplier data here. Pre-filled with defaults."
    ws1['A2'].font = Font(italic=True, color="666666")
    
    # Table 1: PARTS CONFIG
    ws1['A4'] = "TABLE 1: PARTS SUPPLIERS"
    ws1['A4'].font = section_font
    
    parts_headers = ['Part', 'Supplier', 'Lead Time (FN)', 'Cost/Unit', 'Payment Terms (FN)', 'Batch Size']
    for col, h in enumerate(parts_headers, start=1):
        cell = ws1.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 6
    for part, suppliers in DEFAULT_SUPPLIERS.items():
        for supplier in suppliers:
            ws1.cell(row=row, column=1, value=part).border = thin_border
            
            cell = ws1.cell(row=row, column=2, value=supplier['name'])
            cell.border = thin_border
            cell.fill = input_fill
            
            cell = ws1.cell(row=row, column=3, value=supplier['lead_time'])
            cell.border = thin_border
            cell.fill = input_fill
            
            cell = ws1.cell(row=row, column=4, value=supplier['cost'])
            cell.border = thin_border
            cell.fill = input_fill
            cell.number_format = '$#,##0.00'
            
            cell = ws1.cell(row=row, column=5, value=supplier['payment_terms'])
            cell.border = thin_border
            cell.fill = input_fill
            
            cell = ws1.cell(row=row, column=6, value=supplier['batch_size'])
            cell.border = thin_border
            cell.fill = input_fill
            
            row += 1
    
    parts_config_end = row - 1
    row += 2
    
    # Table 2: PIECES CONFIG
    ws1.cell(row=row, column=1, value="TABLE 2: PIECES CONFIGURATION").font = section_font
    row += 1
    
    pieces_headers = ['Piece Name', 'Cost/Unit', 'Batch Size']
    for col, h in enumerate(pieces_headers, start=1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    pieces_start_row = row
    for piece, config in DEFAULT_PIECES_CONFIG.items():
        ws1.cell(row=row, column=1, value=piece).border = thin_border
        
        cell = ws1.cell(row=row, column=2, value=config['cost'])
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0.00'
        
        cell = ws1.cell(row=row, column=3, value=config['batch_size'])
        cell.border = thin_border
        cell.fill = input_fill
        
        row += 1
    
    # Column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 12
    
    # =========================================================================
    # TAB 2: COST_ANALYSIS
    # =========================================================================
    ws2 = wb.create_sheet("COST_ANALYSIS")
    
    ws2['A1'] = "COST ANALYSIS - Batch Size Efficiency"
    ws2['A1'].font = title_font
    
    ws2['A3'] = "PREVIOUS PERIOD COSTS"
    ws2['A3'].font = section_font
    
    # Cost data
    ordering_cost = cost_data.get('ordering_cost', 0)
    holding_cost = cost_data.get('holding_cost', 0)
    total_cost = ordering_cost + holding_cost
    
    ws2.cell(row=5, column=1, value="Ordering Cost (Total)").border = thin_border
    cell = ws2.cell(row=5, column=2, value=ordering_cost)
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    ws2.cell(row=6, column=1, value="Holding Cost (Total)").border = thin_border
    cell = ws2.cell(row=6, column=2, value=holding_cost)
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    ws2.cell(row=7, column=1, value="Total Cost").border = thin_border
    cell = ws2.cell(row=7, column=2, value="=B5+B6")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '$#,##0'
    
    # Ordering Cost Ratio
    ws2['A9'] = "EFFICIENCY ANALYSIS"
    ws2['A9'].font = section_font
    
    ws2.cell(row=11, column=1, value="Ordering Cost Ratio").border = thin_border
    cell = ws2.cell(row=11, column=2, value="=IF(B7>0, B5/B7, 0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    
    # Add Data Bar for Efficiency
    rule = DataBarRule(start_type='min', end_type='max', color="638EC6", showValue=None, minLength=None, maxLength=None)
    ws2.conditional_formatting.add("B11", rule)
    
    # Efficiency Flag
    ws2.cell(row=13, column=1, value="Efficiency Flag").border = thin_border
    cell = ws2.cell(row=13, column=2)
    cell.value = '=IF(B11>0.7,"CRITICAL: Ordering too often. INCREASE BATCH SIZE.",IF(B11<0.3,"CRITICAL: Holding too much. DECREASE BATCH SIZE/JIT.","OK: Balanced"))'
    cell.border = thin_border
    cell.font = Font(bold=True)
    
    # Strategic Advice Box
    ws2['A16'] = "STRATEGIC ADVICE"
    ws2['A16'].font = section_font
    
    ws2.merge_cells('A17:D20')
    cell = ws2['A17']
    cell.value = """Based on your Ordering Cost Ratio:
• > 70%: You're placing too many small orders. Consolidate orders into larger batches.
• < 30%: You're holding too much inventory. Consider Just-In-Time ordering or smaller batches.
• 30-70%: Good balance between ordering frequency and inventory holding."""
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.fill = calc_fill
    
    ws2.row_dimensions[17].height = 80
    
    # Column widths
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    
    # =========================================================================
    # TAB 3: MRP_ENGINE
    # =========================================================================
    ws3 = wb.create_sheet("MRP_ENGINE")
    
    ws3['A1'] = "MRP ENGINE - Material Requirements Planning"
    ws3['A1'].font = title_font
    
    # Section A: Production Demand
    ws3['A3'] = "SECTION A: PRODUCTION DEMAND (Input from Production Manager)"
    ws3['A3'].font = section_font
    
    # Headers
    ws3.cell(row=5, column=1, value="Metric").font = header_font
    ws3.cell(row=5, column=1).fill = header_fill
    for fn in FORTNIGHTS:
        cell = ws3.cell(row=5, column=1+fn, value=f"FN{fn}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Target Production input row
    ws3.cell(row=6, column=1, value="Target Production").border = thin_border
    for fn in FORTNIGHTS:
        cell = ws3.cell(row=6, column=1+fn, value=0)
        cell.border = thin_border
        cell.fill = input_fill
    
    row = 9
    
    # Section B: Net Requirements for each Part
    ws3.cell(row=row, column=1, value="SECTION B: NET REQUIREMENTS CALCULATION").font = section_font
    row += 2
    
    part_rows = {}
    order_rows = {}
    
    chart_start_row = row 
    
    for part in PARTS:
        ws3.cell(row=row, column=1, value=f"{part.upper()}").font = Font(bold=True, color="2F5496")
        row += 1
        
        opening_inv = materials_data['parts'].get(part, {}).get('final_inventory', 0)
        
        # Gross Requirement
        ws3.cell(row=row, column=1, value="Gross Requirement").border = thin_border
        for fn in FORTNIGHTS:
            cell = ws3.cell(row=row, column=1+fn, value=f"=$B$6")  # Links to target production
            cell.border = thin_border
            cell.fill = ref_fill
        gross_row = row
        row += 1
        
        # Scheduled Arrivals (user inputs based on orders placed)
        ws3.cell(row=row, column=1, value="Scheduled Arrivals").border = thin_border
        for fn in FORTNIGHTS:
            cell = ws3.cell(row=row, column=1+fn, value=0)
            cell.border = thin_border
            cell.fill = input_fill
        arrivals_row = row
        row += 1
        
        # Projected Inventory
        ws3.cell(row=row, column=1, value="Projected Inventory").border = thin_border
        for fn in FORTNIGHTS:
            col = 1 + fn
            if fn == 1:
                # First fortnight: Opening + Arrivals - Gross
                cell = ws3.cell(row=row, column=col, 
                    value=f"={opening_inv}+{get_column_letter(col)}{arrivals_row}-{get_column_letter(col)}{gross_row}")
            else:
                # Subsequent: Previous Inv + Arrivals - Gross
                prev_col = get_column_letter(col - 1)
                curr_col = get_column_letter(col)
                cell = ws3.cell(row=row, column=col,
                    value=f"={prev_col}{row}+{curr_col}{arrivals_row}-{curr_col}{gross_row}")
            cell.border = thin_border
            cell.fill = calc_fill
        proj_row = row
        part_rows[part] = {'arrivals': arrivals_row, 'projected': proj_row}
        row += 1
        
        # Net Deficit
        ws3.cell(row=row, column=1, value="Net Deficit (if negative)").border = thin_border
        for fn in FORTNIGHTS:
            col = 1 + fn
            cell = ws3.cell(row=row, column=col, 
                value=f'=IF({get_column_letter(col)}{proj_row}<0, -{get_column_letter(col)}{proj_row}, 0)')
            cell.border = thin_border
            cell.fill = output_fill
        row += 2
        
        # ORDERS
        order_rows[part] = {}
        ws3.cell(row=row, column=1, value=f"ORDERS FOR {part.upper()}").font = Font(bold=True)
        row += 1
        
        suppliers = DEFAULT_SUPPLIERS.get(part, [])
        for supplier in suppliers:
            name = supplier['name']
            lead = supplier['lead_time']
            batch = supplier['batch_size']
            
            ws3.cell(row=row, column=1, value=f"Order {name} (Lead:{lead}, Batch:{batch})").border = thin_border
            for fn in FORTNIGHTS:
                cell = ws3.cell(row=row, column=1+fn, value=0)
                cell.border = thin_border
                cell.fill = input_fill
            order_rows[part][name] = row
            row += 1
            
        row += 2
    
    # Add conditional formatting for negative inventory (Red Fill / White Text)
    for part, rows in part_rows.items():
        ws3.conditional_formatting.add(
            f'B{rows["projected"]}:I{rows["projected"]}',
            FormulaRule(formula=[f'B{rows["projected"]}<0'], fill=red_fill, font=Font(bold=True, color="FFFFFF"))
        )

    # -------------------------------------------------------------
    # CHART: Inventory Sawtooth & Stockout Risk
    # -------------------------------------------------------------
    
    # Create hidden row for Stockout Line (0s)
    stockout_row = row + 5
    ws3.cell(row=stockout_row, column=1, value="Stockout Limit")
    for f in range(1, 9):
        ws3.cell(row=stockout_row, column=1+f, value=0)
    
    # Hide the row? Openpyxl row dimensions hidden=True
    ws3.row_dimensions[stockout_row].hidden = True
    
    chart = LineChart()
    chart.title = "Inventory Sawtooth & Stockout Risk"
    chart.style = 12
    chart.y_axis.title = "Inventory Units"
    chart.x_axis.title = "Fortnight"
    chart.height = 12
    chart.width = 20
    
    # Series
    # Part A
    part_a_row = part_rows["Part A"]["projected"]
    data_a = Reference(ws3, min_col=2, min_row=part_a_row, max_col=9)
    s1 = Series(data_a, title="Part A Inv")
    chart.append(s1)

    # Part B
    part_b_row = part_rows["Part B"]["projected"]
    data_b = Reference(ws3, min_col=2, min_row=part_b_row, max_col=9)
    s2 = Series(data_b, title="Part B Inv")
    chart.append(s2)
    
    # Stockout Line (Red)
    data_zero = Reference(ws3, min_col=2, min_row=stockout_row, max_col=9)
    s3 = Series(data_zero, title="Stockout Line")
    s3.graphicalProperties.line.solidFill = "FF0000"
    s3.graphicalProperties.line.width = 20000 # Thick
    chart.append(s3)
    
    # Categories (FN1..FN8 from header)
    cats = Reference(ws3, min_col=2, min_row=5, max_col=9)
    chart.set_categories(cats)
    
    ws3.add_chart(chart, "K2")
    
    # Column widths
    ws3.column_dimensions['A'].width = 35
    for col in range(2, 10):
        ws3.column_dimensions[get_column_letter(col)].width = 10
    
    # =========================================================================
    # TAB 4: CASH_FLOW_PREVIEW
    # =========================================================================
    ws4 = wb.create_sheet("CASH_FLOW_PREVIEW")
    
    ws4['A1'] = "CASH FLOW PREVIEW - Procurement Spending"
    ws4['A1'].font = title_font
    
    ws4['A3'] = "ESTIMATED OUTFLOW BY SUPPLIER"
    ws4['A3'].font = section_font
    
    # Headers
    ws4.cell(row=5, column=1, value="Supplier").font = header_font
    ws4.cell(row=5, column=1).fill = header_fill
    for fn in FORTNIGHTS:
        cell = ws4.cell(row=5, column=1+fn, value=f"FN{fn}")
        cell.font = header_font
        cell.fill = header_fill
    cell = ws4.cell(row=5, column=10, value="Total")
    cell.font = header_font
    cell.fill = header_fill
    
    row = 6
    supplier_data_rows = {}
    
    for supplier_name in SUPPLIERS:
        ws4.cell(row=row, column=1, value=supplier_name).border = thin_border
        
        # Calculate sum of orders for this supplier across Part A and Part B
        # Link to MRP_ENGINE!{Col}{Row}
        # Ref: order_rows[part][supplier_name]
        
        for fn in FORTNIGHTS:
            col_letter = get_column_letter(1+fn)
            # Formula: =SUM(MRP_ENGINE!{col}{PartA_Row}, MRP_ENGINE!{col}{PartB_Row})
            refs = []
            for part in PARTS:
                r = order_rows.get(part, {}).get(supplier_name)
                if r:
                    refs.append(f"MRP_ENGINE!{col_letter}{r}")
            
            formula = "=" + "+".join(refs) if refs else "=0"
            cell = ws4.cell(row=row, column=1+fn, value=formula)
            cell.border = thin_border
            cell.fill = calc_fill
            cell.number_format = '$#,##0'
        
        # Total
        cell = ws4.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
        
        supplier_data_rows[supplier_name] = row
        row += 1
        
    # Total row
    ws4.cell(row=row, column=1, value="TOTAL SPEND").font = Font(bold=True)
    for fn in FORTNIGHTS:
        cell = ws4.cell(row=row, column=1+fn, value=f"=SUM({get_column_letter(1+fn)}6:{get_column_letter(1+fn)}{row-1})")
        cell.border = thin_border
        cell.fill = output_fill
        cell.font = Font(bold=True)
        cell.number_format = '$#,##0'
    cell = ws4.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
    cell.fill = output_fill
    cell.font = Font(bold=True)
    cell.number_format = '$#,##0'
    total_row = row
    
    # -------------------------------------------------------------
    # CHART: Cash Drain Heatmap (Stacked Bar)
    # -------------------------------------------------------------
    chart_cash = BarChart()
    chart_cash.title = "Cash Drain Heatmap by Supplier"
    chart_cash.type = "col"
    chart_cash.style = 10
    chart_cash.grouping = "stacked"
    chart_cash.overlap = 100
    chart_cash.y_axis.title = "Cash Outflow ($)"
    chart_cash.x_axis.title = "Fortnight"
    chart_cash.height = 10
    chart_cash.width = 15
    
    # Series
    # Rows 6, 7, 8 are Supplier A, B, C
    categories = Reference(ws4, min_col=2, min_row=5, max_col=9)
    
    for i, supp in enumerate(SUPPLIERS):
        r = supplier_data_rows[supp]
        data_s = Reference(ws4, min_col=2, min_row=r, max_col=9)
        ser = Series(data_s, title=supp)
        chart_cash.append(ser)

    chart_cash.set_categories(categories)
    ws4.add_chart(chart_cash, "A12") # Place below table

    
    # Budget tracking
    row += 20
    ws4.cell(row=row, column=1, value="BUDGET TRACKING").font = section_font
    row += 1
    
    ws4.cell(row=row, column=1, value="Total Budget").border = thin_border
    cell = ws4.cell(row=row, column=2, value=100000)
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    row += 1
    
    ws4.cell(row=row, column=1, value="Total Projected Spend").border = thin_border
    cell = ws4.cell(row=row, column=2, value=f"=J{total_row}")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '$#,##0'
    row += 1
    
    ws4.cell(row=row, column=1, value="Remaining Budget").border = thin_border
    cell = ws4.cell(row=row, column=2, value=f"=B{row-2}-B{row-1}")
    cell.border = thin_border
    cell.fill = output_fill
    cell.font = Font(bold=True)
    cell.number_format = '$#,##0'
    
    # Column widths
    ws4.column_dimensions['A'].width = 20
    for col in range(2, 11):
        ws4.column_dimensions[get_column_letter(col)].width = 12
    
    # =========================================================================
    # TAB 5: UPLOAD_READY_PROCUREMENT
    # =========================================================================
    ws5 = wb.create_sheet("UPLOAD_READY_PROCUREMENT")
    
    ws5['A1'] = "PROCUREMENT DECISIONS - ExSim Upload Format (Side-by-Side)"
    ws5['A1'].font = title_font
    ws5['A2'] = "This matches the exact ExSim Procurement upload layout"
    ws5['A2'].font = Font(italic=True, color="666666")
    
    # PARTS section (Columns A-K)
    ws5['A4'] = "Parts"
    ws5['A4'].font = section_font
    
    # Parts headers
    parts_headers = ['Zone', 'Supplier', 'Component'] + [str(fn) for fn in FORTNIGHTS]
    for col, h in enumerate(parts_headers, start=1):
        cell = ws5.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Parts data rows (Zone -> Supplier -> Part, with FN1-8 values)
    row = 6
    for zone in ZONES:
        for supplier in SUPPLIERS:
            for part in PARTS:
                ws5.cell(row=row, column=1, value=zone).border = thin_border
                ws5.cell(row=row, column=2, value=supplier).border = thin_border
                ws5.cell(row=row, column=3, value=part).border = thin_border
                
                # Link to MRP_ENGINE orders (only for Center zone)
                if zone == "Center":
                    mrp_row = order_rows.get(part, {}).get(supplier, 0)
                    for fn in FORTNIGHTS:
                        if mrp_row:
                            cell = ws5.cell(row=row, column=3+fn, value=f"=MRP_ENGINE!{get_column_letter(1+fn)}{mrp_row}")
                        else:
                            cell = ws5.cell(row=row, column=3+fn, value=0)
                        cell.border = thin_border
                        cell.fill = input_fill
                else:
                    # Other zones - manual input
                    for fn in FORTNIGHTS:
                        cell = ws5.cell(row=row, column=3+fn, value=0)
                        cell.border = thin_border
                        cell.fill = input_fill
                row += 1
    
    # PIECES section (Columns R-U, starting at column 18)
    pieces_col_start = 18
    
    ws5.cell(row=4, column=pieces_col_start, value="Pieces").font = section_font
    
    pieces_headers = ['Zone', 'Supplier', 'Component', 'Order']
    for col, h in enumerate(pieces_headers):
        cell = ws5.cell(row=5, column=pieces_col_start+col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Pieces data rows
    row = 6
    for zone in ZONES:
        for piece in PIECES:
            ws5.cell(row=row, column=pieces_col_start, value=zone).border = thin_border
            ws5.cell(row=row, column=pieces_col_start+1, value="Pieces").border = thin_border
            ws5.cell(row=row, column=pieces_col_start+2, value=piece).border = thin_border
            cell = ws5.cell(row=row, column=pieces_col_start+3, value=0)
            cell.border = thin_border
            cell.fill = input_fill
            row += 1
    
    # Column widths
    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 12
    for col in range(4, 12):
        ws5.column_dimensions[get_column_letter(col)].width = 8
    
    # Pieces columns
    ws5.column_dimensions[get_column_letter(pieces_col_start)].width = 10
    ws5.column_dimensions[get_column_letter(pieces_col_start+1)].width = 10
    ws5.column_dimensions[get_column_letter(pieces_col_start+2)].width = 10
    ws5.column_dimensions[get_column_letter(pieces_col_start+3)].width = 8
    
    # Save
    wb.save(OUTPUT_FILE)
    print(f"[SUCCESS] Created '{OUTPUT_FILE}'")


def main():
    """Main function."""
    print("ExSim Purchasing Dashboard Generator v2")
    print("=" * 50)
    
    print("\n[*] Loading data files...")
    print(f"    Primary source: {REPORTS_FOLDER}")
    print(f"    Fallback source: {LOCAL_DATA_FOLDER}")
    
    # Raw Materials
    materials_path = get_data_path("raw_materials.xlsx")
    if materials_path:
        materials_data = load_raw_materials(materials_path)
        print(f"  [OK] Loaded raw materials from {materials_path.parent.name}/")
    else:
        materials_data = load_raw_materials(None)
        print("  [!] Using default inventory data")
    
    # Production Costs
    costs_path = get_data_path("production.xlsx")
    if costs_path:
        cost_data = load_production_costs(costs_path)
        print(f"  [OK] Loaded production costs")
    else:
        cost_data = load_production_costs(None)
        print("  [!] Using default cost data")
    
    # Procurement Template
    template_path = get_data_path("Procurement Decisions.xlsx")
    if template_path:
        template_data = load_procurement_template(template_path)
        print(f"  [OK] Loaded procurement template")
    else:
        template_data = load_procurement_template(None)
        print("  [!] Using default template layout")
    
    print("\n[*] Creating dashboard...")
    
    create_purchasing_dashboard(materials_data, cost_data, template_data)
    
    print("\nSheets created:")
    print("  * SUPPLIER_CONFIG (Supplier & Batch Settings)")
    print("  * COST_ANALYSIS (Ordering vs Holding Efficiency)")
    print("  * MRP_ENGINE (Material Requirements Calculator)")
    print("  * CASH_FLOW_PREVIEW (Procurement Spending)")
    print("  * UPLOAD_READY_PROCUREMENT (ExSim Format)")


if __name__ == "__main__":
    main()
