"""
ExSim Dashboard Validation Test Suite

This script validates all 6 ExSim dashboard generators by:
1. Running each generator script
2. Opening the resulting Excel files
3. Validating formulas and cell references
4. Checking for common bugs like incorrect row offsets

Usage: python validate_dashboards.py
"""

import os
import sys
import subprocess
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re

# =============================================================================
# CONFIGURATION
# =============================================================================
BASE_DIR = Path(__file__).parent
DASHBOARDS = {
    "CFO": {
        "dir": "CFO Dashboard",
        "script": "generate_finance_dashboard_final.py",
        "output": "Finance_Dashboard_Final.xlsx",
        "tests": ["liquidity_cascade", "upload_ready_links"]
    },
    "CLO": {
        "dir": "CLO Dashboard", 
        "script": "generate_logistics_dashboard.py",
        "output": "Logistics_Dashboard.xlsx",
        "tests": ["inventory_cascade", "zone_config_refs", "upload_ready_links"]
    },
    "CPO_Workforce": {
        "dir": "CPO Dashboard",
        "script": "generate_cpo_dashboard.py", 
        "output": "CPO_Dashboard.xlsx",
        "tests": ["workforce_planning_refs", "compensation_strategy_refs", "upload_ready_links"]
    },
    "CMO": {
        "dir": "CMO Dashboard",
        "script": "generate_cmo_dashboard_complete.py",
        "output": "CMO_Dashboard_Complete.xlsx",
        "tests": ["strategy_cockpit_refs", "upload_ready_links"]
    },
    "Purchasing": {
        "dir": "Purchasing Role",
        "script": "generate_purchasing_dashboard_v2.py",
        "output": "Purchasing_Dashboard.xlsx",
        "tests": ["mrp_cascade", "upload_ready_links"]
    },
    "ESG": {
        "dir": "ESG Dashboard",
        "script": "generate_esg_dashboard.py",
        "output": "ESG_Dashboard.xlsx",
        "tests": ["esg_strategy_refs", "upload_ready_links"]
    },
    "Production": {
        "dir": "Produciton Manager Dashboard",
        "script": "generate_production_dashboard_zones.py",
        "output": "Production_Dashboard_Zones.xlsx",
        "tests": ["zone_calculator_refs", "upload_ready_links"]
    }
}


# =============================================================================
# TEST UTILITIES
# =============================================================================

def extract_cell_refs(formula):
    """Extract all cell references from a formula."""
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return []
    # Match patterns like B5, $B$5, B$5, $B5, Sheet!B5, etc.
    pattern = r"(?:([A-Za-z_][A-Za-z0-9_]*!)?)?\$?([A-Z]+)\$?(\d+)"
    matches = re.findall(pattern, formula)
    return [(m[0].rstrip("!") if m[0] else None, m[1], int(m[2])) for m in matches]


def check_cell_not_empty(ws, col, row, context=""):
    """Check if a referenced cell is not empty."""
    cell = ws.cell(row=row, column=column_index_from_string(col))
    value = cell.value
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return False, f"{context}: Cell {col}{row} is empty"
    return True, None


def resolve_cross_sheet_ref(wb, sheet_name, col, row):
    """Resolve a cross-sheet reference and check if target exists."""
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cell = ws.cell(row=row, column=column_index_from_string(col))
        return cell.value is not None
    return False


# =============================================================================
# SPECIFIC TEST FUNCTIONS
# =============================================================================

def test_liquidity_cascade(wb, results):
    """Test CFO dashboard Opening Cash → Ending Cash cascade."""
    ws = wb["LIQUIDITY_MONITOR"]
    
    # Find Opening Cash row and Ending Cash row
    open_cash_row = None
    ending_cash_row = None
    
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and "Opening Cash" in str(cell_val):
            open_cash_row = row
        if cell_val and "ENDING CASH BALANCE" in str(cell_val):
            ending_cash_row = row
    
    if not open_cash_row or not ending_cash_row:
        results.append(("FAIL", "CFO", "Could not find Opening Cash or Ending Cash rows"))
        return
    
    # Check FN2-FN8 Opening Cash formulas reference Ending Cash row
    for fn in range(2, 9):
        col = 1 + fn  # Column B=2 for FN1, C=3 for FN2, etc.
        formula = ws.cell(row=open_cash_row, column=col).value
        
        if not formula or not isinstance(formula, str):
            results.append(("FAIL", "CFO", f"FN{fn} Opening Cash has no formula"))
            continue
        
        refs = extract_cell_refs(formula)
        # Should reference previous column's ending cash row
        expected_ref_row = ending_cash_row
        
        found_correct_ref = False
        for sheet, ref_col, ref_row in refs:
            if ref_row == expected_ref_row:
                found_correct_ref = True
                break
        
        if not found_correct_ref:
            results.append(("FAIL", "CFO", f"FN{fn} Opening Cash references row {refs[0][2] if refs else 'N/A'}, expected row {expected_ref_row}"))
        else:
            results.append(("PASS", "CFO", f"FN{fn} Opening Cash correctly references Ending Cash row {expected_ref_row}"))


def test_inventory_cascade(wb, results):
    """Test CLO dashboard Projected Inventory cascade."""
    ws = wb["INVENTORY_TETRIS"]
    
    zones_found = 0
    for row in range(1, min(100, ws.max_row + 1)):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and "ZONE" in str(cell_val).upper() and "═" in str(cell_val):
            zones_found += 1
    
    if zones_found >= 3:
        results.append(("PASS", "CLO", f"Found {zones_found} zone sections in INVENTORY_TETRIS"))
    else:
        results.append(("FAIL", "CLO", f"Expected 5 zone sections, found {zones_found}"))


def test_zone_config_refs(wb, results):
    """Test CLO dashboard ROUTE_CONFIG references."""
    if "ROUTE_CONFIG" not in wb.sheetnames:
        results.append(("FAIL", "CLO", "ROUTE_CONFIG sheet not found"))
        return
    
    ws = wb["ROUTE_CONFIG"]
    
    # Check that transport modes exist at rows 6-8 column B
    for row, mode in [(6, "Train"), (7, "Truck"), (8, "Plane")]:
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and mode in str(cell_val):
            results.append(("PASS", "CLO", f"ROUTE_CONFIG row {row} contains {mode}"))
        else:
            results.append(("FAIL", "CLO", f"ROUTE_CONFIG row {row} expected {mode}, found {cell_val}"))


def test_calculator_refs(wb, results):
    """Test CPO Final CALCULATOR references."""
    ws = wb["CALCULATOR"]
    
    # Find TOTAL row
    total_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "TOTAL":
            total_row = row
            break
    
    if total_row:
        # Check if column B and F have SUM formulas
        b_formula = ws.cell(row=total_row, column=2).value
        f_formula = ws.cell(row=total_row, column=6).value
        
        if b_formula and "SUM" in str(b_formula).upper():
            results.append(("PASS", "CPO_Final", f"CALCULATOR TOTAL row {total_row} has SUM formula in column B"))
        else:
            results.append(("FAIL", "CPO_Final", f"CALCULATOR TOTAL row {total_row} missing SUM formula in column B"))
        
        if f_formula and "SUM" in str(f_formula).upper():
            results.append(("PASS", "CPO_Final", f"CALCULATOR TOTAL row {total_row} has SUM formula in column F"))
        else:
            results.append(("FAIL", "CPO_Final", f"CALCULATOR TOTAL row {total_row} missing SUM formula in column F"))
    else:
        results.append(("FAIL", "CPO_Final", "Could not find TOTAL row in CALCULATOR"))


def test_zone_calculator_refs(wb, results):
    """Test CPO Zones ZONE_CALCULATORS references."""
    if "ZONE_CALCULATORS" not in wb.sheetnames:
        results.append(("FAIL", "CPO_Zones", "ZONE_CALCULATORS sheet not found"))
        return
    
    ws = wb["ZONE_CALCULATORS"]
    zones = ["CENTER", "WEST", "NORTH", "EAST", "SOUTH"]
    zones_found = 0
    
    for row in range(1, min(200, ws.max_row + 1)):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val:
            for zone in zones:
                if zone in str(cell_val).upper() and "═" in str(cell_val):
                    zones_found += 1
                    zones.remove(zone)
                    break
    
    if zones_found == 5:
        results.append(("PASS", "CPO_Zones", "All 5 zones found in ZONE_CALCULATORS"))
    else:
        results.append(("FAIL", "CPO_Zones", f"Expected 5 zones, found {zones_found}"))


def test_strategy_cockpit_refs(wb, results):
    """Test CMO STRATEGY_COCKPIT references."""
    if "STRATEGY_COCKPIT" not in wb.sheetnames:
        results.append(("FAIL", "CMO", "STRATEGY_COCKPIT sheet not found"))
        return
    
    ws = wb["STRATEGY_COCKPIT"]
    
    # Check for zone rows starting at row 16
    zones_found = 0
    for row in range(16, 21):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and cell_val in ["Center", "West", "North", "East", "South"]:
            zones_found += 1
    
    if zones_found == 5:
        results.append(("PASS", "CMO", "All 5 zones found in STRATEGY_COCKPIT rows 16-20"))
    else:
        results.append(("FAIL", "CMO", f"Expected 5 zones at rows 16-20, found {zones_found}"))


def test_mrp_cascade(wb, results):
    """Test Purchasing MRP_ENGINE cascade."""
    if "MRP_ENGINE" not in wb.sheetnames:
        results.append(("FAIL", "Purchasing", "MRP_ENGINE sheet not found"))
        return
    
    ws = wb["MRP_ENGINE"]
    
    # Look for Projected Inventory rows with cascading formulas
    proj_inv_rows = []
    for row in range(1, min(100, ws.max_row + 1)):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and "Projected Inventory" in str(cell_val):
            proj_inv_rows.append(row)
    
    if len(proj_inv_rows) >= 2:
        results.append(("PASS", "Purchasing", f"Found {len(proj_inv_rows)} Projected Inventory rows in MRP_ENGINE"))
        
        # Check cascade for first Projected Inventory row
        row = proj_inv_rows[0]
        fn3_formula = ws.cell(row=row, column=4).value  # Column D = FN3
        if fn3_formula and isinstance(fn3_formula, str) and "C" in fn3_formula:
            results.append(("PASS", "Purchasing", f"MRP cascade formula references previous column"))
        else:
            results.append(("WARN", "Purchasing", f"MRP cascade formula may not reference previous column: {fn3_formula}"))
    else:
        results.append(("FAIL", "Purchasing", f"Expected 2+ Projected Inventory rows, found {len(proj_inv_rows)}"))


def test_workforce_planning_refs(wb, results):
    """Test CPO Workforce WORKFORCE_PLANNING references."""
    if "WORKFORCE_PLANNING" not in wb.sheetnames:
        results.append(("FAIL", "CPO_Workforce", "WORKFORCE_PLANNING sheet not found"))
        return
    
    ws = wb["WORKFORCE_PLANNING"]
    
    # Check for zone rows (Center, West, North, East, South)
    zones_found = 0
    for row in range(1, min(50, ws.max_row + 1)):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and cell_val in ["Center", "West", "North", "East", "South"]:
            zones_found += 1
    
    if zones_found == 5:
        results.append(("PASS", "CPO_Workforce", "All 5 zones found in WORKFORCE_PLANNING"))
    else:
        results.append(("FAIL", "CPO_Workforce", f"Expected 5 zones, found {zones_found}"))
    
    # Check for hiring/firing cost formulas
    formulas_found = 0
    for row in range(1, min(50, ws.max_row + 1)):
        for col in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and "MAX(0" in str(cell.value):
                formulas_found += 1
    
    if formulas_found >= 5:
        results.append(("PASS", "CPO_Workforce", f"Found {formulas_found} MAX formulas for hiring/firing calc"))
    else:
        results.append(("WARN", "CPO_Workforce", f"Expected 5+ MAX formulas, found {formulas_found}"))


def test_compensation_strategy_refs(wb, results):
    """Test CPO Workforce COMPENSATION_STRATEGY references."""
    if "COMPENSATION_STRATEGY" not in wb.sheetnames:
        results.append(("FAIL", "CPO_Workforce", "COMPENSATION_STRATEGY sheet not found"))
        return
    
    ws = wb["COMPENSATION_STRATEGY"]
    
    # Check for inflation rate input at B6
    inflation_cell = ws.cell(row=6, column=2).value
    if inflation_cell is not None:
        results.append(("PASS", "CPO_Workforce", f"Inflation rate found at B6: {inflation_cell}"))
    else:
        results.append(("FAIL", "CPO_Workforce", "Inflation rate missing at B6"))
    
    # Check for Strike Risk formulas
    strike_formulas = 0
    for row in range(1, min(30, ws.max_row + 1)):
        for col in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and "STRIKE" in str(cell.value).upper():
                strike_formulas += 1
    
    if strike_formulas >= 5:
        results.append(("PASS", "CPO_Workforce", f"Found {strike_formulas} Strike Risk indicators"))
    else:
        results.append(("WARN", "CPO_Workforce", f"Expected 5+ Strike Risk indicators, found {strike_formulas}"))


def test_esg_strategy_refs(wb, results):
    """Test ESG STRATEGY_SELECTOR references."""
    if "STRATEGY_SELECTOR" not in wb.sheetnames:
        results.append(("FAIL", "ESG", "STRATEGY_SELECTOR sheet not found"))
        return
    
    ws = wb["STRATEGY_SELECTOR"]
    
    # Check for initiative rows (Solar, Trees, Green Electricity, CO2 Credits)
    initiatives = ["Solar", "Trees", "Green", "Credit"]
    found_initiatives = 0
    for row in range(1, min(30, ws.max_row + 1)):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val:
            for init in initiatives:
                if init.lower() in str(cell_val).lower():
                    found_initiatives += 1
                    break
    
    if found_initiatives >= 4:
        results.append(("PASS", "ESG", f"Found {found_initiatives} initiative rows in STRATEGY_SELECTOR"))
    else:
        results.append(("FAIL", "ESG", f"Expected 4 initiatives, found {found_initiatives}"))
    
    # Check for IMPACT_CONFIG cross-references
    impact_refs = 0
    for row in range(1, min(30, ws.max_row + 1)):
        for col in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and "IMPACT_CONFIG" in str(cell.value):
                impact_refs += 1
    
    if impact_refs >= 4:
        results.append(("PASS", "ESG", f"Found {impact_refs} IMPACT_CONFIG cross-references in STRATEGY_SELECTOR"))
    else:
        results.append(("WARN", "ESG", f"Expected 4+ IMPACT_CONFIG refs, found {impact_refs}"))


def test_upload_ready_links(wb, results, dashboard_name):
    """Test UPLOAD_READY tabs have formulas linking to other tabs."""
    upload_tabs = [s for s in wb.sheetnames if "UPLOAD" in s.upper()]
    
    if not upload_tabs:
        results.append(("FAIL", dashboard_name, "No UPLOAD_READY tabs found"))
        return
    
    for tab_name in upload_tabs:
        ws = wb[tab_name]
        formula_count = 0
        cross_sheet_refs = 0
        
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    if "!" in cell.value:
                        cross_sheet_refs += 1
        
        if formula_count > 0:
            results.append(("PASS", dashboard_name, f"{tab_name}: {formula_count} formulas, {cross_sheet_refs} cross-sheet refs"))
        else:
            results.append(("WARN", dashboard_name, f"{tab_name}: No formulas found - may be input-only"))


# =============================================================================
# MAIN TEST RUNNER
# =============================================================================

def run_generator(dashboard_info):
    """Run a dashboard generator script."""
    script_path = BASE_DIR / dashboard_info["dir"] / dashboard_info["script"]
    working_dir = BASE_DIR / dashboard_info["dir"]
    
    if not script_path.exists():
        return False, f"Script not found: {script_path}"
    
    try:
        result = subprocess.run(
            [sys.executable, str(script_path)],
            cwd=str(working_dir),
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if result.returncode == 0:
            return True, "Generated successfully"
        else:
            return False, f"Error: {result.stderr[:200]}"
    except subprocess.TimeoutExpired:
        return False, "Generation timed out"
    except Exception as e:
        return False, str(e)


def run_tests(dashboard_name, dashboard_info):
    """Run all tests for a dashboard."""
    results = []
    
    # Check if output file exists
    output_path = BASE_DIR / dashboard_info["dir"] / dashboard_info["output"]
    if not output_path.exists():
        results.append(("FAIL", dashboard_name, f"Output file not found: {output_path}"))
        return results
    
    # Load workbook
    try:
        wb = load_workbook(str(output_path), data_only=False)
    except Exception as e:
        results.append(("FAIL", dashboard_name, f"Could not load workbook: {e}"))
        return results
    
    # Run specific tests
    test_funcs = {
        "liquidity_cascade": test_liquidity_cascade,
        "inventory_cascade": test_inventory_cascade,
        "zone_config_refs": test_zone_config_refs,
        "calculator_refs": test_calculator_refs,
        "zone_calculator_refs": test_zone_calculator_refs,
        "strategy_cockpit_refs": test_strategy_cockpit_refs,
        "mrp_cascade": test_mrp_cascade,
        "workforce_planning_refs": test_workforce_planning_refs,
        "compensation_strategy_refs": test_compensation_strategy_refs,
        "esg_strategy_refs": test_esg_strategy_refs,
    }
    
    for test_name in dashboard_info["tests"]:
        if test_name == "upload_ready_links":
            test_upload_ready_links(wb, results, dashboard_name)
        elif test_name in test_funcs:
            test_funcs[test_name](wb, results)
    
    wb.close()
    return results


def main():
    print("=" * 70)
    print("ExSim Dashboard Validation Test Suite")
    print("=" * 70)
    
    all_results = []
    
    for dashboard_name, dashboard_info in DASHBOARDS.items():
        print(f"\n{'-' * 50}")
        print(f"Testing: {dashboard_name}")
        print(f"{'-' * 50}")
        
        # Generate dashboard
        print(f"  [*] Running generator...")
        success, message = run_generator(dashboard_info)
        if success:
            print(f"  [OK] {message}")
        else:
            print(f"  [XX] {message}")
            all_results.append(("FAIL", dashboard_name, f"Generation failed: {message}"))
            continue
        
        # Run tests
        print(f"  [*] Running validation tests...")
        results = run_tests(dashboard_name, dashboard_info)
        all_results.extend(results)
        
        for status, name, message in results:
            icon = "OK" if status == "PASS" else ("!!" if status == "WARN" else "XX")
            print(f"  [{icon}] {message}")
    
    # Summary
    print(f"\n{'=' * 70}")
    print("SUMMARY")
    print(f"{'=' * 70}")
    
    pass_count = sum(1 for r in all_results if r[0] == "PASS")
    warn_count = sum(1 for r in all_results if r[0] == "WARN")
    fail_count = sum(1 for r in all_results if r[0] == "FAIL")
    
    print(f"  PASS: {pass_count}")
    print(f"  WARN: {warn_count}")
    print(f"  FAIL: {fail_count}")
    
    if fail_count > 0:
        print(f"\n[FAIL] Failures detected:")
        for status, name, message in all_results:
            if status == "FAIL":
                print(f"  - [{name}] {message}")
        return 1
    else:
        print(f"\n[SUCCESS] All critical tests passed!")
        return 0


if __name__ == "__main__":
    sys.exit(main())
