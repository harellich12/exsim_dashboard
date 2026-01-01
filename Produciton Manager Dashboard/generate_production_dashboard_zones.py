"""
ExSim Production Dashboard Zones - Zone-Specific Production Planning

Handles physical separation of resources (Machines/Inventory) across
different geographic zones (Center, West, North, East, South).

Required libraries: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference, Series
import warnings
import sys

# Add parent directory to path to import case_parameters
sys.path.append(str(Path(__file__).parent.parent))
try:
    from case_parameters import PRODUCTION
except ImportError:
    print("Warning: Could not import case_parameters.py. Using defaults.")
    PRODUCTION = {}

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Required input files from centralized Reports folder
REQUIRED_FILES = [
    'raw_materials.xlsx',
    'finished_goods_inventory.xlsx',
    'workers_balance_overtime.xlsx',
    'machine_spaces.xlsx',
    'Production Decisions.xlsx'
]

# Data source: Primary = Reports folder at project root, Fallback = local /data
# Can be overridden by EXSIM_REPORTS_PATH environment variable for testing
import os
REPORTS_FOLDER = Path(os.environ.get('EXSIM_REPORTS_PATH', Path(__file__).parent.parent / "Reports"))
LOCAL_DATA_FOLDER = Path(__file__).parent / "data"

def get_data_path(filename):
    """Get data file path, checking Reports folder first, then local fallback."""
    primary = REPORTS_FOLDER / filename
    fallback = LOCAL_DATA_FOLDER / filename
    if primary.exists():
        return primary
    elif fallback.exists():
        return fallback
    return None

OUTPUT_FILE = "Production_Dashboard_Zones.xlsx"

FORTNIGHTS = list(range(1, 9))  # 1-8
ZONES = ["Center", "West", "North", "East", "South"]
SECTIONS = ["Section 1", "Section 2", "Section 3"]
MACHINE_TYPES = ["M1", "M2", "M3-alpha", "M3-beta", "M4"]

# Default production parameters
# Default production parameters from Case
MACHINES = PRODUCTION.get('MACHINES', {})
DEFAULT_NOMINAL_RATE = MACHINES.get('M1', {}).get('CAPACITY', 200)
# Variable cost estimate (Material + Labor + OH). Case doesn't give single number, keeping estimate.
DEFAULT_VARIABLE_COST = 40
DEFAULT_OT_CAPACITY_PCT = PRODUCTION.get('WORKFORCE', {}).get('OVERTIME_CAPACITY_PCT', 0.20)
DEFAULT_OT_MULTIPLIER = 1.4
DEFAULT_OT_COST_PREMIUM = 20 # Placeholder approximation of 1.4x Labor


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_numeric(value):
    """Parse formatted number strings."""
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
    try:
        return float(cleaned)
    except:
        return 0


def load_excel_file(filepath, sheet_name=None):
    """Load Excel file."""
    try:
        if sheet_name:
            return pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


# =============================================================================
# DATA LOADING (Zone-Specific)
# =============================================================================

def load_raw_materials_by_zone(filepath):
    """Load raw materials inventory grouped by zone."""
    df = load_excel_file(filepath)

    # Initialize all zones with defaults
    data = {zone: {'part_a': 0, 'part_b': 0} for zone in ZONES}

    if df is None:
        # Default: Only Center has inventory
        data['Center'] = {'part_a': 4000, 'part_b': 1000}
        return data

    current_zone = None
    current_part = None

    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''

        # Detect zone from section headers
        for zone in ZONES:
            if zone.lower() in first_val.lower() and 'section' in first_val.lower():
                current_zone = zone
                current_part = None # Reset part when zone/section changes
                break

        # Detect part type
        if 'part a' in first_val.lower():
            current_part = 'part_a'
        elif 'part b' in first_val.lower():
            current_part = 'part_b'
        elif 'piece' in first_val.lower() or 'assembly' in first_val.lower():
             # It's something else, so reset current_part so we don't attribute its inventory to Part A/B
             current_part = None

        # Get final inventory
        if 'final' in first_val.lower() and 'inventory' in first_val.lower():
            val = parse_numeric(row.iloc[8]) if len(row) > 8 else 0
            if current_zone and current_part:
                data[current_zone][current_part] = val

    return data


def load_finished_goods_by_zone(filepath):
    """Load finished goods inventory grouped by zone."""
    df = load_excel_file(filepath)

    data = {zone: {'capacity': 0, 'inventory': 0} for zone in ZONES}

    if df is None:
        data['Center'] = {'capacity': 4800, 'inventory': 500}
        return data

    current_zone = None

    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''

        # Detect zone
        for zone in ZONES:
            if zone.lower() in first_val.lower() and ('section' in first_val.lower() or 'warehouse' in first_val.lower()):
                current_zone = zone
                break

        if 'capacity' in first_val.lower():
            import re
            match = re.search(r'(\d+)', first_val.replace(',', ''))
            if match and current_zone:
                data[current_zone]['capacity'] = int(match.group(1))

        if 'final' in first_val.lower() and 'inventory' in first_val.lower():
            val = parse_numeric(row.iloc[8]) if len(row) > 8 else 0
            if current_zone:
                data[current_zone]['inventory'] = val

    return data


def load_workers_by_zone(filepath):
    """Load workers balance grouped by zone."""
    df = load_excel_file(filepath)

    # Default: Workers only in Center and West
    data = {zone: {'workers': 0, 'absenteeism': 0} for zone in ZONES}

    if df is None:
        data['Center'] = {'workers': 219, 'absenteeism': 0}
        data['West'] = {'workers': 71, 'absenteeism': 0}
        return data

    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''

        if 'workers assigned' in first_val.lower():
            # Columns: Center, West, North, East, South, Total
            for col_idx, zone in enumerate(['Center', 'West', 'North', 'East', 'South'], start=1):
                if col_idx < len(row):
                    data[zone]['workers'] = parse_numeric(row.iloc[col_idx])

        if 'absenteeism' in first_val.lower():
            for col_idx, zone in enumerate(['Center', 'West', 'North', 'East', 'South'], start=1):
                if col_idx < len(row):
                    data[zone]['absenteeism'] = parse_numeric(row.iloc[col_idx])

    return data


def load_machines_by_zone(filepath):
    """Load machine counts and modules grouped by zone."""
    df = load_excel_file(filepath)

    # Default: All machines in Center
    data = {zone: {'machines': 0, 'modules': 0, 'modules_used': 0} for zone in ZONES}

    if df is None:
        data['Center'] = {'machines': 57, 'modules': 72, 'modules_used': 69}
        return data

    # For now, assume all machines are in Center (most common scenario)
    total_machines = 0
    modules_available = 72
    modules_occupied = 0

    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''

        # Get machine counts
        for mt in MACHINE_TYPES:
            if first_val == mt:
                for col_idx in reversed(range(1, min(11, len(row)))):
                    val = parse_numeric(row.iloc[col_idx])
                    if val > 0:
                        total_machines += int(val)
                        break
                break

        if 'available' in first_val.lower():
            modules_available = parse_numeric(row.iloc[1]) if len(row) > 1 else 72

        if 'occupied' in first_val.lower():
            modules_occupied = parse_numeric(row.iloc[1]) if len(row) > 1 else 0

    # Assign all to Center (default scenario)
    data['Center'] = {
        'machines': total_machines,
        'modules': int(modules_available),
        'modules_used': int(modules_occupied)
    }

    return data


def load_production_template(filepath):
    """Load production decisions template."""
    try:
        df = pd.read_excel(filepath, sheet_name='Production', header=None)
        return {'df': df, 'exists': True}
    except:
        return {'df': None, 'exists': False}


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def create_zones_dashboard(materials_data, fg_data, workers_data,
                           machines_data, template_data):
    """Create the Zone-Specific Production Dashboard."""

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5496")
    title_font = Font(bold=True, size=14, color="2F5496")
    zone_font = Font(bold=True, size=11, color="FFFFFF")
    zone_fills = {
        'Center': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'West': PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        'North': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'East': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        'South': PatternFill(start_color="9E480E", end_color="9E480E", fill_type="solid"),
    }
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ref_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Track zone output rows for linking
    zone_output_rows = {}

    # =========================================================================
    # TAB 1: ZONE_CALCULATORS
    # =========================================================================
    ws1 = wb.active
    ws1.title = "ZONE_CALCULATORS"

    ws1['A1'] = "ZONE-SPECIFIC PRODUCTION CALCULATORS"
    ws1['A1'].font = title_font
    ws1['A2'] = "Each zone has its own resources. Resources in Center do NOT count towards West capacity."
    ws1['A2'].font = Font(italic=True, color="666666")

    row = 4

    for zone in ZONES:
        zone_data = {
            'machines': machines_data.get(zone, {}).get('machines', 0),
            'materials': materials_data.get(zone, {}).get('part_a', 0),
            'workers': workers_data.get(zone, {}).get('workers', 0),
        }

        # Zone Header
        ws1.merge_cells(f'A{row}:J{row}')
        cell = ws1.cell(row=row, column=1, value=f"═══ {zone.upper()} ZONE ═══")
        cell.font = zone_font
        cell.fill = zone_fills[zone]
        cell.alignment = Alignment(horizontal='center')
        chart_anchor_row = row  # Anchor for charts
        row += 1

        # Zone Parameters
        params = [
            ("Machines in Zone", zone_data['machines']),
            ("Material Stock (Part A)", zone_data['materials']),
            ("Workers in Zone", zone_data['workers']),
            ("Nominal Rate/Machine", DEFAULT_NOMINAL_RATE),
        ]

        for i, (label, value) in enumerate(params):
            ws1.cell(row=row, column=1, value=label).border = thin_border
            cell = ws1.cell(row=row, column=2, value=value)
            cell.border = thin_border
            cell.fill = ref_fill
            row += 1

        params_end = row - 1
        row += 1

        # Production Schedule Headers
        calc_headers = ['Fortnight', 'Target', 'Overtime',
                        'Local Capacity', 'Max OT Pot.', 'Material Cap',
                        'REAL OUTPUT', 'Est. Unit Cost', 'Shipment?']

        for col, h in enumerate(calc_headers, start=1):
            cell = ws1.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = zone_fills[zone]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        row += 1

        data_start = row
        for fn in FORTNIGHTS:
            ws1.cell(row=row, column=1, value=f"FN{fn}").border = thin_border

            # Target (input)
            cell = ws1.cell(row=row, column=2, value=0 if zone_data['machines'] == 0 else 500)
            cell.border = thin_border
            cell.fill = input_fill

            # Overtime Y/N (input)
            cell = ws1.cell(row=row, column=3, value="N")
            cell.border = thin_border
            cell.fill = input_fill
            cell.alignment = Alignment(horizontal='center')

            # Local Capacity (Nominal) = Machines × Rate
            cell = ws1.cell(row=row, column=4, value=f"=$B${params_end-3}*$B${params_end}")
            cell.border = thin_border
            cell.fill = calc_fill

            # Max OT Capacity = Capacity * 0.20
            cell = ws1.cell(row=row, column=5, value=f"=D{row}*{DEFAULT_OT_CAPACITY_PCT}")
            cell.border = thin_border
            cell.font = Font(color="666666", italic=True)

            # Material Cap (simplified)
            cell = ws1.cell(row=row, column=6, value=f"=$B${params_end-2}")
            cell.border = thin_border
            cell.fill = calc_fill

            # REAL OUTPUT = MIN(Target, Nominal + (If OT, OT_Pot, 0), Material)
            # Logic: IF OT="Y", Cap = D + E, else D.
            ot_logic = f'IF(C{row}="Y", D{row}+E{row}, D{row})'
            cell = ws1.cell(row=row, column=7, value=f"=MIN(B{row}, {ot_logic}, F{row})")
            cell.border = thin_border
            cell.fill = output_fill
            cell.font = Font(bold=True)

            # Est Unit Cost
            # If OT="Y", cost = 60, else 40.
            cell = ws1.cell(row=row, column=8, value=f'=IF(C{row}="Y", {DEFAULT_VARIABLE_COST + DEFAULT_OT_COST_PREMIUM}, {DEFAULT_VARIABLE_COST})')
            cell.border = thin_border
            cell.number_format = '$#,##0'

            # Shipment Alert
            cell = ws1.cell(row=row, column=9,
                value=f'=IF(B{row}>F{row}, "SHIPMENT NEEDED!", "OK")')
            cell.border = thin_border

            row += 1

        # Total row
        ws1.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        cell = ws1.cell(row=row, column=2, value=f"=SUM(B{data_start}:B{row-1})")
        cell.fill = input_fill
        cell = ws1.cell(row=row, column=7, value=f"=SUM(G{data_start}:G{row-1})")
        cell.fill = output_fill
        cell.font = Font(bold=True)

        zone_output_rows[zone] = {'total': row, 'data_start': data_start, 'data_end': row-1}

        # ---------------------------------------------------------
        # CHARTS
        # ---------------------------------------------------------

        # 1. Capacity Constraint Stack (Combo)
        # Bar: Nominal (D), OT Pot (E) - Stacked
        # Line: Target (B)

        c1 = BarChart()
        c1.type = "col"
        c1.style = 10
        c1.grouping = "stacked"
        c1.overlap = 100
        c1.title = f"{zone} Capacity Constraints"
        c1.y_axis.title = "Units"
        c1.x_axis.title = "Fortnight"
        c1.height = 10
        c1.width = 15

        # Data
        data_nominal = Reference(ws1, min_col=4, min_row=data_start, max_row=data_start+7)
        data_ot = Reference(ws1, min_col=5, min_row=data_start, max_row=data_start+7)
        cats = Reference(ws1, min_col=1, min_row=data_start, max_row=data_start+7)

        s1 = Series(data_nominal, title="Nominal Cap")
        s1.graphicalProperties.solidFill = "70AD47" # Green
        c1.append(s1)

        s2 = Series(data_ot, title="OT Potential")
        s2.graphicalProperties.solidFill = "ED7D31" # Orange
        c1.append(s2)
        c1.set_categories(cats)

        # Line Chart for Target
        c2 = LineChart()
        data_target = Reference(ws1, min_col=2, min_row=data_start, max_row=data_start+7)
        s3 = Series(data_target, title="Target Demand")
        s3.graphicalProperties.line.solidFill = "FF0000" # Red
        s3.graphicalProperties.line.width = 30000 # Thick
        c2.append(s3)

        c1 += c2 # Combine
        ws1.add_chart(c1, f"K{chart_anchor_row}")

        # 2. Overtime Cost Cliff (Line)
        # X: FN, Y: Unit Cost
        c3 = LineChart()
        c3.title = f"{zone} Unit Cost Analysis"
        c3.y_axis.title = "Unit Cost ($)"
        c3.style = 13
        c3.height = 8
        c3.width = 15

        data_cost = Reference(ws1, min_col=8, min_row=data_start, max_row=data_start+7)
        s4 = Series(data_cost, title="Est. Unit Cost")
        s4.graphicalProperties.line.solidFill = "2F5496" # Blue
        c3.append(s4)
        c3.set_categories(cats)

        ws1.add_chart(c3, f"K{chart_anchor_row + 21}")

        # ---------------------------------------------------------
        # CONDITIONAL FORMATTING
        # ---------------------------------------------------------

        # Real Output < Target -> Red Fill (Bottleneck)
        # Formula: G{row} < B{row}
        # Note: In FormulaRule, openpyxl expects the formula relative to the top-left cell of the range.
        red_fill_fmt = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        dxf = Font(color="9C0006")

        range_string = f"G{data_start}:G{data_start+7}"
        # Start cell is G{data_start}. Formula should be G{data_start} < B{data_start}
        ws1.conditional_formatting.add(range_string,
            FormulaRule(formula=[f"G{data_start}<B{data_start}"], stopIfTrue=True, fill=red_fill_fmt))

        # Cost > 40 -> Red Text
        range_string_cost = f"H{data_start}:H{data_start+7}"
        red_text = Font(color="FF0000", bold=True)
        # Start cell H{data_start}. Formula H{data_start} > 40
        ws1.conditional_formatting.add(range_string_cost,
             FormulaRule(formula=[f"H{data_start}>{DEFAULT_VARIABLE_COST}"], font=red_text))


        row += 20  # Space for next zone (charts take up space)

    # Column widths
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10 # OT
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 14
    ws1.column_dimensions['I'].width = 16

    # =========================================================================
    # TAB 2: RESOURCE_MGR
    # =========================================================================
    ws2 = wb.create_sheet("RESOURCE_MGR")

    ws2['A1'] = "RESOURCE MANAGER - Zone-by-Zone Asset Allocation"
    ws2['A1'].font = title_font

    # Section A: Assignments by Zone
    ws2['A3'] = "SECTION A: MACHINE ASSIGNMENTS BY ZONE"
    ws2['A3'].font = section_font

    assign_headers = ['Zone', 'Section', 'Machines Assigned', 'Workers Needed', 'Status']
    for col, h in enumerate(assign_headers, start=1):
        cell = ws2.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 6
    for zone in ZONES:
        for section in SECTIONS:
            cell = ws2.cell(row=row, column=1, value=zone)
            cell.border = thin_border
            cell.fill = zone_fills[zone]
            cell.font = Font(color="FFFFFF")

            ws2.cell(row=row, column=2, value=section).border = thin_border

            # Default machines (only Center has machines initially)
            default_machines = 0
            if zone == "Center":
                default_machines = {'Section 1': 20, 'Section 2': 20, 'Section 3': 17}.get(section, 10)

            cell = ws2.cell(row=row, column=3, value=default_machines)
            cell.border = thin_border
            cell.fill = input_fill

            cell = ws2.cell(row=row, column=4, value=f"=C{row}*5")
            cell.border = thin_border
            cell.fill = calc_fill

            cell = ws2.cell(row=row, column=5, value="OK")
            cell.border = thin_border
            cell.fill = output_fill

            row += 1

    assign_end = row - 1
    row += 3

    # Section B: Expansion by Zone
    ws2.cell(row=row, column=1, value="SECTION B: EXPANSION RECOMMENDATIONS BY ZONE").font = section_font
    row += 2

    exp_headers = ['Zone', 'Target Capacity', 'Current Machines', 'Capacity Gap', 'Recommendation']
    for col, h in enumerate(exp_headers, start=1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    expansion_start = row
    for zone in ZONES:
        cell = ws2.cell(row=row, column=1, value=zone)
        cell.border = thin_border
        cell.fill = zone_fills[zone]
        cell.font = Font(color="FFFFFF")

        # Link to Zone Calculator total
        # IMPORTANT: Updated logic as cols shifted
        # Target was col 2. Real Output was col 6 (now 7).
        zone_total_row = zone_output_rows.get(zone, {}).get('total', 10)
        cell = ws2.cell(row=row, column=2, value=f"=ZONE_CALCULATORS!B{zone_total_row}")
        cell.border = thin_border
        cell.fill = ref_fill

        # Current machines
        current_machines = machines_data.get(zone, {}).get('machines', 0)
        ws2.cell(row=row, column=3, value=current_machines).border = thin_border

        # Capacity Gap
        cell = ws2.cell(row=row, column=4, value=f"=B{row}-(C{row}*{DEFAULT_NOMINAL_RATE})")
        cell.border = thin_border
        cell.fill = calc_fill

        # Recommendation
        # Recommendation
        # Logic: Compare M1 (Low Cost) vs M3-beta (High efficiency)
        # For simplicity, calculate number of M1s needed.
        m1_cost = MACHINES.get('M1', {}).get('PURCHASE_COST', 10600)
        m3_cost = MACHINES.get('M3_BETA', {}).get('PURCHASE_COST', 155400)
        m3_cap = MACHINES.get('M3_BETA', {}).get('CAPACITY', 1000)
        m1_cap = DEFAULT_NOMINAL_RATE

        # Calculate Machines Needed
        # M1s Needed = Gap / M1_Cap
        # M3s Needed = Gap / M3_Cap

        # We will split Recommendation into "M1 Option" and "M3 Option"
        # We need more columns.
        # Shifted columns manually in header above? No, I need to rewrite header code if I want more columns.
        # Alternative: Just put a complex string in the single Recommendation cell.

        rec_formula = (
            f'=IF(D{row}>0, '
            f'"M1 Opt: "&ROUNDUP(D{row}/{m1_cap},0)&" @ ${m1_cost/1000:.1f}k ea | " &'
            f'"M3 Opt: "&ROUNDUP(D{row}/{m3_cap},0)&" @ ${m3_cost/1000:.1f}k ea", '
            f'"OK")'
        )

        cell = ws2.cell(row=row, column=5, value=rec_formula)
        cell.border = thin_border
        cell.fill = output_fill
        cell.font = Font(size=9) # Smaller font to fit

        row += 1

    row += 3

    # Section C: Real Estate by Zone
    ws2.cell(row=row, column=1, value="SECTION C: REAL ESTATE (MODULES) BY ZONE").font = section_font
    row += 2

    re_headers = ['Zone', 'Machines', 'Module Slots', 'Free Slots', 'Recommendation']
    for col, h in enumerate(re_headers, start=1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for zone in ZONES:
        zone_machines = machines_data.get(zone, {})

        cell = ws2.cell(row=row, column=1, value=zone)
        cell.border = thin_border
        cell.fill = zone_fills[zone]
        cell.font = Font(color="FFFFFF")

        ws2.cell(row=row, column=2, value=zone_machines.get('machines', 0)).border = thin_border
        ws2.cell(row=row, column=3, value=zone_machines.get('modules', 0)).border = thin_border

        cell = ws2.cell(row=row, column=4, value=f"=C{row}-B{row}")
        cell.border = thin_border
        cell.fill = calc_fill

        cell = ws2.cell(row=row, column=5,
            value=f'=IF(D{row}<5, "Buy module in {zone}", "OK")')
        cell.border = thin_border
        cell.fill = output_fill

        row += 1

    # Column widths
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 28

    # =========================================================================
    # TAB 3: UPLOAD_READY_PRODUCTION
    # =========================================================================
    ws3 = wb.create_sheet("UPLOAD_READY_PRODUCTION")

    ws3['A1'] = "PRODUCTION DECISIONS - ExSim Upload (Zone-Mapped)"
    ws3['A1'].font = title_font
    ws3['A2'] = "Values linked from Zone Calculators"
    ws3['A2'].font = Font(italic=True, color="666666")

    # Block 1: Production Targets by Zone
    ws3['A4'] = "Production Targets"
    ws3['A4'].font = section_font

    headers = ['Zone', 'Product', 'Target', 'Overtime']
    for col, h in enumerate(headers, start=1):
        cell = ws3.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 6
    for zone in ZONES:
        cell = ws3.cell(row=row, column=1, value=zone)
        cell.border = thin_border
        cell.fill = zone_fills[zone]
        cell.font = Font(color="FFFFFF")

        ws3.cell(row=row, column=2, value="A").border = thin_border

        # Link to zone total output
        # Updated: Real Output was Col 7 in Zone Calc
        # Total Real Output is in Col 7
        zone_total = zone_output_rows.get(zone, {}).get('total', 10)
        cell = ws3.cell(row=row, column=3, value=f"=ZONE_CALCULATORS!G{zone_total}")
        cell.border = thin_border
        cell.fill = calc_fill

        # Link to first FN overtime of zone
        zone_data_start = zone_output_rows.get(zone, {}).get('data_start', 10)
        cell = ws3.cell(row=row, column=4, value=f"=ZONE_CALCULATORS!C{zone_data_start}")
        cell.border = thin_border
        cell.fill = calc_fill

        row += 1

    row += 2

    # Block 2: Machine Purchases by Zone
    ws3.cell(row=row, column=1, value="Machine Purchases").font = section_font
    row += 1

    headers = ['Zone', 'Machine Type', 'Quantity']
    for col, h in enumerate(headers, start=1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for zone in ZONES:
        for mt in ['M1', 'M2']:  # Most common machine types
            cell = ws3.cell(row=row, column=1, value=zone)
            cell.border = thin_border
            cell.fill = zone_fills[zone]
            cell.font = Font(color="FFFFFF")

            ws3.cell(row=row, column=2, value=mt).border = thin_border

            cell = ws3.cell(row=row, column=3, value=0)
            cell.border = thin_border
            cell.fill = input_fill

            row += 1

    row += 2

    # Block 3: Section Assignments by Zone
    ws3.cell(row=row, column=1, value="Section Assignments").font = section_font
    row += 1

    headers = ['Zone', 'Section', 'Machines', 'Workers']
    for col, h in enumerate(headers, start=1):
        cell = ws3.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    # Link to RESOURCE_MGR assignments
    assign_row = 6
    for zone in ZONES:
        for section in SECTIONS:
            cell = ws3.cell(row=row, column=1, value=zone)
            cell.border = thin_border
            cell.fill = zone_fills[zone]
            cell.font = Font(color="FFFFFF")

            ws3.cell(row=row, column=2, value=section).border = thin_border

            cell = ws3.cell(row=row, column=3, value=f"=RESOURCE_MGR!C{assign_row}")
            cell.border = thin_border
            cell.fill = calc_fill

            cell = ws3.cell(row=row, column=4, value=f"=RESOURCE_MGR!D{assign_row}")
            cell.border = thin_border
            cell.fill = calc_fill

            assign_row += 1
            row += 1

    # Column widths
    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 15

    # Save
    wb.save(OUTPUT_FILE)
    print(f"[SUCCESS] Created '{OUTPUT_FILE}'")


def main():
    """Main function."""
    print("ExSim Production Dashboard Zones Generator")
    print("=" * 50)

    print("\n[*] Loading zone-specific data files...")
    print(f"    Primary source: {REPORTS_FOLDER}")
    print(f"    Fallback source: {LOCAL_DATA_FOLDER}")

    # Raw Materials by Zone
    materials_path = get_data_path("raw_materials.xlsx")
    if materials_path:
        materials_data = load_raw_materials_by_zone(materials_path)
        print(f"  [OK] Loaded materials from {materials_path.parent.name}/")
    else:
        materials_data = load_raw_materials_by_zone(None)
        print("  [!] Using default materials data")

    # Finished Goods by Zone
    fg_path = get_data_path("finished_goods_inventory.xlsx")
    if fg_path:
        fg_data = load_finished_goods_by_zone(fg_path)
        print(f"  [OK] Loaded finished goods by zone")
    else:
        fg_data = load_finished_goods_by_zone(None)
        print("  [!] Using default FG data")

    # Workers by Zone
    workers_path = get_data_path("workers_balance_overtime.xlsx")
    if workers_path:
        workers_data = load_workers_by_zone(workers_path)
        print(f"  [OK] Loaded workers by zone")
    else:
        workers_data = load_workers_by_zone(None)
        print("  [!] Using default workers data")

    # Machines by Zone
    machines_path = get_data_path("machine_spaces.xlsx")
    if machines_path:
        machines_data = load_machines_by_zone(machines_path)
        print(f"  [OK] Loaded machines by zone")
    else:
        machines_data = load_machines_by_zone(None)
        print("  [!] Using default machines data")

    # Template
    template_path = get_data_path("Production Decisions.xlsx")
    template_data = load_production_template(template_path)
    if template_data['exists']:
        print(f"  [OK] Loaded production template")
    else:
        print("  [!] Using default template layout")

    print("\n[*] Creating dashboard...")

    create_zones_dashboard(materials_data, fg_data, workers_data,
                           machines_data, template_data)

    print("\nSheets created:")
    print("  * ZONE_CALCULATORS (5 Zone-Specific Production Blocks)")
    print("  * RESOURCE_MGR (Assignments/Expansion/Modules by Zone)")
    print("  * UPLOAD_READY_PRODUCTION (ExSim Format)")


if __name__ == "__main__":
    main()
