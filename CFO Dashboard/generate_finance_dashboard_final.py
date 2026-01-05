"""
ExSim Finance Dashboard Final - Financial Control & Liquidity

Integrates cash flow monitoring, income statement projection,
balance sheet health tracking, and debt management.

Required libraries: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference, Series
import warnings
import sys

# Add parent directory to path to import case_parameters
sys.path.append(str(Path(__file__).parent.parent))
try:
    from case_parameters import FINANCIAL, COMMON
    from config import get_data_path, OUTPUT_DIR
except ImportError:
    print("Warning: Could not import case_parameters.py or config.py. Using defaults.")
    FINANCIAL = {}
    COMMON = {}
    # Fallback for config if not found (though it should be providing we are in right structure)
    OUTPUT_DIR = Path(__file__).parent
    def get_data_path(f): return Path(f)

# Import shared outputs for inter-dashboard communication
try:
    from shared_outputs import export_dashboard_data, import_dashboard_data
except ImportError:
    export_dashboard_data = None
    import_dashboard_data = None

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Required input files from centralized Reports folder
REQUIRED_FILES = [
    'initial_cash_flow.xlsx',
    'results_and_balance_statements.xlsx',
    'sales_admin_expenses.xlsx',
    'accounts_receivable_payable.xlsx',
    'Finance Decisions.xlsx'
]

OUTPUT_FILE = OUTPUT_DIR / "Finance_Dashboard_Final.xlsx"

FORTNIGHTS = COMMON.get('FORTNIGHTS', list(range(1, 9)))  # From centralized config

# Default financial parameters
DEF_LOANS = FINANCIAL.get('LOANS', {})
ST_LIMIT = DEF_LOANS.get('SHORT_TERM', {}).get('LIMIT', 500000)
LT_LIMIT = DEF_LOANS.get('LONG_TERM', {}).get('LIMIT', 2000000)
DEFAULT_INTEREST_RATE = DEF_LOANS.get('SHORT_TERM', {}).get('INTEREST_RATE_ANNUAL', 0.08)
DEFAULT_TAX_RATE = FINANCIAL.get('TAX_RATE', 0.25)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_numeric(value):
    """Parse formatted number strings."""
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
    # Handle parentheses as negative
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned)
    except:
        return 0


def load_excel_file(filepath, sheet_name=None):
    """Load Excel file."""
    try:
        if sheet_name:
            return pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


# =============================================================================
# DATA LOADING
# =============================================================================

def load_initial_cash_flow(filepath):
    """Load initial cash flow for starting position."""
    df = load_excel_file(filepath)
    
    data = {
        'final_cash': 0,
        'tax_payments': 0
    }
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'final cash' in first_val:
            for col_idx in range(1, min(10, len(row))):
                val = parse_numeric(row.iloc[col_idx])
                if val != 0:
                    data['final_cash'] = val
                    break
        
        if 'tax' in first_val and 'payment' in first_val:
            for col_idx in range(1, min(10, len(row))):
                val = parse_numeric(row.iloc[col_idx])
                if val != 0:
                    data['tax_payments'] = abs(val)
                    break
    
    return data


def load_balance_statements(filepath):
    """Load results and balance statements for historical data."""
    df = load_excel_file(filepath)
    
    data = {
        'net_sales': 0,
        'cogs': 0,
        'gross_income': 0,
        'net_profit': 0,
        'total_assets': 0,
        'total_liabilities': 0,
        'equity': 0,
        'retained_earnings': 0,
        'depreciation': 0,
        'gross_margin_pct': 0,
        'net_margin_pct': 0
    }
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'net sales' in first_val or 'revenue' in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val > 0:
                data['net_sales'] = val
        
        if 'cost of goods sold' in first_val or 'cogs' in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val != 0:
                data['cogs'] = abs(val)
        
        if 'gross' in first_val and ('income' in first_val or 'profit' in first_val or 'margin' in first_val):
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val != 0:
                data['gross_income'] = val
        
        if ('net profit' in first_val or 'net income' in first_val) and 'before' not in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val != 0:
                data['net_profit'] = val
        
        if 'total assets' in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val > 0:
                data['total_assets'] = val
        
        if 'total liabilities' in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val != 0:
                data['total_liabilities'] = abs(val)
        
        if first_val == 'equity' or 'total equity' in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val != 0:
                data['equity'] = val
        
        if 'retained earnings' in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            data['retained_earnings'] = val
        
        if 'depreciation' in first_val:
            val = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
            if val != 0:
                data['depreciation'] = abs(val)
    
    # Calculate margins (with division protection)
    if data['net_sales'] > 0:
        data['gross_margin_pct'] = data['gross_income'] / data['net_sales']
        data['net_margin_pct'] = data['net_profit'] / data['net_sales']
    
    return data


def load_sales_admin_expenses(filepath):
    """Load S&A expenses for overhead estimation."""
    df = load_excel_file(filepath)
    
    data = {'total_sa_expenses': 0}
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'total' in first_val and ('sales' in first_val or 'admin' in first_val or 's&a' in first_val):
            for col_idx in range(1, min(12, len(row))):
                val = parse_numeric(row.iloc[col_idx])
                if val > 0:
                    data['total_sa_expenses'] = val
                    break
    
    return data


def load_receivables_payables(filepath):
    """Load accounts receivable and payable."""
    df = load_excel_file(filepath)
    
    data = {
        'receivables': [0] * 8,
        'payables': [0] * 8
    }
    
    if df is None:
        return data
    
    # Parse by fortnight from cash flow report if available
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        # Receivables (Cash In)
        if 'receipts' in first_val and 'customer' in first_val:
            for i in range(8):
                if i + 1 < len(row):
                    data['receivables'][i] = parse_numeric(row.iloc[i+1])
        
        # Payables (Cash Out)
        if 'payment' in first_val and 'supplier' in first_val:
            for i in range(8):
                if i + 1 < len(row):
                    data['payables'][i] = abs(parse_numeric(row.iloc[i+1]))
    
    return data


def load_finance_template(filepath):
    """Load finance decisions template."""
    try:
        df = pd.read_excel(filepath, sheet_name='Finance', header=None)
        return {'df': df, 'exists': True}
    except:
        return {'df': None, 'exists': False}


# =============================================================================
# HARD DATA LOADERS - NEW PRECISE EXTRACTORS
# =============================================================================

def load_machine_depreciation(filepath):
    """
    Extract Period Amortization / Leasing from machine_spaces.xlsx.
    Sum to get total depreciation expense.
    """
    if filepath is None or not filepath.exists():
        return 0.0
    
    try:
        df = pd.read_excel(filepath, header=None)
        
        # Find column with "Period Amortization"
        for idx, row in df.iterrows():
            values = [str(v).strip() if pd.notna(v) else "" for v in row]
            if any("Period Amortization" in v or "Amortization" in v for v in values):
                # Found header row - get column index
                amort_col = None
                for col_idx, val in enumerate(values):
                    if "Period Amortization" in val or "Amortization" in val:
                        amort_col = col_idx
                        break
                
                if amort_col:
                    total = 0.0
                    for offset in range(1, 15):
                        if idx + offset < len(df):
                            data_row = df.iloc[idx + offset]
                            machine = str(data_row.iloc[0]).strip() if pd.notna(data_row.iloc[0]) else ""
                            if "Total" in machine and any(m in machine for m in ["M1", "M2", "M3", "M4"]):
                                total += parse_numeric(data_row.iloc[amort_col])
                    return total
                break
    except Exception as e:
        print(f"  [!] Error in load_machine_depreciation: {e}")
    
    return 0.0


def load_initial_cash_precise(filepath):
    """
    Extract exact "Final cash (at the start of the first fortnight)" value.
    This is the TRUE starting cash for FN1.
    """
    if filepath is None or not filepath.exists():
        return 0.0
    
    try:
        df = pd.read_excel(filepath, header=None)
        
        for idx, row in df.iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if "Final cash" in label and "start of the first fortnight" in label:
                return parse_numeric(row.iloc[1]) if len(row) > 1 else 0.0
        
        # Fallback: look for just "Final cash"
        for idx, row in df.iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if "Final cash" in label:
                return parse_numeric(row.iloc[1]) if len(row) > 1 else 0.0
                
    except Exception as e:
        print(f"  [!] Error in load_initial_cash_precise: {e}")
    
    return 0.0


def load_hard_schedule_precise(filepath):
    """
    Extract Receivables and Payables for Fortnights 1-8.
    Returns dict: {1: {'receivables': x, 'payables': y}, ...}
    """
    data = {fn: {'receivables': 0.0, 'payables': 0.0} for fn in range(1, 9)}
    
    if filepath is None or not filepath.exists():
        return data
    
    try:
        df = pd.read_excel(filepath, header=None)
        
        # Find header row with "Fortnight", "Receivables", "Payables"
        for idx, row in df.iterrows():
            values = [str(v).strip() if pd.notna(v) else "" for v in row]
            if "Fortnight" in values:
                rec_col = None
                pay_col = None
                for col_idx, val in enumerate(values):
                    if "Receivables" in val:
                        rec_col = col_idx
                    if "Payables" in val:
                        pay_col = col_idx
                
                # Extract data for FN 1-8
                for fn in range(1, 9):
                    if idx + fn < len(df):
                        data_row = df.iloc[idx + fn]
                        fn_val = int(parse_numeric(data_row.iloc[0]))
                        if fn_val == fn:
                            if rec_col is not None:
                                data[fn]['receivables'] = parse_numeric(data_row.iloc[rec_col])
                            if pay_col is not None:
                                data[fn]['payables'] = parse_numeric(data_row.iloc[pay_col])
                break
                
    except Exception as e:
        print(f"  [!] Error in load_hard_schedule_precise: {e}")
    
    return data


def load_retained_earnings(filepath):
    """
    Extract Retained Earnings from results_and_balance_statements.xlsx.
    This is the max dividend capacity.
    """
    if filepath is None or not filepath.exists():
        return 0.0
    
    try:
        df = pd.read_excel(filepath, header=None)
        
        for idx, row in df.iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if "Retained Earnings" in label:
                # Check columns for the value
                for col_idx in range(1, min(10, len(row))):
                    val = parse_numeric(row.iloc[col_idx])
                    if val != 0:
                        return val
                        
    except Exception as e:
        print(f"  [!] Error in load_retained_earnings: {e}")
    
    return 0.0


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def create_finance_dashboard(cash_data, balance_data, sa_data, ar_ap_data, template_data, hard_data=None, output_buffer=None):
    """
    Create the Finance Dashboard with Hard Data injection.
    
    Args:
        cash_data: Cash flow data dict
        balance_data: Balance sheet data dict
        sa_data: Sales & Admin expenses dict
        ar_ap_data: Accounts receivable/payable dict
        template_data: Template configuration
        hard_data: Hard data values dict (optional)
        output_buffer: io.BytesIO buffer for output (optional). If provided, returns
                      the buffer instead of saving to disk.
    
    Returns:
        BytesIO buffer if output_buffer provided, None otherwise
    """
    
    # Handle missing hard_data
    if hard_data is None:
        hard_data = {
            'depreciation': 0,
            'starting_cash': 0,
            'schedule': {fn: {'receivables': 0, 'payables': 0} for fn in range(1, 9)},
            'retained_earnings': 0
        }
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5496")
    title_font = Font(bold=True, size=14, color="2F5496")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ref_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # NEW: Hard data fill - gray background for read-only hard data
    hard_data_fill = PatternFill(start_color="B4B4B4", end_color="B4B4B4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # =========================================================================
    # TAB 1: LIQUIDITY_MONITOR
    # =========================================================================
    ws1 = wb.active
    ws1.title = "LIQUIDITY_MONITOR"
    
    ws1['A1'] = "LIQUIDITY MONITOR - Cash Flow Engine"
    ws1['A1'].font = title_font
    
    # Section A: Initialization
    ws1['A3'] = "SECTION A: INITIALIZATION (Initial Cash Flow Bridge)"
    ws1['A3'].font = section_font
    
    ws1.cell(row=5, column=1, value="Cash at End of Last Period").border = thin_border
    cell = ws1.cell(row=5, column=2, value=cash_data.get('final_cash', 0))
    cell.border = thin_border
    cell.fill = ref_fill
    cell.number_format = '$#,##0'
    
    ws1.cell(row=6, column=1, value="Less: Tax Payments").border = thin_border
    cell = ws1.cell(row=6, column=2, value=cash_data.get('tax_payments', 0))
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    
    ws1.cell(row=7, column=1, value="Less: Dividend Payments").border = thin_border
    cell = ws1.cell(row=7, column=2, value=0)
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    
    ws1.cell(row=8, column=1, value="Less: Asset Purchases").border = thin_border
    cell = ws1.cell(row=8, column=2, value=0)
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    
    ws1.cell(row=9, column=1, value="STARTING CASH FOR FN1 (HARD)").font = Font(bold=True)
    # Use hard_data starting cash if available, otherwise formula
    starting_cash_value = hard_data.get('starting_cash', 0)
    if starting_cash_value > 0:
        cell = ws1.cell(row=9, column=2, value=starting_cash_value)
        cell.fill = hard_data_fill  # Gray = read-only hard data
    else:
        cell = ws1.cell(row=9, column=2, value="=B5-B6-B7-B8")
        cell.fill = output_fill
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.number_format = '$#,##0'
    ws1.cell(row=9, column=3, value="← From initial_cash_flow.xlsx").font = Font(italic=True, color="666666")
    
    # Section B: Operational Cash Flow
    row = 12
    ws1.cell(row=row, column=1, value="SECTION B: OPERATIONAL CASH FLOW").font = section_font
    row += 2
    
    # Headers
    ws1.cell(row=row, column=1, value="Item").font = header_font
    ws1.cell(row=row, column=1).fill = header_fill
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=f"FN{fn}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Opening Cash
    # The offset calculation needs to account for ALL rows between open_cash_row and ending_cash_row.
    # After adding Debt Capacity rows, the offset has increased.
    # Layout from open_cash_row:
    #   +1 Sales, +2 Procurement, +3 S&A, +4 Receivables, +5 Payables
    #   +6 gap, +7 gap (extra), +8 Section C header, +9 Credit, +10 ST Debt Balance
    #   +11 Investments, +12 Mortgage, +13 Dividends, +14 gap
    #   +15 Section D header, +16 Net Flow, +17 Ending Cash
    ending_cash_offset = 18  # Updated: open_cash_row + 18 = ending_cash_row
    
    ws1.cell(row=row, column=1, value="Opening Cash").border = thin_border
    open_cash_row = row
    for fn in FORTNIGHTS:
        if fn == 1:
            # FN1 Opening = Starting Cash from Section A
            cell = ws1.cell(row=row, column=1+fn, value="=$B$9")
        else:
            # FN2+ Opening = Previous FN's Ending Cash
            # Previous column is (1+fn-1) = fn, ending cash is at open_cash_row + ending_cash_offset
            prev_col = get_column_letter(fn)  # Previous fortnight column
            cell = ws1.cell(row=row, column=1+fn, value=f"={prev_col}${row + ending_cash_offset}")
        cell.border = thin_border
        cell.fill = ref_fill
        cell.number_format = '$#,##0'
    row += 1
    
    # Sales Receipts
    ws1.cell(row=row, column=1, value="Sales Receipts (Est.)").border = thin_border
    sales_row = row
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=100000)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
    row += 1
    
    # Procurement Spend
    ws1.cell(row=row, column=1, value="Procurement Spend (Est.)").border = thin_border
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
    procurement_row = row
    row += 1
    
    # Fixed Overhead (S&A)
    sa_per_fn = sa_data.get('total_sa_expenses', 0) / 8
    ws1.cell(row=row, column=1, value="Fixed Overhead (S&A)").border = thin_border
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=int(sa_per_fn))
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
    sa_row = row
    row += 1
    
    # Receivables - inject HARD DATA
    ws1.cell(row=row, column=1, value="Receivables (HARD)").border = thin_border
    recv_row = row
    for fn in FORTNIGHTS:
        # Use hard_data schedule if available
        hard_recv = hard_data.get('schedule', {}).get(fn, {}).get('receivables', 0)
        if hard_recv > 0:
            cell = ws1.cell(row=row, column=1+fn, value=hard_recv)
            cell.fill = hard_data_fill  # Gray = read-only
        else:
            cell = ws1.cell(row=row, column=1+fn, value=ar_ap_data['receivables'][fn-1])
            cell.fill = ref_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
    row += 1
    
    # Payables - inject HARD DATA
    ws1.cell(row=row, column=1, value="Payables (HARD)").border = thin_border
    pay_row = row
    for fn in FORTNIGHTS:
        # Use hard_data schedule if available (as negative for outflow)
        hard_pay = hard_data.get('schedule', {}).get(fn, {}).get('payables', 0)
        if hard_pay > 0:
            cell = ws1.cell(row=row, column=1+fn, value=-hard_pay)  # Negative = outflow
            cell.fill = hard_data_fill  # Gray = read-only
        else:
            cell = ws1.cell(row=row, column=1+fn, value=ar_ap_data['payables'][fn-1])
            cell.fill = ref_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
    row += 2
    
    row += 2
    
    # Section C: Financing Decisions
    ws1.cell(row=row, column=1, value="SECTION C: FINANCING DECISIONS").font = section_font
    
    # NEW: Debt Capacity Meter (Header)
    ws1.cell(row=row, column=3, value="DEBT CAPACITY (Short Term)").font = Font(bold=True)
    ws1.cell(row=row, column=4, value=f"Limit: ${ST_LIMIT:,.0f}").font = Font(italic=True)
    
    row += 1
    
    ws1.cell(row=row, column=1, value="Change in Credit Line (+/-)").border = thin_border
    credit_row = row
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        # Validation: Flag if exceeds limit (simplified check: assume starting 0 used)
        # Real check needs cumulative balance. 
        # For now, just add a tip/warning if input > 500k.
        # Ideally, we track 'Short Term Debt Balance' row.
    
    # Add ST Debt Balance Row (Hidden calculation or explicit?)
    # Let's add it explicitly to track capacity.
    row += 1
    ws1.cell(row=row, column=1, value="Proj. ST Debt Balance").font = Font(italic=True)
    balance_st_row = row
    for fn in FORTNIGHTS:
        col_let = get_column_letter(1+fn)
        prev_bal = 0 if fn == 1 else f"{get_column_letter(fn)}{row}" # Simplified: assumes 0 start or hardcoded start
        # Formula: Previous + New Borrowing
        ws1.cell(row=row, column=1+fn, value=f"={prev_bal}+{col_let}{credit_row}").number_format = '$#,##0'
        
        # Conditional Format for Limit
        # Red if > Limit
        
    row += 1
    

    
    ws1.cell(row=row, column=1, value="Change in Investments (+/-)").border = thin_border
    invest_row = row
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
    row += 1
    
    ws1.cell(row=row, column=1, value="New Mortgage Inflow").border = thin_border
    mortgage_row = row
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
    row += 1
    
    ws1.cell(row=row, column=1, value="Dividends Paid").border = thin_border
    dividend_row = row
    for fn in FORTNIGHTS:
        cell = ws1.cell(row=row, column=1+fn, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
    row += 2
    
    # Section D: The Balance
    ws1.cell(row=row, column=1, value="SECTION D: CASH BALANCE").font = section_font
    row += 1
    
    ws1.cell(row=row, column=1, value="Net Cash Flow").border = thin_border
    net_flow_row = row
    for fn in FORTNIGHTS:
        col = 1 + fn
        col_letter = get_column_letter(col)
        # Net = Sales + Receivables + Credit + Mortgage - Procurement - S&A - Payables - Investments - Dividends
        cell = ws1.cell(row=row, column=col,
            value=f"={col_letter}{sales_row}+{col_letter}{recv_row}+{col_letter}{credit_row}+{col_letter}{mortgage_row}-{col_letter}{procurement_row}-{col_letter}{sa_row}-{col_letter}{pay_row}-{col_letter}{invest_row}-{col_letter}{dividend_row}")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
    row += 1
    
    ws1.cell(row=row, column=1, value="ENDING CASH BALANCE").font = Font(bold=True)
    ending_cash_row = row
    for fn in FORTNIGHTS:
        col = 1 + fn
        col_letter = get_column_letter(col)
        cell = ws1.cell(row=row, column=col,
            value=f"={col_letter}{open_cash_row}+{col_letter}{net_flow_row}")
        cell.border = thin_border
        cell.fill = output_fill
        cell.font = Font(bold=True)
        cell.number_format = '$#,##0'
    row += 1
    
    ws1.cell(row=row, column=1, value="Solvency Check").border = thin_border
    for fn in FORTNIGHTS:
        col = 1 + fn
        col_letter = get_column_letter(col)
        cell = ws1.cell(row=row, column=col,
            value=f'=IF({col_letter}{ending_cash_row}<0,"INSOLVENT!",IF({col_letter}{ending_cash_row}>200000,"Excess Cash","OK"))')
        cell.border = thin_border
    
    # Add conditional formatting
    ws1.conditional_formatting.add(
        f'B{ending_cash_row}:I{ending_cash_row}',
        FormulaRule(formula=[f'B{ending_cash_row}<0'], fill=red_fill)
    )
    ws1.conditional_formatting.add(
        f'B{ending_cash_row}:I{ending_cash_row}',
        FormulaRule(formula=[f'B{ending_cash_row}>200000'], fill=green_fill)
    )
    
    # ---------------------------------------------------------
    # CHART: Liquidity Forecast (Line Chart)
    # ---------------------------------------------------------
    chart_liq = LineChart()
    chart_liq.title = "Liquidity Forecast (Cash Balance)"
    chart_liq.style = 12
    chart_liq.y_axis.title = "Cash Balance ($)"
    chart_liq.x_axis.title = "Fortnight"
    chart_liq.height = 10
    chart_liq.width = 15

    # Data: Ending Cash (Row ending_cash_row, Cols B-I)
    # Fortnights are columns 2 to 9
    data_cash = Reference(ws1, min_col=2, min_row=ending_cash_row, max_col=9)
    s_cash = Series(data_cash, title="Ending Cash")
    chart_liq.append(s_cash)

    # Categories (FN1..FN8 headers at Row 14)
    cats_liq = Reference(ws1, min_col=2, min_row=14, max_col=9)
    chart_liq.set_categories(cats_liq)

    ws1.add_chart(chart_liq, "H2")
    
    # Column widths
    ws1.column_dimensions['A'].width = 28
    for col in range(2, 10):
        ws1.column_dimensions[get_column_letter(col)].width = 14
    
    # =========================================================================
    # TAB 2: PROFIT_CONTROL
    # =========================================================================
    ws2 = wb.create_sheet("PROFIT_CONTROL")
    
    ws2['A1'] = "PROFIT CONTROL - Income Statement Forecast vs Actuals"
    ws2['A1'].font = title_font
    
    ws2['A3'] = "HISTORICAL MARGINS (From results_and_balance_statements)"
    ws2['A3'].font = section_font
    
    ws2.cell(row=5, column=1, value="Historical Gross Margin %").border = thin_border
    cell = ws2.cell(row=5, column=2, value=balance_data.get('gross_margin_pct', 0))
    cell.border = thin_border
    cell.fill = ref_fill
    cell.number_format = '0.0%'
    
    ws2.cell(row=6, column=1, value="Historical Net Margin %").border = thin_border
    cell = ws2.cell(row=6, column=2, value=balance_data.get('net_margin_pct', 0))
    cell.border = thin_border
    cell.fill = ref_fill
    cell.number_format = '0.0%'
    
    # Income Statement Comparison
    row = 9
    ws2.cell(row=row, column=1, value="INCOME STATEMENT COMPARISON").font = section_font
    row += 1
    
    headers = ['Line Item', 'Last Round Actuals', 'This Round Projected', 'Variance %']
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    # Revenue
    ws2.cell(row=row, column=1, value="Net Sales / Revenue").border = thin_border
    ws2.cell(row=row, column=2, value=balance_data.get('net_sales', 0)).border = thin_border
    ws2['B' + str(row)].fill = ref_fill
    ws2['B' + str(row)].number_format = '$#,##0'
    cell = ws2.cell(row=row, column=3, value=balance_data.get('net_sales', 0))
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    cell = ws2.cell(row=row, column=4, value=f"=IF(B{row}>0,(C{row}-B{row})/B{row},0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    revenue_row = row
    row += 1
    
    # COGS
    ws2.cell(row=row, column=1, value="Cost of Goods Sold").border = thin_border
    ws2.cell(row=row, column=2, value=balance_data.get('cogs', 0)).border = thin_border
    ws2['B' + str(row)].fill = ref_fill
    ws2['B' + str(row)].number_format = '$#,##0'
    cell = ws2.cell(row=row, column=3, value=f"=C{revenue_row}*(1-$B$5)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '$#,##0'
    cell = ws2.cell(row=row, column=4, value=f"=IF(B{row}>0,(C{row}-B{row})/B{row},0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    cogs_row = row
    row += 1
    
    # Gross Margin
    ws2.cell(row=row, column=1, value="Gross Margin").border = thin_border
    ws2.cell(row=row, column=2, value=balance_data.get('gross_income', 0)).border = thin_border
    ws2['B' + str(row)].fill = ref_fill
    ws2['B' + str(row)].number_format = '$#,##0'
    cell = ws2.cell(row=row, column=3, value=f"=C{revenue_row}-C{cogs_row}")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '$#,##0'
    cell = ws2.cell(row=row, column=4, value=f"=IF(B{row}>0,(C{row}-B{row})/B{row},0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    gross_margin_row = row
    row += 1
    
    # S&A Expenses
    ws2.cell(row=row, column=1, value="S&A Expenses").border = thin_border
    ws2.cell(row=row, column=2, value=sa_data.get('total_sa_expenses', 0)).border = thin_border
    ws2['B' + str(row)].fill = ref_fill
    ws2['B' + str(row)].number_format = '$#,##0'
    cell = ws2.cell(row=row, column=3, value=sa_data.get('total_sa_expenses', 0))
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    cell = ws2.cell(row=row, column=4, value=f"=IF(B{row}>0,(C{row}-B{row})/B{row},0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    sa_expense_row = row
    row += 1
    
    # Depreciation - inject HARD DATA from machine_spaces.xlsx
    ws2.cell(row=row, column=1, value="Depreciation (HARD)").border = thin_border
    ws2.cell(row=row, column=2, value=balance_data.get('depreciation', 0)).border = thin_border
    ws2['B' + str(row)].fill = ref_fill
    ws2['B' + str(row)].number_format = '$#,##0'
    # Use hard_data depreciation if available
    hard_depr = hard_data.get('depreciation', 0)
    if hard_depr > 0:
        cell = ws2.cell(row=row, column=3, value=hard_depr)
        cell.fill = hard_data_fill  # Gray = read-only hard data
    else:
        cell = ws2.cell(row=row, column=3, value=balance_data.get('depreciation', 0))
        cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    cell = ws2.cell(row=row, column=4, value=f"=IF(B{row}>0,(C{row}-B{row})/B{row},0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    # Add note about data source
    ws2.cell(row=row, column=5, value="← Auto-calc from machine_spaces.xlsx").font = Font(italic=True, color="666666")
    depreciation_row = row
    row += 1
    
    # Interest Expense
    ws2.cell(row=row, column=1, value="Interest Expense").border = thin_border
    ws2.cell(row=row, column=2, value=0).border = thin_border
    ws2['B' + str(row)].fill = ref_fill
    ws2['B' + str(row)].number_format = '$#,##0'
    cell = ws2.cell(row=row, column=3, value=0)
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    interest_row = row
    row += 1
    
    # Net Income
    ws2.cell(row=row, column=1, value="EST. NET INCOME").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=balance_data.get('net_profit', 0)).border = thin_border
    ws2['B' + str(row)].fill = ref_fill
    ws2['B' + str(row)].font = Font(bold=True)
    ws2['B' + str(row)].number_format = '$#,##0'
    cell = ws2.cell(row=row, column=3, 
        value=f"=C{gross_margin_row}-C{sa_expense_row}-C{depreciation_row}-C{interest_row}")
    cell.border = thin_border
    cell.fill = output_fill
    cell.font = Font(bold=True)
    cell.number_format = '$#,##0'
    cell = ws2.cell(row=row, column=4, value=f"=IF(B{row}>0,(C{row}-B{row})/B{row},0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    net_income_row = row
    row += 2
    
    # Accuracy Check
    ws2.cell(row=row, column=1, value="ACCURACY CHECK").font = section_font
    row += 1
    
    ws2.cell(row=row, column=1, value="Projected Net Margin").border = thin_border
    cell = ws2.cell(row=row, column=2, value=f"=IF(C{revenue_row}>0,C{net_income_row}/C{revenue_row},0)")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    proj_margin_row = row
    row += 1
    
    ws2.cell(row=row, column=1, value="Profit Realism Flag").border = thin_border
    cell = ws2.cell(row=row, column=2, 
        value=f'=IF(B{proj_margin_row}>$B$6+0.05,"WARNING: Unrealistic profit jump!","Projection OK")')
    cell.border = thin_border
    cell.font = Font(bold=True)
    
    # Add conditional formatting for net income (Text Color instead of Fill)
    ws2.conditional_formatting.add(
        f'C{net_income_row}',
        FormulaRule(formula=[f'C{net_income_row}<0'], font=Font(color="C00000", bold=True))
    )
    ws2.conditional_formatting.add(
        f'C{net_income_row}',
        FormulaRule(formula=[f'C{net_income_row}>0'], font=Font(color="00B050", bold=True))
    )
    
    # Variance % Formatting (Yellow if > 10% deviation)
    # Variance is in Column 4 (D)
    yellow_fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")
    ws2.conditional_formatting.add(
        f'D12:D{net_income_row}',
        FormulaRule(formula=[f'ABS(D12)>0.1'], fill=yellow_fill)
    )
    
    # Column widths
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 14
    
    # =========================================================================
    # TAB 3: BALANCE_SHEET_HEALTH
    # =========================================================================
    ws3 = wb.create_sheet("BALANCE_SHEET_HEALTH")
    
    ws3['A1'] = "BALANCE SHEET HEALTH - Solvency & Debt Control"
    ws3['A1'].font = title_font
    
    ws3['A3'] = "CURRENT POSITION (From results_and_balance_statements)"
    ws3['A3'].font = section_font
    
    ws3.cell(row=5, column=1, value="Total Assets").border = thin_border
    cell = ws3.cell(row=5, column=2, value=balance_data.get('total_assets', 0))
    cell.border = thin_border
    cell.fill = ref_fill
    cell.number_format = '$#,##0'
    
    ws3.cell(row=6, column=1, value="Total Liabilities").border = thin_border
    cell = ws3.cell(row=6, column=2, value=balance_data.get('total_liabilities', 0))
    cell.border = thin_border
    cell.fill = ref_fill
    cell.number_format = '$#,##0'
    
    ws3.cell(row=7, column=1, value="Total Equity").border = thin_border
    cell = ws3.cell(row=7, column=2, value=balance_data.get('equity', 0))
    cell.border = thin_border
    cell.fill = ref_fill
    cell.number_format = '$#,##0'
    
    ws3.cell(row=8, column=1, value="Retained Earnings").border = thin_border
    cell = ws3.cell(row=8, column=2, value=balance_data.get('retained_earnings', 0))
    cell.border = thin_border
    cell.fill = ref_fill
    cell.number_format = '$#,##0'
    
    row = 11
    ws3.cell(row=row, column=1, value="DEBT ANALYSIS").font = section_font
    row += 2
    
    ws3.cell(row=row, column=1, value="Current Debt Ratio").border = thin_border
    cell = ws3.cell(row=row, column=2, value="=B6/B5")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '0.0%'
    current_debt_row = row
    row += 2
    
    ws3.cell(row=row, column=1, value="Projected New Credit Lines").border = thin_border
    cell = ws3.cell(row=row, column=2, value=0)
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    new_credit_row = row
    row += 1
    
    ws3.cell(row=row, column=1, value="Projected New Mortgages").border = thin_border
    cell = ws3.cell(row=row, column=2, value=0)
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    new_mortgage_row = row
    row += 1
    
    ws3.cell(row=row, column=1, value="Total New Debt").border = thin_border
    cell = ws3.cell(row=row, column=2, value=f"=B{new_credit_row}+B{new_mortgage_row}")
    cell.border = thin_border
    cell.fill = calc_fill
    cell.number_format = '$#,##0'
    total_new_debt_row = row
    row += 2
    
    ws3.cell(row=row, column=1, value="Est. Post-Decision Debt Ratio").border = thin_border
    cell = ws3.cell(row=row, column=2, value=f"=(B6+B{total_new_debt_row})/B5")
    cell.border = thin_border
    cell.fill = output_fill
    cell.font = Font(bold=True)
    cell.number_format = '0.0%'
    post_debt_row = row
    row += 2
    
    # Warning Flags
    ws3.cell(row=row, column=1, value="WARNING FLAGS").font = section_font
    row += 1
    
    ws3.cell(row=row, column=1, value="Debt Level Check").border = thin_border
    cell = ws3.cell(row=row, column=2, 
        value=f'=IF(B{post_debt_row}>0.6,"CRITICAL: Debt too high. Credit Rating Risk.","Health OK")')
    cell.border = thin_border
    cell.font = Font(bold=True)
    row += 1
    
    cell = ws3.cell(row=row, column=2, 
        value='=IF(B8<0,"CRITICAL: Equity Erosion. Retained earnings negative.","Equity OK")')
    cell.border = thin_border
    cell.font = Font(bold=True)
    
    # ---------------------------------------------------------
    # CHART: Solvency Gauge (Bar Chart with Limit Line)
    # ---------------------------------------------------------
    
    # Helper Data for Chart (Hidden Columns H-J)
    ws3['H12'] = "Metric"
    ws3['I12'] = "Ratio"
    ws3['J12'] = "Limit"
    
    ws3['H13'] = "Current"
    ws3['I13'] = f"=B{current_debt_row}"
    ws3['J13'] = 0.6
    
    ws3['H14'] = "Post-Decision"
    ws3['I14'] = f"=B{post_debt_row}"
    ws3['J14'] = 0.6
    
    # Chart
    c_bar = BarChart()
    c_bar.title = "Solvency Gauge (Debt Ratio)"
    c_bar.style = 10
    c_bar.height = 10
    c_bar.width = 10
    
    data_bar = Reference(ws3, min_col=9, min_row=12, max_row=14) # Includes header
    c_bar.add_data(data_bar, titles_from_data=True)
    c_bar.set_categories(Reference(ws3, min_col=8, min_row=13, max_row=14))
    
    c_line = LineChart()
    data_line = Reference(ws3, min_col=10, min_row=12, max_row=14) # Includes header
    c_line.add_data(data_line, titles_from_data=True)
    c_line.y_axis.axId = 200 # Separate axis? No, same axis for comparison.
    # Actually if we want same axis we don't set axId or we set it to same.
    # By default openpyxl combo puts them on same axis if not specified otherwise usually.
    
    # Style line
    s_line = c_line.series[0]
    s_line.graphicalProperties.line.solidFill = "FF0000"
    s_line.graphicalProperties.line.width = 20000
    
    c_bar += c_line
    
    ws3.add_chart(c_bar, "D12")
    
    # Column widths
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 50
    
    # =========================================================================
    # TAB 4: DEBT_MANAGER
    # =========================================================================
    ws4 = wb.create_sheet("DEBT_MANAGER")
    
    ws4['A1'] = "DEBT MANAGER - Mortgage Calculator"
    ws4['A1'].font = title_font
    
    ws4['A3'] = "MORTGAGE BLOCK"
    ws4['A3'].font = section_font
    
    mortgage_headers = ['Loan #', 'Amount', 'Interest Rate', 'Payment Period 1', 'Payment Period 2', 'Total Payments']
    for col, h in enumerate(mortgage_headers, start=1):
        cell = ws4.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Pre-fill 3 loan rows
    for loan in range(1, 4):
        row = 5 + loan
        ws4.cell(row=row, column=1, value=f"Loan {loan}").border = thin_border
        
        cell = ws4.cell(row=row, column=2, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        cell = ws4.cell(row=row, column=3, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '0.0%'
        
        cell = ws4.cell(row=row, column=4, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        cell = ws4.cell(row=row, column=5, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        cell = ws4.cell(row=row, column=6, value=f"=D{row}+E{row}")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
    
    # Totals
    row = 10
    ws4.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws4.cell(row=row, column=2, value="=SUM(B6:B8)").fill = calc_fill
    ws4.cell(row=row, column=6, value="=SUM(F6:F8)").fill = calc_fill
    
    # === DIVIDEND CONTROL SECTION ===
    row = 13
    ws4.cell(row=row, column=1, value="DIVIDEND CONTROL").font = section_font
    row += 1
    
    # Max Dividend Capacity (HARD from retained earnings)
    ws4.cell(row=row, column=1, value="Max Dividend Capacity (HARD)").border = thin_border
    hard_retained = hard_data.get('retained_earnings', 0)
    cell = ws4.cell(row=row, column=2, value=hard_retained)
    cell.border = thin_border
    cell.fill = hard_data_fill  # Gray = read-only
    cell.number_format = '$#,##0'
    ws4.cell(row=row, column=3, value="← Retained Earnings from balance sheet").font = Font(italic=True, color="666666")
    max_div_row = row
    row += 1
    
    # User Dividend Input
    ws4.cell(row=row, column=1, value="User Dividend Input").border = thin_border
    cell = ws4.cell(row=row, column=2, value=0)
    cell.border = thin_border
    cell.fill = input_fill
    cell.number_format = '$#,##0'
    user_div_row = row
    row += 1
    
    # Dividend Status
    ws4.cell(row=row, column=1, value="Dividend Status").border = thin_border
    cell = ws4.cell(row=row, column=2, value=f'=IF(B{user_div_row}>B{max_div_row},"❌ ILLEGAL: Exceeds Retained Earnings","✓ LEGAL")')
    cell.border = thin_border
    row += 1
    
    # Conditional format: Turn dividend input RED if > max capacity
    red_rule = FormulaRule(
        formula=[f'$B${user_div_row}>$B${max_div_row}'],
        fill=red_fill
    )
    ws4.conditional_formatting.add(f'B{user_div_row}', red_rule)
    
    # Column widths
    for col in range(1, 7):
        ws4.column_dimensions[get_column_letter(col)].width = 18
    
    # =========================================================================
    # TAB 5: UPLOAD_READY_FINANCE
    # =========================================================================
    ws5 = wb.create_sheet("UPLOAD_READY_FINANCE")
    
    ws5['A1'] = "FINANCE DECISIONS - ExSim Upload Format"
    ws5['A1'].font = title_font
    ws5['A2'] = "Copy these values to ExSim Finance upload"
    ws5['A2'].font = Font(italic=True, color="666666")
    
    # Credit Lines
    ws5['A4'] = "Credit Lines"
    ws5['A4'].font = section_font
    
    credit_headers = ['FN1', 'FN2', 'FN3', 'FN4', 'FN5', 'FN6', 'FN7', 'FN8']
    for col, h in enumerate(credit_headers, start=2):
        cell = ws5.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    ws5.cell(row=6, column=1, value="Amount").border = thin_border
    for fn in FORTNIGHTS:
        cell = ws5.cell(row=6, column=1+fn, value=f"=LIQUIDITY_MONITOR!{get_column_letter(1+fn)}{credit_row}")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
    
    # Investments
    row = 9
    ws5.cell(row=row, column=1, value="Investments").font = section_font
    row += 1
    
    for col, h in enumerate(credit_headers, start=2):
        cell = ws5.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    ws5.cell(row=row, column=1, value="Amount").border = thin_border
    for fn in FORTNIGHTS:
        cell = ws5.cell(row=row, column=1+fn, value=f"=LIQUIDITY_MONITOR!{get_column_letter(1+fn)}{invest_row}")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
    
    # Mortgages
    row += 3
    ws5.cell(row=row, column=1, value="Mortgages").font = section_font
    row += 1
    
    mortgage_upload_headers = ['Loan', 'Amount', 'Payment 1', 'Payment 2']
    for col, h in enumerate(mortgage_upload_headers, start=1):
        cell = ws5.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    for loan in range(1, 4):
        ws5.cell(row=row, column=1, value=f"Loan {loan}").border = thin_border
        ws5.cell(row=row, column=2, value=f"=DEBT_MANAGER!B{5+loan}").border = thin_border
        ws5.cell(row=row, column=3, value=f"=DEBT_MANAGER!D{5+loan}").border = thin_border
        ws5.cell(row=row, column=4, value=f"=DEBT_MANAGER!E{5+loan}").border = thin_border
        row += 1
    
    # Dividends
    row += 2
    ws5.cell(row=row, column=1, value="Dividends").font = section_font
    row += 1
    
    # Link to per-fortnight dividends row
    for fn in FORTNIGHTS:
        cell = ws5.cell(row=row, column=1+fn, value=f"=LIQUIDITY_MONITOR!{get_column_letter(1+fn)}{dividend_row}")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
    
    # Column widths
    for col in range(1, 10):
        ws5.column_dimensions[get_column_letter(col)].width = 14
    
    # =========================================================================
    # TAB 6: CROSS_REFERENCE (Upstream Dashboard KPIs)
    # =========================================================================
    ws6 = wb.create_sheet("CROSS_REFERENCE")
    
    ws6['A1'] = "CROSS-REFERENCE SUMMARY - Upstream Dashboard KPIs"
    ws6['A1'].font = title_font
    ws6['A2'] = "Key metrics from other dashboards affecting cash flow. Data from shared_outputs.json."
    ws6['A2'].font = Font(italic=True, color="666666")
    
    # Load shared outputs
    try:
        shared_data = import_dashboard_data('CMO') if import_dashboard_data else {}
        cmo_data_shared = shared_data or {}
        prod_data_shared = (import_dashboard_data('Production') or {}) if import_dashboard_data else {}
        purch_data_shared = (import_dashboard_data('Purchasing') or {}) if import_dashboard_data else {}
        clo_data_shared = (import_dashboard_data('CLO') or {}) if import_dashboard_data else {}
        cpo_data_shared = (import_dashboard_data('CPO') or {}) if import_dashboard_data else {}
        esg_data_shared = (import_dashboard_data('ESG') or {}) if import_dashboard_data else {}
    except:
        cmo_data_shared = prod_data_shared = purch_data_shared = clo_data_shared = cpo_data_shared = esg_data_shared = {}
    
    row = 4
    
    # CMO Section (Revenue & Marketing)
    ws6.cell(row=row, column=1, value="CMO (Revenue & Marketing)").font = section_font
    ws6.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws6.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    rev = cmo_data_shared.get('est_revenue', 0)
    
    cmo_metrics = [
        ("Projected Revenue", rev, "$#,##0"),
        ("Marketing Spend", cmo_data_shared.get('marketing_spend', 0) if cmo_data_shared else 0, "$#,##0"),
        ("Innovation Costs", cmo_data_shared.get('innovation_costs', 0) if cmo_data_shared else 0, "$#,##0"),
    ]
    for label, value, fmt in cmo_metrics:
        ws6.cell(row=row, column=1, value=label).border = thin_border
        cell = ws6.cell(row=row, column=2, value=value)
        cell.border = thin_border
        cell.fill = ref_fill
        cell.number_format = fmt
        row += 1
    
    row += 1
    
    # Production Section
    ws6.cell(row=row, column=1, value="PRODUCTION (Output)").font = section_font
    ws6.cell(row=row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws6.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    prod_plan = prod_data_shared.get('production_plan', {}) if prod_data_shared else {}
    
    # Handle dict values (new format) or int values (legacy)
    total_production = 0
    if prod_plan:
        for v in prod_plan.values():
            if isinstance(v, dict):
                total_production += v.get('Target', 0)
            else:
                total_production += v

    unit_costs = prod_data_shared.get('unit_costs', {}) if prod_data_shared else {}
    avg_cost = sum(unit_costs.values()) / len(unit_costs) if unit_costs else 0
    
    # Handle capacity_utilization dict
    cap_util = prod_data_shared.get('capacity_utilization', 0) if prod_data_shared else 0
    if isinstance(cap_util, dict):
        cap_util_val = cap_util.get('mean', 0)
    else:
        cap_util_val = cap_util

    prod_metrics = [
        ("Total Production", total_production, "#,##0"),
        ("Avg Unit Cost", avg_cost, "$#,##0.00"),
        ("Utilization", cap_util_val, "0.0%"),
    ]
    for label, value, fmt in prod_metrics:
        ws6.cell(row=row, column=1, value=label).border = thin_border
        cell = ws6.cell(row=row, column=2, value=value)
        cell.border = thin_border
        cell.fill = ref_fill
        cell.number_format = fmt
        row += 1
    
    row += 1
    
    # Cost Aggregation (COGS Drivers)
    ws6.cell(row=row, column=1, value="COST DRIVERS (Variable)").font = section_font
    ws6.cell(row=row, column=1).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws6.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    labor_cost = cpo_data_shared.get('payroll_forecast', 0) # Using payroll forecast as proxy for total labor cost
    material_cost = purch_data_shared.get('supplier_spend', 0)
    logistics_cost = clo_data_shared.get('logistics_costs', 0)
    
    cogs_metrics = [
        ("Labor (CPO)", labor_cost, "$#,##0"),
        ("Materials (Purchasing)", material_cost, "$#,##0"),
        ("Logistics (CLO)", logistics_cost, "$#,##0"),
        ("TOTAL VARIABLE EST.", labor_cost + material_cost + logistics_cost, "$#,##0")
    ]
    
    for label, value, fmt in cogs_metrics:
        ws6.cell(row=row, column=1, value=label).border = thin_border
        cell = ws6.cell(row=row, column=2, value=value)
        cell.border = thin_border
        cell.fill = ref_fill
        cell.number_format = fmt
        row += 1

    row += 1

    # ESG Section
    ws6.cell(row=row, column=1, value="ESG (Sustainability)").font = section_font
    ws6.cell(row=row, column=1).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    ws6.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    esg_metrics = [
        ("CO2 Emissions", esg_data_shared.get('co2_emissions', 0) if esg_data_shared else 0, "#,##0"),
        ("Tax Liability", esg_data_shared.get('tax_liability', 0) if esg_data_shared else 0, "$#,##0"),
    ]
    for label, value, fmt in esg_metrics:
        ws6.cell(row=row, column=1, value=label).border = thin_border
        cell = ws6.cell(row=row, column=2, value=value)
        cell.border = thin_border
        cell.fill = ref_fill
        cell.number_format = fmt
        row += 1

    # Formatting
    for col in ['A', 'B']:
        ws6.column_dimensions[col].width = 35

    # Save
    if output_buffer:
        wb.save(output_buffer)
        return output_buffer
    else:
        wb.save(OUTPUT_FILE)
        print(f"[SUCCESS] Created '{OUTPUT_FILE}'")
        return None


def main():
    """Main function."""
    print("ExSim Finance Dashboard Final Generator")
    print("=" * 50)
    
    print("\n[*] Loading data files...")
    from config import REPORTS_DIR, DATA_DIR
    print(f"    Primary source: {REPORTS_DIR}")
    print(f"    Fallback source: {DATA_DIR}")
    
    # Initial Cash Flow
    cash_path = get_data_path("initial_cash_flow.xlsx")
    if cash_path:
        cash_data = load_initial_cash_flow(cash_path)
        print(f"  [OK] Loaded initial cash from {cash_path.parent.name}/")
    else:
        cash_data = load_initial_cash_flow(None)
        print("  [!] Using default cash data")
    
    # Balance Statements
    balance_path = get_data_path("results_and_balance_statements.xlsx")
    if balance_path:
        balance_data = load_balance_statements(balance_path)
        print(f"  [OK] Loaded balance data")
    else:
        balance_data = load_balance_statements(None)
        print("  [!] Using default balance data")
    
    # S&A Expenses
    sa_path = get_data_path("sales_admin_expenses.xlsx")
    if sa_path:
        sa_data = load_sales_admin_expenses(sa_path)
        print(f"  [OK] Loaded S&A expenses")
    else:
        sa_data = load_sales_admin_expenses(None)
        print("  [!] Using default S&A data")
    
    # Receivables/Payables
    ar_ap_path = get_data_path("accounts_receivable_payable.xlsx")
    if ar_ap_path:
        ar_ap_data = load_receivables_payables(ar_ap_path)
        print(f"  [OK] Loaded AR/AP data")
    else:
        ar_ap_data = load_receivables_payables(None)
        print("  [!] Using default AR/AP data")
    
    # Template
    template_path = get_data_path("Finance Decisions.xlsx")
    template_data = load_finance_template(template_path)
    if template_data['exists']:
        print(f"  [OK] Loaded finance template")
    else:
        print("  [!] Using default template layout")
    
    # === NEW: HARD DATA LOADERS ===
    print("\n[*] Loading HARD DATA for injection...")
    
    # Machine Depreciation
    machine_path = get_data_path("machine_spaces.xlsx")
    if machine_path:
        hard_depreciation = load_machine_depreciation(machine_path)
        print(f"  [HARD] Depreciation: ${hard_depreciation:,.0f} (from machine_spaces.xlsx)")
    else:
        hard_depreciation = 0.0
        print("  [!] No machine_spaces.xlsx - using 0 for depreciation")
    
    # Precise Starting Cash
    if cash_path:
        hard_starting_cash = load_initial_cash_precise(cash_path)
        print(f"  [HARD] Starting Cash FN1: ${hard_starting_cash:,.0f}")
    else:
        hard_starting_cash = 0.0
    
    # Precise AR/AP Schedule
    if ar_ap_path:
        hard_schedule = load_hard_schedule_precise(ar_ap_path)
        total_rec = sum(d['receivables'] for d in hard_schedule.values())
        total_pay = sum(d['payables'] for d in hard_schedule.values())
        print(f"  [HARD] Receivables: ${total_rec:,.0f} | Payables: ${total_pay:,.0f}")
    else:
        hard_schedule = {fn: {'receivables': 0, 'payables': 0} for fn in range(1, 9)}
    
    # Retained Earnings (for dividend limit)
    if balance_path:
        hard_retained_earnings = load_retained_earnings(balance_path)
        print(f"  [HARD] Retained Earnings (Dividend Limit): ${hard_retained_earnings:,.0f}")
    else:
        hard_retained_earnings = 0.0
    
    # Bundle hard data
    hard_data = {
        'depreciation': hard_depreciation,
        'starting_cash': hard_starting_cash,
        'schedule': hard_schedule,
        'retained_earnings': hard_retained_earnings
    }
    
    print("\n[*] Generating Finance Dashboard...")
    
    create_finance_dashboard(cash_data, balance_data, sa_data, ar_ap_data, template_data, hard_data)
    
    print("\nSheets created:")
    print("  * LIQUIDITY_MONITOR (Cash Flow Engine)")
    print("  * PROFIT_CONTROL (Income Statement Projection)")
    print("  * BALANCE_SHEET_HEALTH (Solvency & Debt)")
    print("  * DEBT_MANAGER (Mortgage Calculator)")
    print("  * UPLOAD_READY_FINANCE (ExSim Format)")
    
    # Export key metrics for use by other systems
    if export_dashboard_data:
        export_dashboard_data('CFO', {
            'cash_flow_projection': cash_data,
            'debt_levels': balance_data.get('total_debt', 0),
            'liquidity_status': 'OK' if cash_data.get('ending_cash', 0) > 0 else 'LOW'
        })


if __name__ == "__main__":
    main()
