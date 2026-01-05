import sys
import os
import io
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../Purchasing Role')))

from generate_purchasing_dashboard_v2 import create_purchasing_dashboard

def seed_shared_outputs():
    try:
        from shared_outputs import SharedOutputManager
        manager = SharedOutputManager()
        
        # Mock Production Data (Production Plan)
        # Assuming aggregated plan or per-zone. 
        # Purchasing needs total part requirements or per-zone?
        # MRP Engine Row 6 in dashboard is "Target Production" (Global or Center?).
        # Let's seed per-zone and see what we can pull.
        manager.export('Production', {
            'production_plan': {
                'Center': {'Target': 1500}, 
                'West': {'Target': 800}, 
                'North': {'Target': 600}, 
                'East': {'Target': 900}, 
                'South': {'Target': 700}
            },
            'capacity_utilization': {'mean': 0.85},
            'overtime_hours': 100,
            'unit_costs': {}
        })
        print("Seeded Shared Outputs successfully (Production data).")
    except Exception as e:
        print(f"Failed to seed shared outputs: {e}")

def test_purchasing_generation_with_overrides():
    print("Testing Purchasing Dashboard Generation...")
    
    # Mock Input Data
    materials_data = {
        'parts': {
            'Part A': {'final_inventory': 500},
            'Part B': {'final_inventory': 300}
        },
        'pieces': {
            'Piece 1': {'final_inventory': 1000},
            'Piece 2': {'final_inventory': 1000}
        }
    }
    cost_data = {'ordering_cost': 5000, 'holding_cost': 2000, 'consumption_cost': 50000}
    template_data = {'df': None, 'exists': False}
    
    # Decision Overrides
    # Override order quantities
    overrides = {
        'Part A': {
            'Supplier A': {1: 1000, 2: 2000} # FN1, FN2
        }
    }
    
    output = io.BytesIO()
    create_purchasing_dashboard(
        materials_data, cost_data, template_data,
        output_buffer=output,
        decision_overrides=overrides
    )
    
    output.seek(0)
    wb = load_workbook(output)
    
    # 1. Verify MRP Engine
    ws_mrp = wb["MRP ENGINE"]
    
    # Check Target Production Import (Row 6)
    # Ideally should be a formula pointing to Cross Reference
    prod_target_fn1 = ws_mrp['B6'].value
    print(f"Target Production Link (FN1): {prod_target_fn1} (Expected: ='CROSS REFERENCE'!B11)")
    
    # Check Order Override (Part A, Supplier A)
    # Find row for Part A, Supplier A
    # Part A starts around row 11. 
    # Order Supplier A is likely row 18 (Gross, Arrivals, Proj, Deficit lines before it).
    # Let's trust my visual scan or just search?
    # Scanning implies Row 18 is Supplier A Order.
    # Col B is FN1.
    order_val = ws_mrp['B18'].value
    print(f"Order Part A/Supp A (FN1): {order_val} (Expected: 1000)")
    
    # 2. Verify Cross Reference (Production Data)
    print("\nChecking Cross Reference Tab...")
    if "CROSS REFERENCE" in wb.sheetnames:
        ws_xref = wb["CROSS REFERENCE"]
        # Basic check
        plan_val = ws_xref['B5'].value
        print(f"Production Plan Ref: {plan_val}")
    else:
        print("FAIL: Cross Reference tab missing!")

    print("Test Complete.")

if __name__ == "__main__":
    seed_shared_outputs()
    test_purchasing_generation_with_overrides()
