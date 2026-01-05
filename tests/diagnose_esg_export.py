import sys
import os
import io
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../ESG Dashboard')))

from generate_esg_dashboard import create_esg_dashboard

def seed_shared_outputs():
    try:
        from shared_outputs import SharedOutputManager
        manager = SharedOutputManager()
        
        # Mock Upstream Data
        manager.export('Production', {
            'production_plan': {'Center': {'Target': 2000}}, 
            'capacity_utilization': {'mean': 0.75}
        })
        manager.export('CLO', {
            'logistics_costs': 45000
        })
        
        print("Seeded Shared Outputs successfully.")
    except Exception as e:
        print(f"Failed to seed shared outputs: {e}")

def test_esg_generation():
    print("Testing ESG Dashboard Generation...")
    
    # Mock Input Data
    esg_data = {'emissions': 1000, 'tax_paid': 30000, 'energy_consumption': 50000}
    prod_data = {'total_production': 2000}
    
    # Overrides
    overrides = {
        'Solar PV Panels': 10,
        'Green Electricity': 0.5 # 50%
    }
    
    output = io.BytesIO()
    create_esg_dashboard(
        esg_data, prod_data,
        output_buffer=output,
        decision_overrides=overrides
    )
    
    output.seek(0)
    wb = load_workbook(output)
    
    # 1. Verify Strategy Selector (Tab 2)
    ws_strat = wb["STRATEGY SELECTOR"]
    
    # Check Solar PV (Row 13?)
    # "Solar PV Panels" is usually first key in DEFAULT_INITIATIVES.
    # Row 13, Col 2 = Quantity.
    solar_qty = ws_strat['B13'].value
    print(f"Solar PV Qty: {solar_qty} (Expected: 10)")
    
    # Check Green Electricity (Row 15?)
    # Keys: Solar, Trees, Green Elec... (Row 13 labeled name at A13)
    # Let's find row for "Green Electricity"
    ge_row = -1
    for r in range(13, 17):
        if ws_strat[f'A{r}'].value == "Green Electricity":
            ge_row = r
            break
            
    if ge_row > 0:
        ge_qty = ws_strat[f'B{ge_row}'].value
        print(f"Green Elec %: {ge_qty} (Expected: 0.5)")
    else:
        print("FAIL: Could not find Green Electricity row")
    
    # 2. Verify Cross Reference (Tab 5)
    print("\nChecking Cross Reference Tab...")
    if "CROSS REFERENCE" in wb.sheetnames:
        ws_xref = wb["CROSS REFERENCE"]
        print("PASS: Cross Reference tab exists.")
        
        # Verify Production
        # Label "Total Production"
        found = False
        for row in range(1, 20):
            val = ws_xref[f'A{row}'].value
            if val == "Total Production":
                prod_val = ws_xref[f'B{row}'].value
                print(f"XRef Production: {prod_val} (Expected: 2000)")
                found = True
        
        # Verify logistics
        for row in range(1, 20):
            val = ws_xref[f'A{row}'].value
            if val == "Logistics Costs":
                log_val = ws_xref[f'B{row}'].value
                print(f"XRef Logistics: {log_val} (Expected: 45000 or similar string)")

    else:
        print("FAIL: Cross Reference tab missing!")
        
    # 3. Verify Upload Ready (Tab 3)
    ws_up = wb["UPLOAD READY ESG"]
    # Check Header
    if ws_up['A1'].value == "UPLOAD READY DATA - DO NOT EDIT":
        print("PASS: Upload Ready tab exists.")

    print("Test Complete.")

if __name__ == "__main__":
    seed_shared_outputs()
    test_esg_generation()
