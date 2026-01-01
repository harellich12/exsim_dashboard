"""
ExSim Dashboard Generation Test Suite

Tests data loaders with in-memory fixtures and validates dashboard generation
produces correct Excel output.

Run with: pytest tests/test_dashboard_generation.py -v
"""

import pytest
import pandas as pd
import io
import sys
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook

# Add project root to path for imports
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "CFO Dashboard"))
sys.path.insert(0, str(PROJECT_ROOT / "CMO Dashboard"))


# =============================================================================
# FIXTURES - In-Memory Excel Data
# =============================================================================

@pytest.fixture
def mock_initial_cash_xlsx():
    """
    Create in-memory Excel file with initial cash flow data.
    Simulates messy row: "Final cash (at the start of the first fortnight)", "$56,421.00"
    """
    wb = Workbook()
    ws = wb.active
    
    # Header rows (simulating real format)
    ws['A1'] = "Company X"
    ws['A2'] = "Initial Cash Flow Report"
    ws['A3'] = ""
    
    # Data rows
    ws['A5'] = "Opening Cash Balance"
    ws['B5'] = "$100,000.00"
    
    ws['A6'] = "Collections from customers"
    ws['B6'] = "$50,000.00"
    
    ws['A7'] = "Payments to suppliers"
    ws['B7'] = "-$43,579.00"
    
    ws['A8'] = "Tax Payment"
    ws['B8'] = "-$25,000.00"
    
    # The key row we're testing
    ws['A9'] = "Final cash (at the start of the first fortnight)"
    ws['B9'] = "$56,421.00"
    
    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def mock_hard_schedule_xlsx():
    """
    Create in-memory Excel file with receivables/payables schedule.
    8 rows for FN 1-8 with Fortnight, Receivables, Payables columns.
    """
    wb = Workbook()
    ws = wb.active
    
    # Header
    ws['A1'] = "Fortnight"
    ws['B1'] = "Receivables"
    ws['C1'] = "Payables"
    
    # 8 fortnight rows
    schedule_data = [
        (1, 50000, 30000),
        (2, 55000, 32000),
        (3, 48000, 28000),
        (4, 52000, 31000),
        (5, 60000, 35000),
        (6, 58000, 33000),
        (7, 62000, 36000),
        (8, 65000, 38000),
    ]
    
    for i, (fn, rec, pay) in enumerate(schedule_data, start=2):
        ws.cell(row=i, column=1, value=fn)
        ws.cell(row=i, column=2, value=rec)
        ws.cell(row=i, column=3, value=pay)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def mock_sales_admin_xlsx():
    """
    Create in-memory Excel file with sales/admin expenses.
    Includes TV and Radio advertising with spots and amounts.
    """
    wb = Workbook()
    ws = wb.active
    
    # Header
    ws['A1'] = "Sales & Administrative Expenses"
    ws['A2'] = ""
    
    # Expense rows - format: Label, Details, Amount
    ws['A4'] = "TV Advertising Expenses"
    ws['B4'] = "10 spots"
    ws['C4'] = 30000  # $30,000 / 10 spots = $3,000 per spot
    
    ws['A5'] = "Radio Advertising Expenses"
    ws['B5'] = "100 spots"
    ws['C5'] = 5000  # $5,000 / 100 spots = $50 per spot
    
    ws['A6'] = "Salespeople Salaries"
    ws['B6'] = "20 people"
    ws['C6'] = 30000  # $30,000 / 20 = $1,500 per person
    
    ws['A7'] = "Salespeople Hiring"
    ws['B7'] = "5 hires"
    ws['C7'] = 5500  # $5,500 / 5 = $1,100 per hire
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@pytest.fixture
def mock_cfo_data_bundle():
    """Bundle of all CFO data dicts with known values for integration tests."""
    return {
        'cash_data': {
            'final_cash': 99999,
            'tax_payments': 5000
        },
        'balance_data': {
            'net_sales': 500000,
            'cogs': 300000,
            'gross_income': 200000,
            'net_profit': 50000,
            'total_assets': 1000000,
            'total_liabilities': 400000,
            'equity': 600000,
            'retained_earnings': 150000,
            'depreciation': 25000,
            'gross_margin_pct': 0.4,
            'net_margin_pct': 0.1
        },
        'sa_data': {
            'total_sa_expense': 75000,
            'tv_spend': 15000,
            'radio_spend': 5000,
            'salespeople_cost': 30000
        },
        'ar_ap_data': {
            'total_receivables': 80000,
            'total_payables': 40000,
            'by_fn': {fn: {'receivables': 10000, 'payables': 5000} for fn in range(1, 9)}
        },
        'template_data': None,
        'hard_data': {
            'depreciation': 25000,
            'starting_cash': 99999,
            'schedule': {fn: {'receivables': 10000 + fn*1000, 'payables': 5000 + fn*500} for fn in range(1, 9)},
            'retained_earnings': 150000
        }
    }


@pytest.fixture
def mock_cmo_intelligence():
    """Marketing intelligence data with known TV spot cost."""
    return {
        'economics': {
            'TV_Cost_Spot': 2500,
            'Radio_Cost_Spot': 50,
            'Salary_Per_Person': 1500,
            'Hiring_Cost': 1100
        },
        'pricing': {
            'Center': 150,
            'West': 145,
            'North': 148,
            'East': 152,
            'South': 147
        }
    }


# =============================================================================
# DATA LOADER UNIT TESTS
# =============================================================================

class TestDataLoaders:
    """Unit tests for data loading functions."""
    
    def test_load_initial_cash_extracts_correct_float(self, mock_initial_cash_xlsx, tmp_path):
        """
        Test that load_initial_cash_flow extracts the correct float from a messy string row.
        Expected: Parses "$56,421.00" from "Final cash (at the start of the first fortnight)"
        """
        # Import the function
        from generate_finance_dashboard_final import load_initial_cash_flow
        
        # Save fixture to temp file (function expects Path)
        test_file = tmp_path / "initial_cash_flow.xlsx"
        test_file.write_bytes(mock_initial_cash_xlsx.read())
        
        # Call the function
        result = load_initial_cash_flow(test_file)
        
        # Assert correct extraction
        assert 'final_cash' in result
        assert result['final_cash'] == 56421.0, f"Expected 56421.0, got {result['final_cash']}"
    
    def test_load_hard_schedule_returns_8_rows(self, mock_hard_schedule_xlsx, tmp_path):
        """
        Test that load_hard_schedule_precise returns a dict with exactly 8 entries (FN 1-8).
        Each entry should have 'receivables' and 'payables' keys.
        """
        from generate_finance_dashboard_final import load_hard_schedule_precise
        
        # Save fixture to temp file
        test_file = tmp_path / "accounts_receivable_payable.xlsx"
        test_file.write_bytes(mock_hard_schedule_xlsx.read())
        
        # Call the function
        result = load_hard_schedule_precise(test_file)
        
        # Assert 8 entries
        assert len(result) == 8, f"Expected 8 entries, got {len(result)}"
        
        # Assert all keys 1-8 present
        for fn in range(1, 9):
            assert fn in result, f"Missing fortnight {fn}"
            assert 'receivables' in result[fn], f"Missing 'receivables' for FN{fn}"
            assert 'payables' in result[fn], f"Missing 'payables' for FN{fn}"
    
    def test_extract_marketing_economics_divides_correctly(self, mock_sales_admin_xlsx, tmp_path):
        """
        Test that load_marketing_intelligence correctly divides Amount by Count.
        Expected: TV $30,000 / 10 spots = $3,000 per spot
        """
        from generate_cmo_dashboard_complete import load_marketing_intelligence
        
        # Save fixture to temp file
        test_file = tmp_path / "sales_admin_expenses.xlsx"
        test_file.write_bytes(mock_sales_admin_xlsx.read())
        
        # Call with sales file (market file can be None)
        result = load_marketing_intelligence(test_file, None)
        
        # Assert correct division
        assert 'economics' in result
        assert result['economics']['TV_Cost_Spot'] == 3000.0, \
            f"Expected 3000.0 (30000/10), got {result['economics']['TV_Cost_Spot']}"
        assert result['economics']['Radio_Cost_Spot'] == 50.0, \
            f"Expected 50.0 (5000/100), got {result['economics']['Radio_Cost_Spot']}"
        assert result['economics']['Salary_Per_Person'] == 1500.0, \
            f"Expected 1500.0 (30000/20), got {result['economics']['Salary_Per_Person']}"


# =============================================================================
# INTEGRATION TESTS - File Generation
# =============================================================================

class TestFinanceDashboardGeneration:
    """Integration tests for CFO Finance Dashboard generation."""
    
    def test_finance_dashboard_cell_b5_matches_input(self, mock_cfo_data_bundle, tmp_path):
        """
        Test that Cell B5 in LIQUIDITY_MONITOR equals the injected cash value.
        """
        from generate_finance_dashboard_final import create_finance_dashboard
        import os
        
        # Change to temp directory so output goes there
        original_cwd = os.getcwd()
        os.chdir(tmp_path)
        
        try:
            # Call the dashboard generator
            create_finance_dashboard(
                cash_data=mock_cfo_data_bundle['cash_data'],
                balance_data=mock_cfo_data_bundle['balance_data'],
                sa_data=mock_cfo_data_bundle['sa_data'],
                ar_ap_data=mock_cfo_data_bundle['ar_ap_data'],
                template_data=mock_cfo_data_bundle['template_data'],
                hard_data=mock_cfo_data_bundle['hard_data']
            )
            
            # Load the generated file
            output_file = tmp_path / "Finance_Dashboard_Final.xlsx"
            assert output_file.exists(), "Dashboard file was not created"
            
            wb = load_workbook(output_file, data_only=True)
            ws = wb["LIQUIDITY_MONITOR"]
            
            # Cell B5 should contain the final_cash value
            actual_value = ws['B5'].value
            expected_value = mock_cfo_data_bundle['cash_data']['final_cash']
            
            assert actual_value == expected_value, \
                f"Cell B5 mismatch: expected {expected_value}, got {actual_value}"
        
        finally:
            os.chdir(original_cwd)
    
    def test_finance_dashboard_starting_cash_uses_hard_data(self, mock_cfo_data_bundle, tmp_path):
        """
        Test that Row 9 (Starting Cash for FN1) uses the hard_data value.
        """
        from generate_finance_dashboard_final import create_finance_dashboard
        import os
        
        original_cwd = os.getcwd()
        os.chdir(tmp_path)
        
        try:
            create_finance_dashboard(
                cash_data=mock_cfo_data_bundle['cash_data'],
                balance_data=mock_cfo_data_bundle['balance_data'],
                sa_data=mock_cfo_data_bundle['sa_data'],
                ar_ap_data=mock_cfo_data_bundle['ar_ap_data'],
                template_data=mock_cfo_data_bundle['template_data'],
                hard_data=mock_cfo_data_bundle['hard_data']
            )
            
            output_file = tmp_path / "Finance_Dashboard_Final.xlsx"
            wb = load_workbook(output_file, data_only=True)
            ws = wb["LIQUIDITY_MONITOR"]
            
            # Row 9 Col B should contain starting_cash from hard_data
            actual_value = ws['B9'].value
            expected_value = mock_cfo_data_bundle['hard_data']['starting_cash']
            
            assert actual_value == expected_value, \
                f"Starting Cash mismatch: expected {expected_value}, got {actual_value}"
        
        finally:
            os.chdir(original_cwd)


# =============================================================================
# CMO LOGIC VERIFICATION
# =============================================================================

class TestCMODashboardLogic:
    """Logic verification tests for CMO Dashboard."""
    
    def test_cmo_tv_spot_cost_appears_in_output(self, mock_cmo_intelligence, tmp_path):
        """
        Test that the TV Spot Cost passed to the generator appears in the output.
        """
        from generate_cmo_dashboard_complete import create_complete_dashboard
        import os
        
        # Create minimal required data structures
        market_data = {
            'by_segment': {
                'High': {zone: {'my_market_share': 10, 'my_awareness': 50, 'comp_avg_awareness': 40,
                               'my_price': 150, 'comp_avg_price': 145, 'my_attractiveness': 60}
                        for zone in ['Center', 'West', 'North', 'East', 'South']},
                'Low': {zone: {'my_market_share': 15, 'my_awareness': 45, 'comp_avg_awareness': 42,
                              'my_price': 120, 'comp_avg_price': 118, 'my_attractiveness': 55}
                       for zone in ['Center', 'West', 'North', 'East', 'South']}
            },
            'zones': {zone: {'my_price': 150, 'comp_avg_price': 145, 'my_awareness': 50,
                            'my_attractiveness': 60, 'my_market_share': 12}
                     for zone in ['Center', 'West', 'North', 'East', 'South']}
        }
        
        innovation_features = ["STAINLESS MATERIAL", "ENERGY EFFICIENCY"]
        
        marketing_template = {
            'df': None,
            'tv_budget': 0,
            'brand_focus': 0,
            'radio_budgets': {zone: 0 for zone in ['Center', 'West', 'North', 'East', 'South']},
            'demand': {zone: 1000 for zone in ['Center', 'West', 'North', 'East', 'South']},
            'prices': {zone: 150 for zone in ['Center', 'West', 'North', 'East', 'South']},
            'payment_terms': {zone: 'A' for zone in ['Center', 'West', 'North', 'East', 'South']},
            'salespeople': {zone: 10 for zone in ['Center', 'West', 'North', 'East', 'South']}
        }
        
        sales_data = {
            'by_zone': {zone: {'units': 500, 'price': 150} for zone in ['Center', 'West', 'North', 'East', 'South']},
            'totals': {'units': 2500, 'tv_spend': 15000, 'radio_spend': 5000, 'salespeople_cost': 30000}
        }
        
        inventory_data = {
            'final_inventory': 1000,
            'is_stockout': False,
            'by_zone': {zone: {'final': 200, 'capacity': 500} for zone in ['Center', 'West', 'North', 'East', 'South']}
        }
        
        original_cwd = os.getcwd()
        os.chdir(tmp_path)
        
        try:
            create_complete_dashboard(
                market_data=market_data,
                innovation_features=innovation_features,
                marketing_template=marketing_template,
                sales_data=sales_data,
                inventory_data=inventory_data,
                marketing_intelligence=mock_cmo_intelligence
            )
            
            output_file = tmp_path / "CMO_Dashboard_Complete.xlsx"
            assert output_file.exists(), "CMO Dashboard file was not created"
            
            wb = load_workbook(output_file, data_only=True)
            ws = wb["STRATEGY_COCKPIT"]
            
            # Find the TV Cost cell - it should be in the Unit Economics section
            # Looking for the value 2500 (from mock_cmo_intelligence)
            found_tv_cost = False
            expected_tv_cost = mock_cmo_intelligence['economics']['TV_Cost_Spot']
            
            # Search the sheet for the TV cost value
            for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=5):
                for cell in row:
                    if cell.value == expected_tv_cost:
                        found_tv_cost = True
                        break
                    # Also check if it's in a "TV" labeled row
                    if 'TV' in str(cell.value).upper() if cell.value else False:
                        # Check adjacent cells for the cost
                        for adj_col in range(cell.column, cell.column + 3):
                            adj_cell = ws.cell(row=cell.row, column=adj_col)
                            if adj_cell.value == expected_tv_cost:
                                found_tv_cost = True
                                break
            
            assert found_tv_cost, \
                f"TV Spot Cost {expected_tv_cost} not found in STRATEGY_COCKPIT sheet"
        
        finally:
            os.chdir(original_cwd)


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
