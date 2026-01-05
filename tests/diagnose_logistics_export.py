import sys
import os
import io
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../CLO Dashboard')))

from generate_logistics_dashboard import create_logistics_dashboard

def seed_shared_outputs():
    try:
        from shared_outputs import SharedOutputManager
        manager = SharedOutputManager()
        
        # Mock Production Data (Production Plan)
        manager.export('Production', {
            'production_plan': {
                'Center': {'Target': 1500}, 
                'West': {'Target': 800}, 
                'North': {'Target': 600}, 
                'East': {'Target': 900}, 
                'South': {'Target': 700}
            }
        })
        
        # Mock CMO Data (Demand Forecast)
        manager.export('CMO', {
            'demand_forecast': {
                'Center': 1200, 'West': 800, 'North': 600, 'East': 900, 'South': 700
            }
        })
        print("Seeded Shared Outputs successfully.")
    except Exception as e:
        print(f"Failed to seed shared outputs: {e}")

def test_logistics_generation():
    print("Testing Logistics Dashboard Generation...")
    
    # Mock Input Data
    inventory_data = {z: {'capacity': 2000, 'inventory': 500} for z in ['Center', 'West', 'North', 'East', 'South']}
    template_data = {'df': None, 'exists': False}
    cost_data = {'total_shipping_cost': 5000}
    
    output = io.BytesIO()
    create_logistics_dashboard(
        inventory_data, template_data, cost_data,
        intelligence_data=None,
        output_buffer=output
    )
    
    output.seek(0)
    wb = load_workbook(output)
    
    # 1. Verify Inventory Tetris Links
    ws_tetris = wb["INVENTORY TETRIS"]
    
    # Check Center Zone Production (FN1)
    # With fixes, this should be a link to Cross Reference Row 6 (Center)
    prod_val = ws_tetris['B8'].value
    print(f"Center Production (FN1): {prod_val} (Expected: ='CROSS REFERENCE'!B6)")
    
    sales_val = ws_tetris['C8'].value
    print(f"Center Sales (FN1): {sales_val} (Expected: ='CROSS REFERENCE'!C6)")
    
    # 2. Verify Cross Reference
    print("\nChecking Cross Reference Tab...")
    if "CROSS REFERENCE" in wb.sheetnames:
        ws_xref = wb["CROSS REFERENCE"]
        print("PASS: Cross Reference tab exists.")
        
        # Verify Center Zone Data (Row 6)
        xref_prod = ws_xref['B6'].value
        xref_sales = ws_xref['C6'].value
        print(f"XRef Center Production: {xref_prod} (Expected: 1500)")
        print(f"XRef Center Demand: {xref_sales} (Expected: 1200)")
    else:
        print("FAIL: Cross Reference tab missing!")

    print("Test Complete.")

if __name__ == "__main__":
    seed_shared_outputs()
    test_logistics_generation()
