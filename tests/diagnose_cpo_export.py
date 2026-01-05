import sys
import os
import io
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../CPO Dashboard')))

from generate_cpo_dashboard import create_cpo_dashboard

def seed_shared_outputs():
    try:
        from shared_outputs import SharedOutputManager
        manager = SharedOutputManager()
        
        # Mock Production Data (Production Plan)
        # Production Plan format: {Zone: {Target: X}}
        manager.export('Production', {
            'production_plan': {
                'Center': {'Target': 1500}, 
            },
            'overtime_hours': 120
        })
        
        # Mock CFO Data (Budget)
        manager.export('CFO', {
            'liquidity_status': 'Healthy', 
            'cash_on_hand': 500000
        })
        print("Seeded Shared Outputs successfully.")
    except Exception as e:
        print(f"Failed to seed shared outputs: {e}")

def test_cpo_generation():
    print("Testing CPO Dashboard Generation...")
    
    # Mock Input Data
    workers_data = {
        'Center': {'workers': 100, 'absenteeism': 0.02},
        'West': {'workers': 50, 'absenteeism': 0.02},
        'North': {'workers': 40, 'absenteeism': 0.02},
        'East': {'workers': 60, 'absenteeism': 0.02},
        'South': {'workers': 70, 'absenteeism': 0.02}
    }
    sales_data = {'headcount': 50, 'avg_salary': 800, 'total_salary': 40000, 'hiring_cost': 1200}
    labor_data = {'total_labor': 100000}
    absenteeism_rate = 0.03
    
    # Overrides
    overrides = {
        'workforce': {
            'Center': {'required': 120, 'turnover': 0.05} # Want to hire 20+
        },
        'salary': {
            'Center': 750 # Raise salary
        }
    }
    
    output = io.BytesIO()
    create_cpo_dashboard(
        workers_data, sales_data, labor_data, absenteeism_rate,
        output_buffer=output,
        decision_overrides=overrides
    )
    
    output.seek(0)
    wb = load_workbook(output)
    
    # 1. Verify Workforce Planning (Tab 1)
    ws_plan = wb["WORKFORCE PLANNING"]
    
    # Check Center Zone (Row 10)
    # Col 3 = Required (Override = 120)
    req_val = ws_plan['C10'].value
    print(f"Center Required Workers: {req_val} (Expected: 120)")
    
    # Col 4 = Turnover (Override = 0.05)
    turn_val = ws_plan['D10'].value
    print(f"Center Turnover: {turn_val} (Expected: 0.05)")
    
    # 2. Verify Compensation Strategy (Tab 2)
    ws_comp = wb["COMPENSATION STRATEGY"]
    
    # Check Center Salary (Row 11) -> Center is first?
    # Index 11 is header? No, Row 10 header. Row 11 = Center.
    # Col 4 = Proposed (Override = 750)
    salary_val = ws_comp['D11'].value
    print(f"Center Proposed Salary: {salary_val} (Expected: 750)")
    
    # 3. Verify Cross Reference (Tab 4)
    # Check if Production Target is visible
    print("\nChecking Cross Reference Tab...")
    if "CROSS REFERENCE" in wb.sheetnames:
        ws_xref = wb["CROSS REFERENCE"]
        print("PASS: Cross Reference tab exists.")
        
        # Row 5: "Total Production Target" -> Value in Col 2 (from logic: sum(targets))
        # Center=1500. Total=1500.
        # Let's find "Total Production Target" label row
        found = False
        for row in range(1, 20):
            val = ws_xref[f'A{row}'].value
            if val == "Total Production Target":
                target_val = ws_xref[f'B{row}'].value
                print(f"XRef Production Target: {target_val} (Expected: 1500)")
                found = True
        if not found:
            print("FAIL: Could not find 'Total Production Target' row.")
            
        # Check CFO Status
        # Label "Liquidity Status"
        for row in range(1, 20):
            val = ws_xref[f'A{row}'].value
            if val == "Liquidity Status":
                liq = ws_xref[f'B{row}'].value
                print(f"XRef Liquidity: {liq} (Expected: Healthy)")

    else:
        print("FAIL: Cross Reference tab missing!")

    print("Test Complete.")

if __name__ == "__main__":
    seed_shared_outputs()
    test_cpo_generation()
