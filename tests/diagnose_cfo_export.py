import sys
import os
import io
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../CFO Dashboard')))

from generate_finance_dashboard_final import create_finance_dashboard

def seed_shared_outputs():
    try:
        from shared_outputs import SharedOutputManager
        manager = SharedOutputManager()
        
        # Mock Upstream Data
        manager.export('CMO', {'est_revenue': 500000, 'marketing_spend': 50000})
        manager.export('Production', {
            'production_plan': {'Center': {'Target': 1000}, 'West': {'Target': 500}}, 
            'unit_costs': {'Center': 100},
            'capacity_utilization': 0.85
        })
        manager.export('Purchasing', {'supplier_spend': 120000})
        manager.export('CLO', {'logistics_costs': 30000})
        manager.export('CPO', {'payroll_forecast': 80000})
        manager.export('ESG', {'co2_emissions': 500, 'tax_liability': 5000})
        
        print("Seeded Shared Outputs successfully.")
    except Exception as e:
        print(f"Failed to seed shared outputs: {e}")

def test_cfo_generation():
    print("Testing CFO Dashboard Generation...")
    
    # Mock Input Data
    cash_data = {'final_cash': 200000, 'tax_payments': 10000}
    balance_data = {
        'net_sales': 1000000, 'cogs': 600000, 'gross_income': 400000,
        'net_profit': 150000, 'total_assets': 2000000, 'total_liabilities': 800000,
        'equity': 1200000, 'retained_earnings': 300000, 'depreciation': 20000
    }
    sa_data = {'total_sa_expenses': 50000}
    ar_ap_data = {
        'receivables': [10000]*8,
        'payables': [5000]*8
    }
    template_data = {'df': None, 'exists': False}
    hard_data = {
        'depreciation': 25000,
        'starting_cash': 190000,
        'schedule': {fn: {'receivables': 12000, 'payables': 6000} for fn in range(1, 9)},
        'retained_earnings': 300000
    }
    
    output = io.BytesIO()
    create_finance_dashboard(
        cash_data, balance_data, sa_data, ar_ap_data, template_data, hard_data,
        output_buffer=output
    )
    
    output.seek(0)
    wb = load_workbook(output)
    
    # 1. Verify Cross Reference (Tab 6)
    print("\nChecking Cross Reference Tab...")
    if "CROSS REFERENCE" in wb.sheetnames:
        ws_xref = wb["CROSS REFERENCE"]
        print("PASS: Cross Reference tab exists.")
        
        # Helper to find value by label
        def get_xref_val(label):
            for row in range(1, 50):
                val = ws_xref[f'A{row}'].value
                if val == label:
                    return ws_xref[f'B{row}'].value
            return None
            
        # Verify CMO
        rev = get_xref_val("Projected Revenue")
        print(f"XRef Revenue: {rev} (Expected: 500000)")
        
        # Verify Production
        prod = get_xref_val("Total Production")
        print(f"XRef Production: {prod} (Expected: 1500)") # 1000+500
        
        # Verify Costs
        labor = get_xref_val("Labor (CPO)")
        print(f"XRef Labor: {labor} (Expected: 80000)")
        
        mat = get_xref_val("Materials (Purchasing)")
        print(f"XRef Material: {mat} (Expected: 120000)")
        
        log = get_xref_val("Logistics (CLO)")
        print(f"XRef Logistics: {log} (Expected: 30000)")
        
        # Verify ESG
        tax = get_xref_val("Tax Liability")
        print(f"XRef ESG Tax: {tax} (Expected: 5000)")

    else:
        print("FAIL: Cross Reference tab missing!")
        
    # 2. Verify Liquidity Monitor Hard Data Injection (Tab 1)
    ws_liq = wb["LIQUIDITY MONITOR"]
    
    # Starting Cash (B9) should be hard_data value (190000)
    start_cash = ws_liq['B9'].value
    print(f"\nStarting Cash (FN1): {start_cash} (Expected: 190000)")
    
    # Receivables FN1 (C21?) - Need to find row 'Receivables (HARD)'
    # It is roughly row 21 based on code logic.
    # Let's search column A
    found_rec = False
    for row in range(15, 30):
        val = ws_liq[f'A{row}'].value
        if val == "Receivables (HARD)":
            rec_val = ws_liq[f'C{row}'].value # FN1 is col C (3)
            print(f"Receivables FN1: {rec_val} (Expected: 12000)")
            found_rec = True
            break
            
    if not found_rec:
        print("FAIL: Could not find 'Receivables (HARD)' row")

    print("Test Complete.")

if __name__ == "__main__":
    seed_shared_outputs()
    test_cfo_generation()
