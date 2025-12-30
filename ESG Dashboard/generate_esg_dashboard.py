"""
ExSim ESG Dashboard - CO2 Abatement Strategy Dashboard

Compares ROI of different green initiatives against the cost of paying CO2 Taxes.
Helps the Sustainability Officer make data-driven environmental decisions.

Required libraries: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, LineChart, Reference, Series
import warnings

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Required input files from centralized Reports folder
REQUIRED_FILES = [
    'esg_report.xlsx',
    'production.xlsx',
    'ESG Decision.xlsx'
]

# Data source: Primary = Reports folder at project root, Fallback = local /data
REPORTS_FOLDER = Path(__file__).parent.parent / "Reports"
LOCAL_DATA_FOLDER = Path(__file__).parent / "data"

def get_data_path(filename):
    """Get data file path, checking Reports folder first, then local fallback."""
    primary = REPORTS_FOLDER / filename
    fallback = LOCAL_DATA_FOLDER / filename
    if primary.exists():
        return primary
    elif fallback.exists():
        return fallback
    return None

OUTPUT_FILE = "ESG_Dashboard.xlsx"

# Default initiative specifications
DEFAULT_INITIATIVES = {
    "Solar PV Panels": {
        "unit_cost": 0,  # $ per panel
        "co2_reduction": 0,  # tons per panel per year
        "unit": "panels",
        "capex": True
    },
    "Trees Planted": {
        "unit_cost": 0,  # $ per tree
        "co2_reduction": 0,  # tons per tree per year
        "unit": "trees",
        "capex": True
    },
    "Green Electricity": {
        "unit_cost": 0,  # $ premium per kWh
        "co2_reduction": 0,  # tons per kWh converted
        "unit": "% of consumption",
        "capex": False
    },
    "CO2 Credits": {
        "unit_cost": 0,  # $ per credit (1 credit = 1 ton)
        "co2_reduction": 0,  # 1 ton per credit
        "unit": "credits",
        "capex": False
    }
}

DEFAULT_CO2_TAX_RATE = 0  # Was 30
DEFAULT_EMISSIONS = 0  # Was 150
DEFAULT_ENERGY_CONSUMPTION = 0  # Was 500000


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_numeric(value):
    """Parse formatted number strings."""
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
    try:
        return float(cleaned)
    except:
        return 0


def load_excel_file(filepath):
    """Load Excel file."""
    try:
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


# =============================================================================
# DATA LOADING
# =============================================================================

def load_esg_report(filepath):
    """Load ESG report data."""
    df = load_excel_file(filepath)
    
    data = {
        'emissions': 0,
        'tax_paid': 0,
        'energy_consumption': 0
    }
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'emission' in first_val and 'total' in first_val:
            data['emissions'] = parse_numeric(row.iloc[1]) if len(row) > 1 else DEFAULT_EMISSIONS
        
        if 'tax' in first_val and ('paid' in first_val or 'bill' in first_val):
            data['tax_paid'] = parse_numeric(row.iloc[1]) if len(row) > 1 else data['emissions'] * DEFAULT_CO2_TAX_RATE
        
        if 'energy' in first_val and 'consumption' in first_val:
            data['energy_consumption'] = parse_numeric(row.iloc[1]) if len(row) > 1 else DEFAULT_ENERGY_CONSUMPTION
    
    return data


def load_production_data(filepath):
    """Load production data for scale reference."""
    df = load_excel_file(filepath)
    
    data = {'total_production': 50000}
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'production' in first_val and 'total' in first_val:
            data['total_production'] = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
    
    return data


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def create_esg_dashboard(esg_data, production_data):
    """Create the ESG Dashboard using openpyxl."""
    
    wb = Workbook()
    
    # Styles
    title_font = Font(bold=True, size=14, color="1F4E79")
    section_font = Font(bold=True, size=12, color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ref_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    best_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    initiative_fills = {
        'Solar PV Panels': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        'Trees Planted': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'Green Electricity': PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
        'CO2 Credits': PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
    }
    
    # =========================================================================
    # TAB 1: IMPACT_CONFIG
    # =========================================================================
    ws1 = wb.active
    ws1.title = "IMPACT_CONFIG"
    
    ws1['A1'] = "IMPACT CONFIGURATION - Initiative Specifications"
    ws1['A1'].font = title_font
    ws1['A2'] = "Define the unit costs and CO2 reduction rates from Case Guide."
    ws1['A2'].font = Font(italic=True, color="666666")
    
    # CO2 Tax Rate
    ws1['A4'] = "CO2 TAX RATE ($/Ton)"
    ws1['A4'].font = section_font
    cell = ws1['B4']
    cell.value = DEFAULT_CO2_TAX_RATE
    cell.fill = input_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    ws1['C4'] = "<-- Get from Case Guide!"
    ws1['C4'].font = Font(bold=True, italic=True, color="C00000")
    
    # Initiative specs table
    ws1['A6'] = "INITIATIVE SPECIFICATIONS"
    ws1['A6'].font = section_font
    
    headers = ['Initiative', 'Unit Cost ($)', 'CO2 Reduction (Tons/Unit)', 'Unit Type']
    for col, h in enumerate(headers, start=1):
        cell = ws1.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    init_start_row = 8
    for idx, (name, specs) in enumerate(DEFAULT_INITIATIVES.items()):
        row = init_start_row + idx
        
        cell = ws1.cell(row=row, column=1, value=name)
        cell.fill = initiative_fills.get(name, ref_fill)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        
        cell = ws1.cell(row=row, column=2, value=specs['unit_cost'])
        cell.fill = input_fill
        cell.border = thin_border
        cell.number_format = '$#,##0.00'
        
        cell = ws1.cell(row=row, column=3, value=specs['co2_reduction'])
        cell.fill = input_fill
        cell.border = thin_border
        cell.number_format = '0.0000'
        
        cell = ws1.cell(row=row, column=4, value=specs['unit'])
        cell.fill = ref_fill
        cell.border = thin_border
    
    init_end_row = init_start_row + len(DEFAULT_INITIATIVES) - 1
    
    # Guidance notes
    ws1['A14'] = "NOTES:"
    ws1['A14'].font = Font(bold=True)
    ws1['A15'] = "- Solar: CAPEX investment, long-term savings"
    ws1['A16'] = "- Trees: Low cost, slow reduction, good for PR"
    ws1['A17'] = "- Green Electricity: Operating cost, immediate impact"
    ws1['A18'] = "- CO2 Credits: Quick fix, no long-term benefit"
    
    # Column widths
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 18
    
    # =========================================================================
    # TAB 2: STRATEGY_SELECTOR
    # =========================================================================
    ws2 = wb.create_sheet("STRATEGY_SELECTOR")
    
    ws2['A1'] = "STRATEGY SELECTOR - CO2 Abatement Calculator"
    ws2['A1'].font = title_font
    ws2['A2'] = "Compare ROI of green initiatives vs. paying CO2 taxes."
    ws2['A2'].font = Font(italic=True, color="666666")
    
    # Section A: Baseline
    ws2['A4'] = "SECTION A: CURRENT BASELINE"
    ws2['A4'].font = section_font
    
    ws2['A6'] = "Current CO2 Emissions (Tons/Year)"
    cell = ws2['B6']
    cell.value = esg_data['emissions']
    cell.fill = ref_fill
    cell.border = thin_border
    
    ws2['A7'] = "Current CO2 Tax Bill ($)"
    cell = ws2['B7']
    cell.value = f'=B6*IMPACT_CONFIG!$B$4'
    cell.fill = calc_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    ws2['A8'] = "Energy Consumption (kWh/Year)"
    cell = ws2['B8']
    cell.value = esg_data['energy_consumption']
    cell.fill = ref_fill
    cell.border = thin_border
    cell.number_format = '#,##0'
    
    # Section B: Investment Simulator
    ws2['A10'] = "SECTION B: INVESTMENT SIMULATOR"
    ws2['A10'].font = section_font
    ws2['A11'] = "Enter quantities in yellow cells to see impact."
    ws2['A11'].font = Font(italic=True, color="666666")
    
    sim_headers = ['Initiative', 'Quantity', 'Investment/Cost', 'CO2 Reduced (Tons)', 
                   'Tax Savings ($)', 'Net Annual Benefit', 'Payback (Years)', 'Cost per Ton']
    
    for col, h in enumerate(sim_headers, start=1):
        cell = ws2.cell(row=12, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    sim_start_row = 13
    initiatives = list(DEFAULT_INITIATIVES.keys())
    
    for idx, name in enumerate(initiatives):
        row = sim_start_row + idx
        init_config_row = init_start_row + idx
        
        # Initiative name
        cell = ws2.cell(row=row, column=1, value=name)
        cell.fill = initiative_fills.get(name, ref_fill)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        
        # Quantity (input)
        if name == "Green Electricity":
            # Percentage input for green electricity
            cell = ws2.cell(row=row, column=2, value=0)  # 0%
            cell.number_format = '0%'
        else:
            cell = ws2.cell(row=row, column=2, value=0)
        cell.fill = input_fill
        cell.border = thin_border
        
        # Investment/Cost calculation
        if name == "Green Electricity":
            # Cost = Energy * Qty% * Premium
            cell = ws2.cell(row=row, column=3, 
                value=f'=$B$8*B{row}*IMPACT_CONFIG!B{init_config_row}')
        else:
            # Cost = Qty * Unit_Cost
            cell = ws2.cell(row=row, column=3, 
                value=f'=B{row}*IMPACT_CONFIG!B{init_config_row}')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # CO2 Reduced
        if name == "Green Electricity":
            # Reduction = Energy * Qty% * Reduction_Rate
            cell = ws2.cell(row=row, column=4, 
                value=f'=$B$8*B{row}*IMPACT_CONFIG!C{init_config_row}')
        else:
            # Reduction = Qty * Reduction_Per_Unit
            cell = ws2.cell(row=row, column=4, 
                value=f'=B{row}*IMPACT_CONFIG!C{init_config_row}')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '#,##0.0'
        
        # Tax Savings
        cell = ws2.cell(row=row, column=5, 
            value=f'=D{row}*IMPACT_CONFIG!$B$4')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Net Annual Benefit (Savings - Annual Cost for OpEx, just Savings for CapEx)
        if DEFAULT_INITIATIVES[name]['capex']:
            # For CAPEX: Annual benefit = full tax savings
            cell = ws2.cell(row=row, column=6, value=f'=E{row}')
        else:
            # For OpEx: Annual benefit = tax savings - cost
            cell = ws2.cell(row=row, column=6, value=f'=E{row}-C{row}')
        cell.fill = calc_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Payback Period
        if DEFAULT_INITIATIVES[name]['capex']:
            # Payback = Investment / Annual Savings
            cell = ws2.cell(row=row, column=7, 
                value=f'=IF(E{row}>0,C{row}/E{row},"N/A")')
        else:
            # No payback for OpEx
            cell = ws2.cell(row=row, column=7, value="N/A (OpEx)")
        cell.fill = calc_fill
        cell.border = thin_border
        if DEFAULT_INITIATIVES[name]['capex']:
            cell.number_format = '0.0'
        
        # Cost per Ton Abated
        # Logic: If Cost ($) / CO2 (Ton), that shows efficiency.
        # But wait, we want "Cost to lower emissions".
        # If I spend $1000 to save 10 tons, cost is $100/ton.
        # If Tax is $30/ton, this is a bad deal (100 > 30).
        # Calculation: Cost / CO2 Reduced
        # For Green Elec (OpEx): Annual Cost / Annual Reduction
        # For Solar (CapEx): This is tricky. Usually annualized cost over lifetime?
        # Or just Investment / Total Lifetime Reduction?
        # Let's simplify: Investment / (Annual Reduction * 10 years)? Not standard.
        # Let's use simple logic: Investment / (Annual Reduction * 1) for OpEx
        # For CapEx: Let's assume a 10 year horizon for standard abatement curve? 
        # Or just stick to Payback.
        # Let's stick to "Cost Per Ton" as defined by: Annual Cost / Annual Reduction (for OpEx) or Investment / Annual Reduction / 10 (for CapEx annualized)
        # Actually simplified View: Cost per Ton = Total Investment / Total Lifetime CO2 (Assume 20 yrs for solar, 50 for trees)
        # Or just follow what the sheet had: "=IF(D{row}>0,C{row}/D{row},0)" which implies Investment / Annual CO2.
        # That's equivalent to "How many years of tax savings to pay back?"
        # Let's stick to the previous formula logic but visualize it.
        # ACTUALLY, "Cost per Ton Abated" in economics is usually Net Present Value.
        # Here we will use: (Annual Cost - Tax Savings) / Tons Reduced?
        # No, simpler: Cost of Initiative / Tons Reduced.
        # Let's use the formula that was already there: C (Cost) / D (Reduction).
        # This gives $ per Annual Ton.
        cell = ws2.cell(row=row, column=8, 
            value=f'=IF(D{row}>0,C{row}/D{row},0)')
        cell.fill = output_fill
        cell.border = thin_border
        cell.number_format = '$#,##0'
        
        # Hidden Helper Column for Chart Threshold (Tax Rate)
        ws2.cell(row=row, column=9, value="=IMPACT_CONFIG!$B$4")
        
    
    sim_end_row = sim_start_row + len(initiatives) - 1
    
    # ---------------------------------------------------------
    # CONDITIONAL FORMATTING
    # ---------------------------------------------------------
    
    # Green Scale for 'Cost per Ton' (Col H) - Dark Green = Lowest Cost
    # We want lower numbers to be "Good" (Green).
    ws2.conditional_formatting.add(
        f'H{sim_start_row}:H{sim_end_row}',
        ColorScaleRule(start_type='min', start_color='63BE7B', # Green
                       mid_type='percentile', mid_value=50, mid_color='FFEB84', # Yellow
                       end_type='max', end_color='F8696B') # Red
    )
    
    # Red Text for Payback > 5 Years (Col G)
    ws2.conditional_formatting.add(
        f'G{sim_start_row}:G{sim_end_row}',
        FormulaRule(formula=[f'AND(ISNUMBER(G{sim_start_row}),G{sim_start_row}>5)'], font=Font(color="C00000", bold=True))
    )
    
    # ---------------------------------------------------------
    # CHART: "Abatement Cost Curve" (Combo Chart)
    # ---------------------------------------------------------
    
    c1 = BarChart()
    c1.type = "col"
    c1.style = 10
    c1.title = "Abatement Cost Curve (Cost vs Tax)"
    c1.y_axis.title = "$ Cost / Ton Avg"
    c1.height = 10
    c1.width = 15
    
    # Series 1: Cost per Ton (Bars)
    data_cost = Reference(ws2, min_col=8, min_row=sim_start_row, max_row=sim_end_row)
    s1 = Series(data_cost, title="Cost Per Ton")
    c1.append(s1)
    
    # Categories
    cats = Reference(ws2, min_col=1, min_row=sim_start_row, max_row=sim_end_row)
    c1.set_categories(cats)
    
    # Series 2: Tax Rate (Line) - Threshold
    c2 = LineChart()
    data_tax = Reference(ws2, min_col=9, min_row=sim_start_row, max_row=sim_end_row)
    s2 = Series(data_tax, title="Tax Rate Threshold")
    s2.graphicalProperties.line.solidFill = "FF0000"
    s2.graphicalProperties.line.width = 20000 
    c2.append(s2)
    
    c1 += c2
    ws2.add_chart(c1, "J10")
    
    
    # Section C: The Verdict
    ws2['A18'] = "SECTION C: THE VERDICT"
    ws2['A18'].font = section_font
    
    # Summary metrics
    ws2['A20'] = "Total CO2 Reduced"
    cell = ws2['B20']
    cell.value = f'=SUM(D{sim_start_row}:D{sim_end_row})'
    cell.fill = output_fill
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.0'
    ws2['C20'] = "tons/year"
    
    ws2['A21'] = "Total Investment Required"
    cell = ws2['B21']
    # Sum only CAPEX items (Solar=row 13, Trees=row 14) -> Dynamic check? 
    # Hardcoded based on order for simplicity, or SUMIF based on config...
    # Let's just sum all Cost column (Investment for Capex, Annual for Opex?)
    # The header says "Investment/Cost".
    cell.value = f'=SUM(C{sim_start_row}:C{sim_end_row})'
    cell.fill = output_fill
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.number_format = '$#,##0'
    
    ws2['A22'] = "New Tax Bill"
    cell = ws2['B22']
    cell.value = f'=MAX(0,B7-B22*IMPACT_CONFIG!B4)' # Warning: Circular logic if B22 refs itself.
    # Logic: New Tax = (Emissions - Reduced) * Rate
    cell.value = f'=MAX(0,(B6-B20)*IMPACT_CONFIG!B4)'
    cell.fill = output_fill
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    ws2['A23'] = "Annual Savings"
    cell = ws2['B23']
    cell.value = f'=B7-B22'
    cell.fill = best_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = thin_border
    cell.number_format = '$#,##0'
    
    # Column widths
    ws2.column_dimensions['A'].width = 30
    for col in range(2, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 15
        
    # Save
    wb.save(OUTPUT_FILE)
    print(f"[SUCCESS] Created '{OUTPUT_FILE}'")


def main():
    """Main function."""
    print("ExSim ESG Dashboard Generator")
    print("=" * 50)
    
    print("\n[*] Loading data files...")
    print(f"    Primary source: {REPORTS_FOLDER}")
    print(f"    Fallback source: {LOCAL_DATA_FOLDER}")
    
    # ESG Report
    esg_path = get_data_path("esg_report.xlsx")
    if esg_path:
        esg_data = load_esg_report(esg_path)
        print(f"  [OK] Loaded ESG report from {esg_path.parent.name}/")
    else:
        esg_data = load_esg_report(None)
        print("  [!] Using default ESG data")
        
    # Production Data
    prod_path = get_data_path("production.xlsx")
    if prod_path:
        prod_data = load_production_data(prod_path)
        print(f"  [OK] Loaded Production data")
    else:
        prod_data = load_production_data(None)
        print("  [!] Using default Production data")
    
    print("\n[*] Creating dashboard...")
    
    create_esg_dashboard(esg_data, prod_data)
    
    print("\nSheets created:")
    print("  * IMPACT_CONFIG (Initiative Specs)")
    print("  * STRATEGY_SELECTOR (ROI Calculator & Abatement Curve)")


if __name__ == "__main__":
    main()
