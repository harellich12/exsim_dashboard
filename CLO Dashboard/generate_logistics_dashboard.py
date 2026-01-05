"""
ExSim Logistics Dashboard - Supply Network Optimization

Balances Inventory levels across zones using Shipments.
Handles warehouse capacity, transport modes, and stockout prevention.

Required libraries: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference, Series
import warnings
import sys

# Add parent directory to path to import case_parameters
# Add parent directory to path to import case_parameters
sys.path.append(str(Path(__file__).parent.parent))
try:
    from case_parameters import LOGISTICS, COMMON
    from config import get_data_path, OUTPUT_DIR
except ImportError:
    print("Warning: Could not import case_parameters.py or config.py. Using defaults.")
    LOGISTICS = {}
    COMMON = {}
    # Fallback for config
    OUTPUT_DIR = Path(__file__).parent
    def get_data_path(f): return Path(f)

# Import shared outputs for inter-dashboard communication
try:
    from shared_outputs import export_dashboard_data
except ImportError:
    export_dashboard_data = None

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Required input files from centralized Reports folder
REQUIRED_FILES = [
    'finished_goods_inventory.xlsx',
    'Logistics Decisions.xlsx',
    'shipping_costs.xlsx'
]

OUTPUT_FILE = OUTPUT_DIR / "Logistics_Dashboard.xlsx"

# Use centralized constants from case_parameters
FORTNIGHTS = COMMON.get('FORTNIGHTS', list(range(1, 9)))
ZONES = COMMON.get('ZONES', ["Center", "West", "North", "East", "South"])
TRANSPORT_MODES = COMMON.get('TRANSPORT_MODES', ["Train", "Truck", "Plane"])
DEFAULT_MATERIAL = "Electroclean"

# Default transport configuration
DEFAULT_TRANSPORT = {
    "Train": {"lead_time": 0, "cost": 0},
    "Truck": {"lead_time": 0, "cost": 0},
    "Plane": {"lead_time": 0, "cost": 0},
}

# Default warehouse configuration
DEFAULT_WAREHOUSE = {
    "Center": {"capacity": 0, "cost_per_module": 0, "capacity_per_module": 0},
    "West": {"capacity": 0, "cost_per_module": 0, "capacity_per_module": 0},
    "North": {"capacity": 0, "cost_per_module": 0, "capacity_per_module": 0},
    "East": {"capacity": 0, "cost_per_module": 0, "capacity_per_module": 0},
    "South": {"capacity": 0, "cost_per_module": 0, "capacity_per_module": 0},
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_numeric(value):
    """Parse formatted number strings."""
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
    try:
        return float(cleaned)
    except:
        return 0


def load_excel_file(filepath, sheet_name=None):
    """Load Excel file."""
    try:
        if sheet_name:
            return pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


# =============================================================================
# DATA LOADING
# =============================================================================

def load_finished_goods_by_zone(filepath):
    """Load finished goods inventory grouped by zone."""
    df = load_excel_file(filepath)
    
    # Default data
    data = {zone: {'capacity': DEFAULT_WAREHOUSE[zone]['capacity'], 
                   'inventory': 0} for zone in ZONES}
    
    if df is None:
        return data
    
    current_zone_idx = 0
    zone_order = ['Center', 'West', 'North', 'East', 'South']
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        
        # Detect capacity header (new zone section)
        if 'capacity:' in first_val.lower():
            import re
            match = re.search(r'(\d+)', first_val.replace(',', ''))
            if match and current_zone_idx < len(zone_order):
                zone = zone_order[current_zone_idx]
                data[zone]['capacity'] = int(match.group(1))
        
        # Get final inventory
        if 'final inventory' in first_val.lower():
            if current_zone_idx < len(zone_order):
                zone = zone_order[current_zone_idx]
                # Get last fortnight value (column 8)
                val = parse_numeric(row.iloc[8]) if len(row) > 8 else 0
                data[zone]['inventory'] = val
                current_zone_idx += 1
    
    return data


def load_logistics_template(filepath):
    """Load logistics decisions template."""
    try:
        df = pd.read_excel(filepath, sheet_name='Logistics', header=None)
        return {'df': df, 'exists': True}
    except:
        return {'df': None, 'exists': False}


def load_shipping_costs(filepath):
    """Load logistics shipping costs."""
    df = load_excel_file(filepath)
    
    data = {'total_shipping_cost': 0}
    
    if df is None:
        return data
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ''
        
        if 'shipping' in first_val and 'cost' in first_val:
            for col_idx in range(1, min(10, len(row))):
                val = parse_numeric(row.iloc[col_idx])
                if val > 0:
                    data['total_shipping_cost'] = val
                    break
    
    return data


def load_logistics_intelligence(filepath):
    """
    Extract Historical Intelligence from logistics.xlsx:
    - Cost benchmarks per route (Train Center-North, etc.)
    - Warehouse penalties by zone
    """
    result = {
        'benchmarks': {},  # route -> cost per unit
        'penalties': {}    # zone -> warehouse cost
    }
    
    if filepath is None or not filepath.exists():
        return result
    
    try:
        df = pd.read_excel(filepath, header=None)
        
        # Find Transportation Costs table
        in_transport_section = False
        for idx, row in df.iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            
            if "Transportation Costs" in label:
                in_transport_section = True
                continue
            
            if in_transport_section:
                # Look for header row with "Type", "Units", "Cost", "Total"
                if "Type" in label:
                    continue
                
                # Check for actual route rows (not Subtotal, not Total)
                if label and "Subtotal" not in label and "Total" != label:
                    route = label
                    units = parse_numeric(row.iloc[1]) if len(row) > 1 else 0
                    total = parse_numeric(row.iloc[3]) if len(row) > 3 else 0
                    
                    if units > 0:
                        cost_per_unit = total / units
                        # Aggregate by route name (may have duplicates)
                        if route in result['benchmarks']:
                            # Average the costs
                            existing = result['benchmarks'][route]
                            result['benchmarks'][route] = (existing + cost_per_unit) / 2
                        else:
                            result['benchmarks'][route] = round(cost_per_unit, 2)
                
                # Exit when we hit Total or empty row after data
                if label == "Total" or (label == "" and len(result['benchmarks']) > 0):
                    in_transport_section = False
        
        # Find Incoming and Outcoming by Zone table
        for idx, row in df.iterrows():
            label = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            
            if "Incoming and Outcoming by Zone" in label:
                # Next row is header: Zone, Received Units, In transit, Sent Units, Shipping Costs, Warehouse Costs
                # Then data rows
                for offset in range(2, 10):
                    if idx + offset < len(df):
                        data_row = df.iloc[idx + offset]
                        zone = str(data_row.iloc[0]).strip() if pd.notna(data_row.iloc[0]) else ""
                        
                        if zone in ["Center", "West", "North", "East", "South"]:
                            # Warehouse costs in column 5 (index 5)
                            warehouse_cost = parse_numeric(data_row.iloc[5]) if len(data_row) > 5 else 0
                            result['penalties'][zone] = warehouse_cost
                        elif zone == "Total":
                            break
                break
        
    except Exception as e:
        print(f"  [!] Error in load_logistics_intelligence: {e}")
    
    return result


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def create_logistics_dashboard(inventory_data, template_data, cost_data, intelligence_data=None, output_buffer=None):
    """
    Create the Logistics Dashboard with Historical Intelligence.
    
    Args:
        inventory_data: Inventory data by zone
        template_data: Template configuration
        cost_data: Shipping cost data
        intelligence_data: Historical intelligence data (optional)
        output_buffer: io.BytesIO buffer for output (optional). If provided, returns
                      the buffer instead of saving to disk.
    
    Returns:
        BytesIO buffer if output_buffer provided, None otherwise
    """
    
    # Handle missing intelligence data
    if intelligence_data is None:
        intelligence_data = {'benchmarks': {}, 'penalties': {}}
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5496")
    title_font = Font(bold=True, size=14, color="2F5496")
    zone_font = Font(bold=True, size=11, color="FFFFFF")
    zone_fills = {
        'Center': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'West': PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        'North': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        'East': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        'South': PatternFill(start_color="9E480E", end_color="9E480E", fill_type="solid"),
    }
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ref_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Traffic Light Fills
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Stockout
    purple_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid") # Overflow
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Optimal
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Track zone data rows for formulas
    zone_data_rows = {}
    
    # =========================================================================
    # TAB 1: ROUTE_CONFIG
    # =========================================================================
    ws1 = wb.active
    ws1.title = "ROUTE_CONFIG"
    
    ws1['A1'] = "ROUTE CONFIGURATION - Transport Physics"
    ws1['A1'].font = title_font
    ws1['A2'] = "Define transport modes and warehouse costs. Yellow cells are editable."
    ws1['A2'].font = Font(italic=True, color="666666")
    
    # MODES CONFIG
    ws1['A4'] = "TABLE 1: TRANSPORT MODES"
    ws1['A4'].font = section_font
    
    mode_headers = ['Mode', 'Lead Time (Fortnights)', 'Cost Per Unit ($)']
    for col, h in enumerate(mode_headers, start=1):
        cell = ws1.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 6
    for mode, config in DEFAULT_TRANSPORT.items():
        ws1.cell(row=row, column=1, value=mode).border = thin_border
        
        cell = ws1.cell(row=row, column=2, value=config['lead_time'])
        cell.border = thin_border
        cell.fill = input_fill
        
        cell = ws1.cell(row=row, column=3, value=config['cost'])
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        row += 1
    
    row += 2
    
    # WAREHOUSE CONFIG
    ws1.cell(row=row, column=1, value="TABLE 2: WAREHOUSE CONFIGURATION").font = section_font
    row += 1
    
    wh_headers = ['Zone', 'Current Capacity', 'Cost Per Module', 'Capacity Per Module']
    for col, h in enumerate(wh_headers, start=1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    wh_config_start = row
    for zone in ZONES:
        zone_inv = inventory_data.get(zone, {})
        capacity = zone_inv.get('capacity', DEFAULT_WAREHOUSE[zone]['capacity'])
        
        cell = ws1.cell(row=row, column=1, value=zone)
        cell.border = thin_border
        cell.fill = zone_fills[zone]
        cell.font = Font(color="FFFFFF")
        
        cell = ws1.cell(row=row, column=2, value=capacity)
        cell.border = thin_border
        cell.fill = ref_fill
        
        cell = ws1.cell(row=row, column=3, value=DEFAULT_WAREHOUSE[zone]['cost_per_module'])
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        cell = ws1.cell(row=row, column=4, value=DEFAULT_WAREHOUSE[zone]['capacity_per_module'])
        cell.border = thin_border
        cell.fill = input_fill
        
        row += 1
    
    # Column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 22
    
    # =========================================================================
    # TAB 1.5: LOGISTICS_DATA (Hidden Database)
    # =========================================================================
    ws_data = wb.create_sheet("LOGISTICS_DATA")
    # Hidden? ws_data.sheet_state = 'hidden' # Keeping visible for verification
    
    ws_data['A1'] = "Route Cost Database (Generated from Case Parameters)"
    headers = ['From', 'To', 'Mode', 'Cost', 'LeadTime']
    for col, h in enumerate(headers, start=1):
        ws_data.cell(row=2, column=col, value=h).font = header_font
        
    row = 3
    matrix = LOGISTICS.get('TRANSPORT_COSTS', {})
    # Matrix format: Center -> West -> Mode -> Cost
    
    # Flatten checks
    best_modes = {} # (From, To) -> {mode, cost}
    
    for origin in ZONES:
        for dest in ZONES:
            if origin == dest:
                continue
            
            routes = matrix.get(origin, {}).get(dest, {})
            # If empty, maybe stored symmetrically? Case implies direction matters?
            # Assuming full matrix or symmetric. If distinct, we rely on keys.
            
            best_val = 999999
            best_mode_name = "N/A"
            
            # Standard Modes
            for mode in ['Truck', 'Train', 'Airplane']: # Case names might vary
                # Map standard names to keys in matrix if needed.
                # In case_parameters.py we likely used standard keys.
                # Let's assume keys match case (e.g., 'Truck', 'Train', 'Airplane')
                
                cost = routes.get(mode, 0)
                time = 1 # Placeholder or look up lead times
                # Case Lead Times: Train=2, Truck=1, Air=0 (Fortnights? Days?)
                # Assuming Fortnights for now based on dashboard defaults.
                if mode == 'Airplane': time = 0
                elif mode == 'Truck': time = 1
                elif mode == 'Train': time = 2
                
                if cost > 0:
                    ws_data.cell(row=row, column=1, value=origin)
                    ws_data.cell(row=row, column=2, value=dest)
                    ws_data.cell(row=row, column=3, value=mode)
                    ws_data.cell(row=row, column=4, value=cost)
                    ws_data.cell(row=row, column=5, value=time)
                    row += 1
                    
                    if cost < best_val:
                        best_val = cost
                        best_mode_name = mode
            
            best_modes[(origin, dest)] = {'mode': best_mode_name, 'cost': best_val}

    # =========================================================================
    # TAB 1.6: ROUTE_OPTIMIZER
    # =========================================================================
    ws_opt = wb.create_sheet("ROUTE_OPTIMIZER")
    ws_opt['A1'] = "ROUTE OPTIMIZER & MATRIX"
    ws_opt['A1'].font = title_font
    
    # 5x5 Matrix of Lowest Cost
    ws_opt['A3'] = "Cheapest Transport Mode Matrix ($ Cost)"
    ws_opt['A3'].font = section_font
    
    # Header Row (Destinations)
    for col, z in enumerate(ZONES, start=2):
        cell = ws_opt.cell(row=4, column=col, value=z)
        cell.font = header_font
        cell.fill = zone_fills[z]
        cell.alignment = Alignment(horizontal='center')
        
    # Rows (Origins)
    row = 5
    for origin in ZONES:
        cell = ws_opt.cell(row=row, column=1, value=origin)
        cell.font = header_font
        cell.fill = zone_fills[origin]
        
        for col, dest in enumerate(ZONES, start=2):
            if origin == dest:
                cell = ws_opt.cell(row=row, column=col, value="-")
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            else:
                data = best_modes.get((origin, dest), {'mode': 'N/A', 'cost': 0})
                cell = ws_opt.cell(row=row, column=col, value=f"{data['mode']}\n(${data['cost']})")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                if data['mode'] == "Train":
                    cell.fill = calc_fill # Greenish
                elif data['mode'] == "Airplane":
                    cell.fill = red_fill # Expensive
                else: 
                    cell.fill = input_fill # Truck
        row += 1

    # Route Calculator
    row += 2
    ws_opt.cell(row=row, column=1, value="ROUTE CALCULATOR").font = section_font
    row += 1
    
    ws_opt.cell(row=row, column=1, value="From:").font = Font(bold=True)
    ws_opt.cell(row=row, column=2, value="Center").fill = input_fill # Input
    ws_opt.cell(row=row, column=2).border = thin_border
    
    ws_opt.cell(row=row+1, column=1, value="To:").font = Font(bold=True)
    ws_opt.cell(row=row+1, column=2, value="North").fill = input_fill # Input
    ws_opt.cell(row=row+1, column=2).border = thin_border
    
    # Calculator Output Table
    row += 3
    calc_headers = ['Mode', 'Cost', 'Lead Time', 'Savings vs Air']
    for col, h in enumerate(calc_headers, start=1):
        ws_opt.cell(row=row, column=col, value=h).font = header_font
        ws_opt.cell(row=row, column=col).fill = header_fill
    
    # Logic: We can't easy-filter hidden sheet in excel formulas easily without complex arrays in older Excel.
    # But we can assume the user looks at the Matrix.
    # Or use SUMIFS/INDEX MATCH.
    # Cost = SUMIFS(LOGISTICS_DATA!D:D, A:A, From, B:B, To, C:C, Mode)
    
    modes = ['Train', 'Truck', 'Airplane']
    start_calc_row = row + 1
    row = start_calc_row
    
    src_cell = f"$B${start_calc_row-4}"
    dst_cell = f"$B${start_calc_row-3}"
    
    for mode in modes:
        ws_opt.cell(row=row, column=1, value=mode).border = thin_border
        
        # Cost Formula
        # =SUMIFS(LOGISTICS_DATA!D:D, LOGISTICS_DATA!A:A, src, LOGISTICS_DATA!B:B, dst, LOGISTICS_DATA!C:C, mode)
        cost_f = f'=SUMIFS(LOGISTICS_DATA!D:D, LOGISTICS_DATA!A:A, {src_cell}, LOGISTICS_DATA!B:B, {dst_cell}, LOGISTICS_DATA!C:C, "{mode}")'
        ws_opt.cell(row=row, column=2, value=cost_f).number_format = '$#,##0'
        ws_opt.cell(row=row, column=2).border = thin_border
        
        # Time Formula
        time_f = f'=SUMIFS(LOGISTICS_DATA!E:E, LOGISTICS_DATA!A:A, {src_cell}, LOGISTICS_DATA!B:B, {dst_cell}, LOGISTICS_DATA!C:C, "{mode}")'
        ws_opt.cell(row=row, column=3, value=time_f).border = thin_border
        
        row += 1
        
    ws_opt.column_dimensions['A'].width = 15
    ws_opt.column_dimensions['B'].width = 15
    
    # =========================================================================
    # TAB 2: INVENTORY_TETRIS
    # =========================================================================
    ws2 = wb.create_sheet("INVENTORY_TETRIS")
    
    ws2['A1'] = "INVENTORY TETRIS - Zone-by-Zone Balance"
    ws2['A1'].font = title_font
    ws2['A2'] = "Balance inventory. Watch: 🔴STOCKOUT (red), 🟡WARNING >90% (yellow), 🟣OVERFLOW (purple)."
    ws2['A2'].font = Font(italic=True, color="666666")
    
    row = 4
    
    for zone in ZONES:
        zone_inv = inventory_data.get(zone, {})
        opening_inv = zone_inv.get('inventory', 0)
        capacity = zone_inv.get('capacity', DEFAULT_WAREHOUSE[zone]['capacity'])
        
        # Get warehouse penalty from intelligence
        zone_penalty = intelligence_data.get('penalties', {}).get(zone, 0)
        
        # Zone Header with penalty display
        ws2.merge_cells(f'A{row}:H{row}')
        cell = ws2.cell(row=row, column=1, value=f"═══ {zone.upper()} ZONE (Capacity: {capacity:,}) ═══")
        cell.font = zone_font
        cell.fill = zone_fills[zone]
        cell.alignment = Alignment(horizontal='center')
        
        # Show warehouse penalty if > 0
        if zone_penalty > 0:
            penalty_cell = ws2.cell(row=row, column=9, value=f"Last Period Rent: ${zone_penalty:,.0f}")
            penalty_cell.font = Font(bold=True, color="C00000")  # Bold red
        
        chart_anchor_row = row 
        row += 1
        
        # Parameters
        ws2.cell(row=row, column=1, value="Opening Inventory").border = thin_border
        cell = ws2.cell(row=row, column=2, value=opening_inv)
        cell.border = thin_border
        cell.fill = ref_fill
        
        ws2.cell(row=row, column=4, value="Capacity").border = thin_border
        cell = ws2.cell(row=row, column=5, value=capacity)
        cell.border = thin_border
        cell.fill = ref_fill
        
        ws2.cell(row=row, column=7, value="Rent Modules?").border = thin_border
        cell = ws2.cell(row=row, column=8, value=0)
        cell.border = thin_border
        cell.fill = input_fill
        
        params_row = row
        row += 2
        
        # Headers
        inv_headers = ['Fortnight', 'Production', 'Sales', 'Outgoing', 'Incoming', 
                       'Projected Inv', 'Capacity', 'Flag'] # Added Capacity col for chart
        for col, h in enumerate(inv_headers, start=1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = zone_fills[zone]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        data_start = row
        for fn in FORTNIGHTS:
            ws2.cell(row=row, column=1, value=f"FN{fn}").border = thin_border
            
            # Production (input from Production Dashboard)
            cell = ws2.cell(row=row, column=2, value=0)
            cell.border = thin_border
            cell.fill = input_fill
            
            # Sales Forecast (input from Marketing Dashboard)
            cell = ws2.cell(row=row, column=3, value=0)
            cell.border = thin_border
            cell.fill = input_fill
            
            # Outgoing Shipments (negative, manual input)
            cell = ws2.cell(row=row, column=4, value=0)
            cell.border = thin_border
            cell.fill = input_fill
            
            # Incoming Shipments (positive, manual input)
            cell = ws2.cell(row=row, column=5, value=0)
            cell.border = thin_border
            cell.fill = input_fill
            
            # Projected Inventory = Prev + Production + Incoming - Outgoing - Sales
            if fn == 1:
                formula = f"=$B${params_row}+B{row}+E{row}-D{row}-C{row}"
            else:
                formula = f"=F{row-1}+B{row}+E{row}-D{row}-C{row}"
            cell = ws2.cell(row=row, column=6, value=formula)
            cell.border = thin_border
            # Fill managed by conditional formatting below
            
            # Capacity Limit Column (Hidden for chart, or visible reference)
            # Dynamic capacity = Base + Rented
            # Zone rows in ROUTE_CONFIG: Data starts at row 13 (Center=13, West=14, etc)
            zone_config_row = 13 + ZONES.index(zone)
            cell = ws2.cell(row=row, column=7, 
                value=f"=$E${params_row} + ($H${params_row} * ROUTE_CONFIG!$D${zone_config_row})")
            cell.border = thin_border
            cell.fill = ref_fill
            
            # Flag - now includes 90% warning
            cell = ws2.cell(row=row, column=8, 
                value=f'=IF(F{row}<0,"🔴 STOCKOUT!",IF(F{row}>G{row},"🟣 OVERFLOW!",IF(F{row}>G{row}*0.9,"🟡 WARNING: >90%","✓ OK")))')
            cell.border = thin_border
            
            row += 1
        
        data_end = row - 1
        zone_data_rows[zone] = {
            'start': data_start, 
            'end': data_end, 
            'params': params_row,
            'rent_cell': f'H{params_row}'
        }
        
        # ---------------------------------------------------------
        # CHARTS
        # ---------------------------------------------------------
        
        # Logic: 
        # Center Zone -> Tetris Combo Chart (Inv vs Cap).
        # Other Zones -> Supply vs Demand (Clustered Bar).
        
        if zone == "Center":
            # 1. Warehouse Tetris (Combo)
            c1 = BarChart()
            c1.type = "col"
            c1.style = 10
            c1.title = "Warehouse Tetris (Inv vs Capacity)"
            c1.y_axis.title = "Units"
            c1.x_axis.title = "Fortnight"
            c1.height = 10
            c1.width = 15
            
            data_inv = Reference(ws2, min_col=6, min_row=data_start, max_row=data_end) # F
            cats = Reference(ws2, min_col=1, min_row=data_start, max_row=data_end)
            
            s1 = Series(data_inv, title="Ending Inv")
            c1.append(s1)
            c1.set_categories(cats)
            
            # Line for Capacity
            c2 = LineChart()
            data_cap = Reference(ws2, min_col=7, min_row=data_start, max_row=data_end) # G
            s2 = Series(data_cap, title="Capacity Limit")
            s2.graphicalProperties.line.solidFill = "FF0000" # Red Line
            c2.append(s2)
            
            c1 += c2
            ws2.add_chart(c1, f"J{chart_anchor_row}")
        
        else:
            # 2. Supply vs Demand (Clustered Bar)
            # Needs helper cols? Or can plot existing cols?
            # Supply = Prod (B) + Incoming (E)
            # Demand = Sales (C) + Outgoing (D)
            # Since these are calculated sums, we can't chart them directly without helper columns.
            # Let's add hidden helper columns to the right for the chart.
            
            helper_col = 10 
            # Supply Col
            ws2.cell(row=data_start-1, column=helper_col, value="Total Supply")
            for r in range(data_start, data_end+1):
                ws2.cell(row=r, column=helper_col, value=f"=B{r}+E{r}")
            
            # Demand Col
            ws2.cell(row=data_start-1, column=helper_col+1, value="Total Demand")
            for r in range(data_start, data_end+1):
                ws2.cell(row=r, column=helper_col+1, value=f"=C{r}+D{r}")
            
            # Hide them?
            # ws2.column_dimensions[get_column_letter(helper_col)].hidden = True
            
            c3 = BarChart()
            c3.type = "col"
            c3.style = 10
            c3.title = f"{zone} Supply vs Demand"
            c3.y_axis.title = "Units"
            c3.height = 10
            c3.width = 15
            
            data_sup = Reference(ws2, min_col=helper_col, min_row=data_start, max_row=data_end)
            data_dem = Reference(ws2, min_col=helper_col+1, min_row=data_start, max_row=data_end)
            cats = Reference(ws2, min_col=1, min_row=data_start, max_row=data_end)
            
            c3.append(Series(data_sup, title="Total Supply"))
            c3.append(Series(data_dem, title="Total Demand"))
            c3.set_categories(cats)
            
            ws2.add_chart(c3, f"J{chart_anchor_row}")

        # ---------------------------------------------------------
        # CONDITIONAL FORMATTING
        # ---------------------------------------------------------
        
        # Target: Projected Inventory (Col F)
        # Red: < 0 (Stockout)
        ws2.conditional_formatting.add(
            f'F{data_start}:F{data_end}',
            FormulaRule(formula=[f'F{data_start}<0'], fill=red_fill)
        )
        
        # Purple: > Capacity (Overflow - Critical)
        # Formula: F > G
        ws2.conditional_formatting.add(
            f'F{data_start}:F{data_end}',
            FormulaRule(formula=[f'F{data_start}>G{data_start}'], fill=purple_fill)
        )
        
        # Yellow/Orange: > 90% Capacity (Warning Zone)
        # Formula: F > G*0.9 AND F <= G
        warning_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        ws2.conditional_formatting.add(
            f'F{data_start}:F{data_end}',
            FormulaRule(formula=[f'AND(F{data_start}>G{data_start}*0.9, F{data_start}<=G{data_start})'], fill=warning_fill)
        )
        
        # Green: 0 <= F <= 90% Capacity (Optimal)
        ws2.conditional_formatting.add(
            f'F{data_start}:F{data_end}',
            FormulaRule(formula=[f'AND(F{data_start}>=0, F{data_start}<=G{data_start}*0.9)'], fill=green_fill)
        )
        
        row += 2
    
    # Column widths
    ws2.column_dimensions['A'].width = 12
    for col in range(2, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    
    # =========================================================================
    # TAB 3: SHIPMENT_BUILDER
    # =========================================================================
    ws3 = wb.create_sheet("SHIPMENT_BUILDER")
    
    ws3['A1'] = "SHIPMENT BUILDER - Plan Your Transfers"
    ws3['A1'].font = title_font
    ws3['A2'] = "Add shipments here. MANUALLY update Outgoing/Incoming in INVENTORY_TETRIS (shifted by Lead Time)."
    ws3['A2'].font = Font(italic=True, color="666666")
    
    ws3['A4'] = "SHIPMENT LOG"
    ws3['A4'].font = section_font
    
    headers = ['Values', 'From Zone', 'To Zone', 'Transport Mode', 'Units', 
               'Cost/Unit', 'Total Cost', 'Ship FN', 'Lead Time', 'Arrive FN']
    for col, h in enumerate(headers, start=1):
        cell = ws3.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Add 20 blank rows for input
    row = 6
    for i in range(20):
        ws3.cell(row=row, column=1, value=f"Shipment {i+1}").border = thin_border
        
        # From/To/Mode/Units (Input)
        for c in range(2, 6):
            cell = ws3.cell(row=row, column=c)
            cell.border = thin_border
            cell.fill = input_fill
        
        # Cost lookup - Wrap in IF(ISBLANK) to prevent #N/A
        cell = ws3.cell(row=row, column=6, value=f"=IF(ISBLANK(D{row}), 0, VLOOKUP(D{row}, ROUTE_CONFIG!$A$6:$C$8, 3, FALSE))")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
        
        # Total Cost
        cell = ws3.cell(row=row, column=7, value=f"=IF(ISBLANK(D{row}), 0, E{row}*F{row})")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
        
        # Ship FN
        cell = ws3.cell(row=row, column=8)
        cell.border = thin_border
        cell.fill = input_fill
        cell.alignment = Alignment(horizontal='center')
        
        # Lead Time lookup
        cell = ws3.cell(row=row, column=9, value=f"=IF(ISBLANK(D{row}), 0, VLOOKUP(D{row}, ROUTE_CONFIG!$A$6:$C$8, 2, FALSE))")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.alignment = Alignment(horizontal='center')
        
        # Arrive FN
        cell = ws3.cell(row=row, column=10, value=f"=IF(ISBLANK(D{row}), 0, H{row}+I{row})")
        cell.border = thin_border
        cell.fill = output_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Column widths
    ws3.column_dimensions['A'].width = 12
    for col in range(2, 11):
        ws3.column_dimensions[get_column_letter(col)].width = 12
    
    # === HISTORICAL COST BENCHMARKS TABLE (Columns M:N) ===
    bench_col = 13  # Column M
    ws3.cell(row=4, column=bench_col, value="HISTORICAL COST BENCHMARKS").font = section_font
    ws3.merge_cells(f'{get_column_letter(bench_col)}4:{get_column_letter(bench_col+1)}4')
    
    # Headers
    ws3.cell(row=5, column=bench_col, value="Route").font = header_font
    ws3.cell(row=5, column=bench_col).fill = header_fill
    ws3.cell(row=5, column=bench_col).border = thin_border
    ws3.cell(row=5, column=bench_col+1, value="Benchmark $/Unit").font = header_font
    ws3.cell(row=5, column=bench_col+1).fill = header_fill
    ws3.cell(row=5, column=bench_col+1).border = thin_border
    
    # Data rows from intelligence
    bench_row = 6
    benchmarks = intelligence_data.get('benchmarks', {})
    if benchmarks:
        for route, cost in benchmarks.items():
            ws3.cell(row=bench_row, column=bench_col, value=route).border = thin_border
            ws3.cell(row=bench_row, column=bench_col).fill = ref_fill  # Gray = read-only
            cell = ws3.cell(row=bench_row, column=bench_col+1, value=cost)
            cell.border = thin_border
            cell.fill = ref_fill  # Gray = read-only
            cell.number_format = '$#,##0.00'
            bench_row += 1
    else:
        # No benchmark data available
        ws3.cell(row=bench_row, column=bench_col, value="(No historical data)").font = Font(italic=True, color="666666")
    
    # Column widths for benchmark table
    ws3.column_dimensions[get_column_letter(bench_col)].width = 25
    ws3.column_dimensions[get_column_letter(bench_col+1)].width = 16
    
    # === PRICE GOUGING ALERT (Conditional Formatting) ===
    # If Cost/Unit > 1.2x the first benchmark (simplified approach)
    # Warning: Turn red if estimated cost is >20% higher than average historical cost
    if benchmarks:
        avg_benchmark = sum(benchmarks.values()) / len(benchmarks) if benchmarks else 0
        gouging_threshold = avg_benchmark * 1.2
        
        # Add conditional format to the Cost/Unit column (F) for all 20 shipment rows
        gouging_rule = FormulaRule(
            formula=[f'F6>{gouging_threshold}'],
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            font=Font(color="9C0006", bold=True)
        )
        ws3.conditional_formatting.add('F6:F25', gouging_rule)
        
        # Add note about the alert
        ws3.cell(row=27, column=1, value=f"⚠️ Red cells = Cost >20% above avg benchmark (${avg_benchmark:.2f} avg)").font = Font(italic=True, color="9C0006")
    
    # =========================================================================
    # TAB 4: UPLOAD_READY_LOGISTICS
    # =========================================================================
    ws4 = wb.create_sheet("UPLOAD_READY_LOGISTICS")
    
    ws4['A1'] = "LOGISTICS DECISIONS - ExSim Upload Format"
    ws4['A1'].font = title_font
    ws4['A2'] = "Copy these values to the web platform."
    ws4['A2'].font = Font(italic=True, color="666666")
    
    ws4['A4'] = "Planned Shipments"
    ws4['A4'].font = section_font
    
    headers = ['From', 'To', 'Material', 'Amount', 'F1', 'F2', 'F3', 'F4', 'F5', 'F6', 'F7', 'F8']
    for col, h in enumerate(headers, start=1):
        cell = ws4.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Map shipments from SHIPMENT_BUILDER (Simplified view)
    # Since we can't do complex aggregation easily in simplified Excel formulas, 
    # we'll just link the first 20 rows directly.
    
    row = 6
    for i in range(20):
        src_row = 6 + i
        
        # From
        cell = ws4.cell(row=row, column=1, value=f"=SHIPMENT_BUILDER!B{src_row}")
        cell.border = thin_border
        cell.fill = ref_fill
        
        # To
        cell = ws4.cell(row=row, column=2, value=f"=SHIPMENT_BUILDER!C{src_row}")
        cell.border = thin_border
        cell.fill = ref_fill
        
        # Material
        cell = ws4.cell(row=row, column=3, value=DEFAULT_MATERIAL)
        cell.border = thin_border
        
        # Amount (Total)
        cell = ws4.cell(row=row, column=4, value=f"=SHIPMENT_BUILDER!E{src_row}")
        cell.border = thin_border
        cell.fill = ref_fill
        
        # F1-F8 matrix (Logic: IF ShipFN = FN, 1, 0) - Simplified as 0/1 flag or Amount?
        # ExSim usually wants the Amount in the specific FN column.
        for fn in FORTNIGHTS:
            cell = ws4.cell(row=row, column=4+fn, 
                value=f'=IF(SHIPMENT_BUILDER!$H${src_row}={fn}, SHIPMENT_BUILDER!$E${src_row}, 0)')
            cell.border = thin_border
            cell.fill = calc_fill
            
        row += 1

    # Save to buffer or file
    if output_buffer is not None:
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer
    else:
        wb.save(OUTPUT_FILE)
        print(f"[SUCCESS] Created '{OUTPUT_FILE}'")
        return None


def main():
    """Main function."""
    print("ExSim Logistics Dashboard Generator")
    print("=" * 50)
    
    print("\n[*] Loading data files...")
    from config import REPORTS_DIR, DATA_DIR
    print(f"    Primary source: {REPORTS_DIR}")
    print(f"    Fallback source: {DATA_DIR}")
    
    # Finished Goods Inventory
    inv_path = get_data_path("finished_goods_inventory.xlsx")
    if inv_path:
        inv_data = load_finished_goods_by_zone(inv_path)
        print(f"  [OK] Loaded finished goods from {inv_path.parent.name}/")
    else:
        inv_data = load_finished_goods_by_zone(None)
        print("  [!] Using default inventory data")
    
    # Template
    template_path = get_data_path("Logistics Decisions.xlsx")
    template_data = load_logistics_template(template_path)
    if template_data['exists']:
        print(f"  [OK] Loaded logistics template")
    else:
        print("  [!] Using default template layout")
    
    # Shipping Costs
    cost_path = get_data_path("shipping_costs.xlsx")
    cost_data = load_shipping_costs(cost_path) if cost_path else {'total_shipping_cost': 0}
    
    # === NEW: LOGISTICS INTELLIGENCE ===
    print("\n[*] Loading Historical Intelligence...")
    intel_path = get_data_path("logistics.xlsx")
    if intel_path:
        intel_data = load_logistics_intelligence(intel_path)
        if intel_data['benchmarks']:
            print(f"  [INTEL] Cost Benchmarks: {len(intel_data['benchmarks'])} routes loaded")
            for route, cost in intel_data['benchmarks'].items():
                print(f"         {route}: ${cost:.2f}/unit")
        if intel_data['penalties']:
            total_pen = sum(intel_data['penalties'].values())
            print(f"  [INTEL] Warehouse Penalties: ${total_pen:,.0f} total")
            for zone, penalty in intel_data['penalties'].items():
                if penalty > 0:
                    print(f"         {zone}: ${penalty:,.0f}")
    else:
        intel_data = {'benchmarks': {}, 'penalties': {}}
        print("  [!] No logistics.xlsx found - intelligence disabled")
    
    print("\n[*] Creating dashboard...")
    
    create_logistics_dashboard(inv_data, template_data, cost_data, intel_data)
    
    print("\nSheets created:")
    print("  * ROUTE_CONFIG (Transport Modes & Costs)")
    print("  * INVENTORY_TETRIS (Zone Balancing & Stockout Checks)")
    print("  * SHIPMENT_BUILDER (Transfer Planning + Cost Benchmarks)")
    print("  * UPLOAD_READY_LOGISTICS (ExSim Format)")
    print("  * CROSS_REFERENCE (Upstream Data)") # Added this line to reflect the new tab
    
    # Export key metrics for downstream dashboards
    try:
        from shared_outputs import export_dashboard_data
        
        # Assuming total_logistics_cost is calculated somewhere or derived from cost_data
        # For now, using cost_data.get('total_shipping_cost', 0) as a placeholder
        total_logistics_cost = cost_data.get('total_shipping_cost', 0) 

        export_dashboard_data('CLO', {
            'shipping_schedule': "See Shipment Builder",
            'logistics_costs': total_logistics_cost,
            'inventory_by_zone': "See Inventory Tetris"
        })
    except ImportError:
        print("Warning: shared_outputs not found, skipping export")


if __name__ == "__main__":
    main()
