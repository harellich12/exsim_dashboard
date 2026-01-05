"""
ExSim CMO Complete Dashboard - Market Allocation & Strategy

Integrates Marketing Decisions, Innovation Decisions, Inventory Checks,
and Segment Analysis into a single cohesive decision-support tool.

Required libraries: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, FormulaRule, IconSetRule
from openpyxl.chart import ScatterChart, BarChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabelList
import warnings
import re
import sys

# Add parent directory to path to import case_parameters
# Add parent directory to path to import case_parameters
sys.path.append(str(Path(__file__).parent.parent))
try:
    from case_parameters import MARKET, COMMON
    from config import get_data_path, OUTPUT_DIR
except ImportError:
    print("Warning: Could not import case_parameters.py or config.py. Using defaults.")
    MARKET = {}
    COMMON = {}  # Add missing COMMON init
    # Fallback for config
    OUTPUT_DIR = Path(__file__).parent
    def get_data_path(f): return Path(f)

# Import shared outputs for inter-dashboard communication
try:
    from shared_outputs import export_dashboard_data
except ImportError:
    export_dashboard_data = None

warnings.filterwarnings('ignore')

# =============================================================================
# CONFIGURATION
# =============================================================================

# Required input files from centralized Reports folder
REQUIRED_FILES = [
    'market-report.xlsx',
    'finished_goods_inventory.xlsx',
    'sales_admin_expenses.xlsx'
]

OUTPUT_FILE = OUTPUT_DIR / "CMO_Dashboard_Complete.xlsx"

# Use centralized constants from case_parameters
MY_COMPANY = COMMON.get('MY_COMPANY', "Company 3")
ZONES = COMMON.get('ZONES', ["Center", "West", "North", "East", "South"])
SEGMENTS = COMMON.get('SEGMENTS', ["High", "Low"])

# Defaults - set to 0 to ensure data comes only from Excel files
DEFAULT_PRICE = 0
DEFAULT_AWARENESS = 0
DEFAULT_ATTRACTIVENESS = 0
DEFAULT_COGS = 0
DEFAULT_SALESPEOPLE_SALARY = 0


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def parse_numeric(value):
    """Parse formatted number strings."""
    if pd.isna(value):
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('$', '').replace(',', '').replace('%', '').replace(' ', '').strip()
    try:
        return float(cleaned)
    except:
        return 0


def load_excel_file(filepath, sheet_name=None):
    """Load Excel file, optionally from specific sheet."""
    try:
        if sheet_name:
            return pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        return pd.read_excel(filepath, header=None)
    except Exception as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


# =============================================================================
# DATA LOADING FUNCTIONS
# =============================================================================

def load_market_report(filepath):
    """Load market report with segment-level data from website export format."""
    df = load_excel_file(filepath)
    
    data = {
        'by_segment': {seg: {zone: {
            'my_market_share': 0,
            'my_awareness': 0,
            'my_attractiveness': 0,
            'my_price': 0,
            'comp_avg_awareness': 0,
            'comp_avg_price': 0
        } for zone in ZONES} for seg in SEGMENTS},
        'zones': {zone: {
            'my_price': 0,
            'comp_avg_price': 0,
            'my_awareness': 0,
            'my_attractiveness': 0,
            'my_market_share': 0
        } for zone in ZONES}
    }
    
    if df is None:
        return data
    
    current_section = None
    my_company_col = None  # Column index for MY_COMPANY
    comp_cols = []  # Column indices for competitors
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        second_val = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
        
        # Detect section headers
        if 'market share' in first_val.lower() and 'segment' in first_val.lower():
            current_section = 'segment_share'
            my_company_col = None
        elif 'market share' in first_val.lower() and 'region' in first_val.lower():
            current_section = 'region_share'
            my_company_col = None
        elif 'awareness' in first_val.lower() and 'segment' in first_val.lower():
            current_section = 'segment_awareness'
            my_company_col = None
        elif 'attractiveness' in first_val.lower():
            current_section = 'attractiveness'
            my_company_col = None
        elif 'price' in first_val.lower() and 'zone' not in first_val.lower():
            current_section = 'price'
            my_company_col = None
        
        # Detect column headers with company names
        if first_val.lower() == 'zone' and my_company_col is None:
            comp_cols = []
            for col_idx in range(len(row)):
                col_val = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                # Find Company 3 column (matches "Company 3 A" or "Company 3")
                if 'Company 3' in col_val or MY_COMPANY in col_val:
                    my_company_col = col_idx
                elif 'Company' in col_val and 'Company 3' not in col_val:
                    comp_cols.append(col_idx)
        
        # Parse zone data rows
        current_zone = None
        current_segment = None
        
        for zone in ZONES:
            if first_val.lower() == zone.lower():
                current_zone = zone
                break
        
        # Also check continuation rows (zone is blank, segment in second column)
        if first_val == '' and second_val.lower() in ['high', 'low']:
            current_segment = second_val.capitalize()
        elif current_zone and second_val.lower() in ['high', 'low']:
            current_segment = second_val.capitalize()
        
        if (current_zone or (first_val == '' and second_val.lower() in ['high', 'low'])) and my_company_col:
            try:
                my_val = parse_numeric(row.iloc[my_company_col])
                
                # Get competitor average
                comp_vals = [parse_numeric(row.iloc[c]) for c in comp_cols if c < len(row)]
                comp_vals = [v for v in comp_vals if v > 0]
                comp_avg = sum(comp_vals) / len(comp_vals) if comp_vals else 0
                
                # Determine which zone to use for continuation rows
                if current_zone:
                    zone_to_use = current_zone
                else:
                    # Find last zone from previous rows
                    zone_to_use = None
                    for prev_idx in range(idx-1, -1, -1):
                        prev_first = str(df.iloc[prev_idx, 0]).strip() if pd.notna(df.iloc[prev_idx, 0]) else ''
                        for z in ZONES:
                            if prev_first.lower() == z.lower():
                                zone_to_use = z
                                break
                        if zone_to_use:
                            break
                
                if zone_to_use is None:
                    continue
                
                # Store data based on section
                if current_section == 'region_share' and my_val > 0:
                    data['zones'][zone_to_use]['my_market_share'] = my_val
                    
                elif current_section == 'segment_share' and current_segment and my_val > 0:
                    data['by_segment'][current_segment][zone_to_use]['my_market_share'] = my_val
                    
                elif current_section == 'price' and my_val > 0:
                    data['zones'][zone_to_use]['my_price'] = my_val
                    if comp_avg > 0:
                        data['zones'][zone_to_use]['comp_avg_price'] = comp_avg
                    for seg in SEGMENTS:
                        data['by_segment'][seg][zone_to_use]['my_price'] = my_val
                        if comp_avg > 0:
                            data['by_segment'][seg][zone_to_use]['comp_avg_price'] = comp_avg
                            
                elif current_section == 'segment_awareness' and current_segment and my_val > 0:
                    data['zones'][zone_to_use]['my_awareness'] = my_val
                    if comp_avg > 0:
                        data['zones'][zone_to_use]['comp_avg_awareness'] = comp_avg
                    data['by_segment'][current_segment][zone_to_use]['my_awareness'] = my_val
                    if comp_avg > 0:
                        data['by_segment'][current_segment][zone_to_use]['comp_avg_awareness'] = comp_avg
                        
                elif current_section == 'attractiveness' and current_segment and my_val > 0:
                    data['zones'][zone_to_use]['my_attractiveness'] = my_val
                    data['by_segment'][current_segment][zone_to_use]['my_attractiveness'] = my_val
                    
            except Exception as e:
                continue
    
    return data


def load_innovation_features(filepath):
    """Load innovation features dynamically."""
    df = load_excel_file(filepath, sheet_name='Innovation')
    
    features = []
    
    if df is None:
        # Default features
        features = [
            "STAINLESS MATERIAL", "RECYCLABLE MATERIALS", "ENERGY EFFICIENCY",
            "LIGHTER AND MORE COMPACT", "IMPACT RESISTANCE", "NOISE REDUCTION",
            "IMPROVED BATTERY CAPACITY", "SELF-CLEANING", "SPEED SETTINGS",
            "DIGITAL CONTROLS", "VOICE ASSISTANCE INTEGRATION",
            "AUTOMATION AND PROGRAMMABILITY", "MULTIFUNCTIONAL ACCESSORIES",
            "MAPPING TECHNOLOGY"
        ]
    else:
        for idx, row in df.iterrows():
            # Look for Improvement column data
            if len(row) > 1:
                improvement = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                if improvement and improvement.upper() not in ['IMPROVEMENT', 'NAN', '']:
                    features.append(improvement)
    
    return features


def get_innovation_cost(feature_name):
    """Get cost dict for a feature from case parameters."""
    # Normalize name for lookup (uppercase, strip)
    name = feature_name.upper().strip()
    costs = MARKET.get("INNOVATION_COSTS", {})
    
    # Try direct match
    if name in costs:
        return costs[name]
    
    # Try partial match
    for key, val in costs.items():
        if key in name or name in key:
            return val
            
    return {"upfront": 0, "variable": 0}


def load_marketing_template(filepath):
    """Load marketing template structure."""
    df = load_excel_file(filepath, sheet_name='Marketing')
    
    template = {
        'df': df,
        'tv_budget': 0,
        'brand_focus': 0,
        'radio_budgets': {zone: 0 for zone in ZONES},
        'demand': {zone: 0 for zone in ZONES},
        'prices': {zone: 0 for zone in ZONES},
        'payment_terms': {zone: '' for zone in ZONES},
        'salespeople': {zone: 0 for zone in ZONES}
    }
    
    if df is not None:
        for idx, row in df.iterrows():
            first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
            
            # TV budget
            if first_val.upper() == 'A' and len(row) > 2:
                channel = str(row.iloc[2]).strip().lower() if pd.notna(row.iloc[2]) else ''
                if channel == 'tv':
                    template['tv_budget'] = parse_numeric(row.iloc[3])
                    template['brand_focus'] = parse_numeric(row.iloc[4])
                elif channel == 'radio':
                    zone = str(row.iloc[1]).strip()
                    if zone in ZONES:
                        template['radio_budgets'][zone] = parse_numeric(row.iloc[3])
            
            # Demand (column 7)
            for zone in ZONES:
                zone_val = str(row.iloc[7]).strip() if len(row) > 7 and pd.notna(row.iloc[7]) else ''
                if zone_val.lower() == zone.lower():
                    template['demand'][zone] = parse_numeric(row.iloc[8]) if len(row) > 8 else 0
    
    return template


def load_sales_admin_expenses(filepath):
    """Load sales and expenses data from website export format."""
    df = load_excel_file(filepath)
    
    data = {
        'by_zone': {zone: {'units': 0, 'price': DEFAULT_PRICE} for zone in ZONES},
        'totals': {'units': 0, 'tv_spend': 0, 'radio_spend': 0, 'salespeople_cost': 0}
    }
    
    if df is None:
        return data
    
    in_sales_section = False
    in_expense_section = False
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        
        # Detect sections
        if 'sales' in first_val.lower() and 'expense' not in first_val.lower() and 'admin' not in first_val.lower():
            in_sales_section = True
            in_expense_section = False
            continue
        elif 'expense' in first_val.lower() or 'admin' in first_val.lower():
            in_sales_section = False
            in_expense_section = True
            continue
        
        # Parse sales data rows (Region, Brand, Units, Local Price, ...)
        if in_sales_section:
            for zone in ZONES:
                if first_val.lower() == zone.lower():
                    # Format: Region, Brand, Units, Local Price, Gross Sales, Discount %, Net Sales
                    units = parse_numeric(row.iloc[2]) if len(row) > 2 else 0
                    price = parse_numeric(row.iloc[3]) if len(row) > 3 else DEFAULT_PRICE
                    
                    if units > 0:
                        data['by_zone'][zone]['units'] = units
                        data['totals']['units'] += units
                    if price > 0:
                        data['by_zone'][zone]['price'] = price
                    break
        
        # Parse expense data rows
        if in_expense_section:
            expense = parse_numeric(row.iloc[2]) if len(row) > 2 else 0
            if 'tv' in first_val.lower() and 'advert' in first_val.lower():
                data['totals']['tv_spend'] = expense
            elif 'radio' in first_val.lower() and 'advert' in first_val.lower():
                data['totals']['radio_spend'] = expense
            elif 'salespeople' in first_val.lower() and 'salar' in first_val.lower():
                data['totals']['salespeople_cost'] = expense
    
    return data


def load_finished_goods_inventory(filepath):
    """Load inventory to detect stockouts from website export format."""
    df = load_excel_file(filepath)
    
    data = {
        'final_inventory': 0, 
        'is_stockout': False,
        'by_zone': {zone: {'final': 0, 'capacity': 0} for zone in ZONES}
    }
    
    if df is None:
        return data
    
    current_zone_idx = 0
    zone_order = ['Center', 'West', 'North', 'East', 'South']
    
    for idx, row in df.iterrows():
        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        
        # Detect capacity header (start of new zone)
        if first_val.startswith('Capacity:'):
            # Extract capacity number
            capacity = parse_numeric(first_val.replace('Capacity:', ''))
            if current_zone_idx < len(zone_order):
                zone = zone_order[current_zone_idx]
                data['by_zone'][zone]['capacity'] = capacity
        
        # Parse final inventory row
        if 'final' in first_val.lower() and 'inventory' in first_val.lower():
            # Get fortnight 8 value (column 9, 0-indexed = column 9)
            final_val = parse_numeric(row.iloc[9]) if len(row) > 9 else 0
            
            if current_zone_idx < len(zone_order):
                zone = zone_order[current_zone_idx]
                data['by_zone'][zone]['final'] = final_val
            
            # Add to total
            data['final_inventory'] += final_val
            if final_val <= 0:
                data['is_stockout'] = True
            
            current_zone_idx += 1
    
    return data


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def load_marketing_intelligence(filepath_sales, filepath_market):
    """
    Load Unit Economics and Competitor Intelligence.
    1. Sales Report -> TV/Radio Cost per Spot, Hiring Fees.
    2. Market Report -> Competitor Pricing per Zone.
    """
    intelligence = {
        'economics': {
            'TV_Cost_Spot': 3000,   # Fallback
            'Radio_Cost_Spot': 300, # Fallback
            'Salary_Per_Person': 1500, # Fallback
            'Hiring_Cost': 1100     # Fallback
        },
        'pricing': {zone: 0 for zone in ZONES}
    }

    # 1. Parse Economics from Sales Report
    if filepath_sales and filepath_sales.exists():
        try:
            df = pd.read_excel(filepath_sales, header=None)
            
            # Scan for costs
            tv_amount = 0; tv_spots = 0
            radio_amount = 0; radio_spots = 0
            salaries = 0; headcount = 0
            hiring_cost = 0; hires = 0
            
            for idx, row in df.iterrows():
                label = str(row.iloc[0]).strip().lower() if pd.notna(row.iloc[0]) else ""
                val_col = 2 # Usually column C
                
                if "tv advertising expenses" in label:
                    tv_amount = parse_numeric(row.iloc[val_col])
                    details = str(row.iloc[1])
                    match = re.search(r'(\d+)\s*spots', details, re.IGNORECASE)
                    if match:
                        tv_spots = int(match.group(1))
                
                if "radio advertising expenses" in label:
                    radio_amount = parse_numeric(row.iloc[val_col])
                    details = str(row.iloc[1])
                    match = re.search(r'(\d+)\s*spots', details, re.IGNORECASE)
                    if match:
                        radio_spots = int(match.group(1))
                        
                if "salespeople salaries" in label:
                    salaries = parse_numeric(row.iloc[val_col])
                    details = str(row.iloc[1])
                    match = re.search(r'(\d+)\s*people', details, re.IGNORECASE)
                    if match:
                        headcount = int(match.group(1))

                if "salespeople hiring" in label:
                    hiring_cost = parse_numeric(row.iloc[val_col])
                    details = str(row.iloc[1])
                    match = re.search(r'(\d+)\s*hires', details, re.IGNORECASE)
                    if match:
                        hires = int(match.group(1))
            
            # Calculate Rates
            if tv_spots > 0:
                intelligence['economics']['TV_Cost_Spot'] = tv_amount / tv_spots
            if radio_spots > 0:
                intelligence['economics']['Radio_Cost_Spot'] = radio_amount / radio_spots
            if headcount > 0:
                intelligence['economics']['Salary_Per_Person'] = salaries / headcount
            if hires > 0:
                intelligence['economics']['Hiring_Cost'] = hiring_cost / hires
                
        except Exception as e:
            print(f"Warning: Could not load economics: {e}")

    # 2. Parse Competitor Pricing from Market Report
    if filepath_market and filepath_market.exists():
        market_data = load_market_report(filepath_market) # Reuse existing parser
        for zone in ZONES:
            intelligence['pricing'][zone] = market_data['zones'][zone].get('comp_avg_price', 0)

    return intelligence

def create_complete_dashboard(market_data, innovation_features, marketing_template, 
                               sales_data, inventory_data, marketing_intelligence, output_buffer=None, decision_overrides=None):
    """Create the complete 5-tab CMO Dashboard."""
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    section_font = Font(bold=True, size=12, color="2F5496")
    title_font = Font(bold=True, size=14, color="2F5496")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    output_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ref_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Seasonality Alert Logic
    current_period = 1 # Default, ideally should be parsed from reports
    next_period = current_period + 1
    peaks = MARKET.get('SEASONALITY_PEAKS', [])
    seasonality_msg = "Normal Demand Expected"
    seasonality_color = output_fill
    
    if next_period in peaks:
        seasonality_msg = f"WARNING: Period {next_period} is a PEAK SEASON! Increase Inventory."
        seasonality_color = red_fill
    elif current_period in peaks:
        seasonality_msg = "Currently in Peak Season. Monitor stockouts closely."
        seasonality_color = orange_fill

    
    # Extract Economics
    econ = marketing_intelligence.get('economics', {})
    tv_cost_spot = econ.get('TV_Cost_Spot', 3000)
    radio_cost_spot = econ.get('Radio_Cost_Spot', 300)
    salary_per_person = econ.get('Salary_Per_Person', 1500)
    hiring_cost = econ.get('Hiring_Cost', 1100)
    
    comp_pricing = marketing_intelligence.get('pricing', {})
    
    # =========================================================================
    # TAB 1: SEGMENT_PULSE
    # =========================================================================
    ws1 = wb.active
    ws1.title = "SEGMENT_PULSE"
    
    ws1['A1'] = "SEGMENT PULSE - Market Allocation Drivers"
    ws1['A1'].font = title_font
    
    # Add Seasonality Alert at top
    ws1['E1'] = seasonality_msg
    ws1['E1'].font = Font(bold=True, color="9C0006" if next_period in peaks else "006100")
    ws1['E1'].fill = seasonality_color

    
    row = 3
    for segment in SEGMENTS:
        ws1.cell(row=row, column=1, value=f"{segment.upper()} SEGMENT ANALYSIS").font = section_font
        row += 1
        
        # Headers
        seg_headers = ['Zone', 'My Market Share', 'Est. Demand', 'Awareness Gap', 'Price Gap', 
                       'Attractiveness', 'Allocation Flag']
        for col, header in enumerate(seg_headers, start=1):
            cell = ws1.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        
        data_start_row = row
        
        # Get Population Data (TAM)
        pop_data = MARKET.get('POPULATION', {})

        
        for zone in ZONES:
            zone_seg = market_data['by_segment'][segment].get(zone, {})
            zone_data = market_data['zones'].get(zone, {})
            
            market_share = zone_seg.get('my_market_share', 0)
            
            # Calculate Penetration/Volume Estimate
            zone_pop = pop_data.get(zone, {}).get(segment, 10000) # Default 10k if missing
            est_units_sold = zone_pop * (market_share / 100)
            
            # FIX: Prefer segment-specific data over zone-level data (matching UI behavior)
            my_awareness = zone_seg.get('my_awareness', zone_data.get('my_awareness', DEFAULT_AWARENESS))
            comp_awareness = zone_seg.get('comp_avg_awareness', zone_data.get('comp_avg_awareness', DEFAULT_AWARENESS))
            awareness_gap = my_awareness - comp_awareness
            
            my_price = zone_seg.get('my_price', zone_data.get('my_price', DEFAULT_PRICE))
            comp_price = zone_seg.get('comp_avg_price', zone_data.get('comp_avg_price', DEFAULT_PRICE))
            price_gap = ((my_price - comp_price) / comp_price * 100) if comp_price > 0 else 0
            
            # FIX: Prefer segment-specific attractiveness (matching UI behavior)
            attractiveness = zone_seg.get('my_attractiveness', zone_data.get('my_attractiveness', DEFAULT_ATTRACTIVENESS))

            
            # Allocation flag logic
            # First check for zones with no market presence
            if market_share == 0 or my_awareness == 0:
                flag = "NO PRESENCE: Zone Not Active"
                flag_fill = ref_fill  # Gray = needs investigation
            elif segment == "High":
                if my_awareness < 30:
                    flag = "CRITICAL: Boost TV for Allocation"
                    flag_fill = red_fill
                else:
                    flag = "OK"
                    flag_fill = output_fill
            else:  # Low segment
                if price_gap > 5:
                    flag = "RISK: Losing Volume to Price"
                    flag_fill = orange_fill
                else:
                    flag = "OK"
                    flag_fill = output_fill
            
            ws1.cell(row=row, column=1, value=zone).border = thin_border
            
            cell = ws1.cell(row=row, column=2, value=market_share)
            cell.border = thin_border
            cell.number_format = '0.0%' if market_share <= 1 else '0.0'
            
            cell = ws1.cell(row=row, column=3, value=est_units_sold)
            cell.border = thin_border
            cell.number_format = '#,##0'
            
            cell = ws1.cell(row=row, column=4, value=awareness_gap)
            cell.border = thin_border
            if awareness_gap < 0:
                cell.fill = red_fill
            
            cell = ws1.cell(row=row, column=5, value=price_gap / 100)
            cell.border = thin_border
            cell.number_format = '0.0%'
            
            ws1.cell(row=row, column=6, value=attractiveness).border = thin_border
            
            cell = ws1.cell(row=row, column=7, value=flag)
            cell.border = thin_border
            cell.fill = flag_fill
            cell.font = Font(bold=True)
            
            row += 1
        
        # Add data bars for market share
        ws1.conditional_formatting.add(
            f'B{data_start_row}:B{row-1}',
            DataBarRule(start_type='num', start_value=0, end_type='num', end_value=50,
                       color="63C384", showValue=True, minLength=None, maxLength=None)
        )
        
        row += 2
    
    # Column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 32
    
    # =========================================================================
    # CHART DATA SECTION (Right side of sheet, starting column H)
    # =========================================================================
    
    # Calculate averages for charts
    high_awareness_avg = sum(market_data['zones'][z].get('my_awareness', DEFAULT_AWARENESS) for z in ZONES) / len(ZONES)
    low_awareness_avg = high_awareness_avg  # Same default for now
    
    my_price_avg = sum(market_data['zones'][z].get('my_price', DEFAULT_PRICE) for z in ZONES) / len(ZONES)
    comp_price_avg = sum(market_data['zones'][z].get('comp_avg_price', DEFAULT_PRICE) for z in ZONES) / len(ZONES)
    my_attract_avg = sum(market_data['zones'][z].get('my_attractiveness', DEFAULT_ATTRACTIVENESS) for z in ZONES) / len(ZONES)
    comp_attract_avg = DEFAULT_ATTRACTIVENESS  # Competitor default
    
    # High segment averages
    high_avg_awareness = sum(market_data['by_segment']['High'][z].get('my_awareness', DEFAULT_AWARENESS) 
                             if 'my_awareness' in market_data['by_segment']['High'][z] 
                             else DEFAULT_AWARENESS for z in ZONES) / len(ZONES)
    high_avg_price_gap = sum(((market_data['zones'][z].get('my_price', DEFAULT_PRICE) - 
                               market_data['zones'][z].get('comp_avg_price', DEFAULT_PRICE)) / 
                              max(1, market_data['zones'][z].get('comp_avg_price', DEFAULT_PRICE)) * 100) 
                             for z in ZONES) / len(ZONES)
    high_avg_attract = my_attract_avg
    
    # Low segment (using same zone data)
    low_avg_awareness = high_avg_awareness
    low_avg_price_gap = high_avg_price_gap
    low_avg_attract = my_attract_avg
    
    # ----- Chart 1: Competitive Positioning Matrix Data -----
    ws1['H1'] = "COMPETITIVE POSITIONING"
    ws1['H1'].font = section_font
    
    # Data table for scatter chart
    ws1['H3'] = "Entity"
    ws1['I3'] = "Price"
    ws1['J3'] = "Attractiveness"
    for c in ['H', 'I', 'J']:
        ws1[f'{c}3'].font = header_font
        ws1[f'{c}3'].fill = header_fill
        ws1[f'{c}3'].border = thin_border
    
    ws1['H4'] = "My Product"
    ws1['I4'] = my_price_avg
    ws1['J4'] = my_attract_avg
    ws1['I4'].number_format = '$#,##0'
    for c in ['H', 'I', 'J']:
        ws1[f'{c}4'].border = thin_border
    
    ws1['H5'] = "Competitors"
    ws1['I5'] = comp_price_avg
    ws1['J5'] = comp_attract_avg
    ws1['I5'].number_format = '$#,##0'
    for c in ['H', 'I', 'J']:
        ws1[f'{c}5'].border = thin_border
    
    # Create Scatter Chart
    chart1 = ScatterChart()
    chart1.title = "Competitive Positioning Matrix"
    chart1.x_axis.title = "Price ($)"
    chart1.y_axis.title = "Attractiveness Score"
    chart1.style = 13
    chart1.height = 10
    chart1.width = 12
    
    # My Product series
    x_values1 = Reference(ws1, min_col=9, min_row=4, max_row=4)
    y_values1 = Reference(ws1, min_col=10, min_row=4, max_row=4)
    series1 = Series(y_values1, x_values1, title="My Product")
    series1.marker = Marker(symbol='circle', size=12)
    series1.graphicalProperties.solidFill = "4472C4"  # Blue
    chart1.series.append(series1)
    
    # Competitor series
    x_values2 = Reference(ws1, min_col=9, min_row=5, max_row=5)
    y_values2 = Reference(ws1, min_col=10, min_row=5, max_row=5)
    series2 = Series(y_values2, x_values2, title="Competitors")
    series2.marker = Marker(symbol='diamond', size=12)
    series2.graphicalProperties.solidFill = "ED7D31"  # Orange
    chart1.series.append(series2)
    
    ws1.add_chart(chart1, "H7")
    
    # ----- Chart 2: High vs Low Segment Gap Data -----
    ws1['H24'] = "HIGH vs LOW SEGMENT GAP"
    ws1['H24'].font = section_font
    
    # Data table for bar chart
    ws1['H26'] = "Metric"
    ws1['I26'] = "High Segment"
    ws1['J26'] = "Low Segment"
    for c in ['H', 'I', 'J']:
        ws1[f'{c}26'].font = header_font
        ws1[f'{c}26'].fill = header_fill
        ws1[f'{c}26'].border = thin_border
    
    ws1['H27'] = "Awareness"
    ws1['I27'] = high_avg_awareness
    ws1['J27'] = low_avg_awareness
    for c in ['H', 'I', 'J']:
        ws1[f'{c}27'].border = thin_border
    
    ws1['H28'] = "Price Competitiveness"
    ws1['I28'] = 100 - abs(high_avg_price_gap)  # Higher = better
    ws1['J28'] = 100 - abs(low_avg_price_gap)
    for c in ['H', 'I', 'J']:
        ws1[f'{c}28'].border = thin_border
    
    ws1['H29'] = "Attractiveness"
    ws1['I29'] = high_avg_attract
    ws1['J29'] = low_avg_attract
    for c in ['H', 'I', 'J']:
        ws1[f'{c}29'].border = thin_border
    
    # Create Clustered Bar Chart
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "High vs Low Segment Comparison"
    chart2.style = 13
    chart2.height = 10
    chart2.width = 12
    
    # Data references
    categories = Reference(ws1, min_col=8, min_row=27, max_row=29)
    data = Reference(ws1, min_col=9, min_row=26, max_col=10, max_row=29)
    
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(categories)
    chart2.shape = 4
    
    ws1.add_chart(chart2, "H31")
    
    # ----- Traffic Light Conditional Formatting -----
    # Apply to awareness data in both High and Low segment sections
    # High segment awareness is in column C, rows 5-9 (approx)
    # Low segment awareness is in column C, rows 13-17 (approx)
    
    # Create icon set rule for awareness columns
    icon_rule = IconSetRule(
        icon_style='3TrafficLights1',
        type='num',
        values=[0, 40, 70],  # Red < 40, Yellow 40-70, Green > 70
        showValue=True,
        reverse=False
    )
    
    # Apply to High segment (rows 5-9, column D - Awareness Gap) 
    ws1.conditional_formatting.add('D5:D9', icon_rule)
    # Apply to Low segment (rows 13-17, column D)
    ws1.conditional_formatting.add('D13:D17', icon_rule)
    
    # Price Gap red text formatting (column E if > 10%)
    price_gap_rule = FormulaRule(
        formula=['E5>0.1'],
        font=Font(bold=True, color="9C0006")
    )
    ws1.conditional_formatting.add('E5:E9', price_gap_rule)
    ws1.conditional_formatting.add('E13:E17', price_gap_rule)
    
    # Additional column widths for chart data
    ws1.column_dimensions['H'].width = 18
    ws1.column_dimensions['I'].width = 14
    ws1.column_dimensions['J'].width = 14

    
    # =========================================================================
    # TAB 2: INNOVATION_LAB
    # =========================================================================
    ws2 = wb.create_sheet("INNOVATION_LAB")
    
    ws2['A1'] = "INNOVATION LAB - Feature Selection"
    ws2['A1'].font = title_font
    
    ws2['A2'] = "Note: Innovations increase Attractiveness. Required for High Segment Allocation."
    ws2['A2'].font = Font(italic=True, color="666666")
    
    # Headers
    innov_headers = ['Feature Name', 'Decision (1=Yes)', 'Est. Cost ($)']
    for col, header in enumerate(innov_headers, start=1):
        cell = ws2.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Dynamic feature list
    row = 5
    for feature in innovation_features:
        ws2.cell(row=row, column=1, value=feature).border = thin_border
        
        # Add Cost Calculation
        costs = get_innovation_cost(feature)
        upfront = costs.get('upfront', 0)
        variable = costs.get('variable', 0)
        
        cost_str = f"${upfront:,.0f} + ${variable:.2f}/unit"
        ws2.cell(row=row, column=3, value=cost_str).border = thin_border
        
        dec_val = 0
        if decision_overrides and 'innovation' in decision_overrides:
             dec_val = decision_overrides['innovation'].get(feature, 0)

        cell = ws2.cell(row=row, column=2, value=dec_val)
        cell.border = thin_border
        cell.fill = input_fill
        cell.alignment = Alignment(horizontal='center')
        
        cell = ws2.cell(row=row, column=3, value=10000)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        # New "Calculated Est. Cost" column
        cell = ws2.cell(row=row, column=4, value=f"=C{row}")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
        # Add Excel Comment
        # from openpyxl.comments import Comment
        # cell.comment = Comment("Approximate value based on case history. Actuals may vary by +/- 10%.", "ExSim")
        # Skipping Comment object import for simplicity, just keeping value.
        
        row += 1
    
    # Total innovation cost
    row += 1
    ws2.cell(row=row, column=1, value="TOTAL INNOVATION COST").font = Font(bold=True)
    cell = ws2.cell(row=row, column=4, value=f'=SUMPRODUCT(B5:B{row-2},C5:C{row-2})') # Update to verify against calc cost? or input? usually input.
    # Actually request says "Link to the static cost".
    # Let's keep sumproduct on Input Cost (Col C) or new Col D? 
    # Usually we pay based on the input cost we agree to, so Col C is fine.
    cell = ws2.cell(row=row, column=3, value=f'=SUMPRODUCT(B5:B{row-2},C5:C{row-2})')
    cell.fill = calc_fill
    cell.font = Font(bold=True)
    cell.number_format = '$#,##0'
    
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 18
    
    ws2.cell(row=4, column=4, value="Calculated Est. Cost").font = header_font
    ws2.cell(row=4, column=4).fill = header_fill
    ws2.cell(row=4, column=4).border = thin_border
    
    innov_cost_cell = f'C{row}'
    
    # =========================================================================
    # TAB 3:    _COCKPIT
    # =========================================================================
    ws3 = wb.create_sheet("STRATEGY_COCKPIT")
    
    ws3['A1'] = "HOW TO USE: Adjust Yellow cells. Check Profit Projection. Go to UPLOAD_READY tabs to copy decisions."
    ws3['A1'].font = Font(italic=True, color="666666")
    
    # Section A: Global Allocations
    # UNIT ECONOMICS CHEAT SHEET (Rows 1-4)
    ws3.insert_rows(1, 4)
    ws3['A1'] = "UNIT ECONOMICS CHEAT SHEET"
    ws3['A1'].font = section_font
    
    ws3['A2'] = "TV Cost/Spot"
    ws3['B2'] = "Radio Cost/Spot"
    ws3['C2'] = "Hiring Fee"
    ws3['D2'] = "Salary/Person"
    
    ws3['A3'] = tv_cost_spot
    ws3['B3'] = radio_cost_spot
    ws3['C3'] = hiring_cost
    ws3['D3'] = salary_per_person
    
    for cell in ws3['A2:D2'][0]:
        cell.font = Font(bold=True, italic=True, color="666666")
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
    for cell in ws3['A3:D3'][0]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.number_format = '$#,##0'

    ws3['A5'] = "HOW TO USE: Adjust Yellow cells. Check Profit Projection. Go to UPLOAD_READY tabs to copy decisions."
    ws3['A5'].font = Font(italic=True, color="666666")

    ws3['A7'] = "SECTION A: GLOBAL ALLOCATIONS"
    ws3['A7'].font = section_font
    
    # TV SPOTS Input (Was Budget)
    ws3.cell(row=9, column=1, value="TV Spots (Qty)").border = thin_border
    # Convert budget to spots approx
    tv_spots_init = int(marketing_template['tv_budget'] / tv_cost_spot) if tv_cost_spot else 0
    
    if decision_overrides and 'tv_spots' in decision_overrides:
         tv_spots_init = decision_overrides['tv_spots']

    cell = ws3.cell(row=9, column=2, value=tv_spots_init)
    cell.border = thin_border
    cell.fill = input_fill
    # cell.number_format = '#,##0' # Spots are integers
    
    # TV Cost formula (actual Excel formula, not display text)
    cell = ws3.cell(row=9, column=3, value=f"=B9*{tv_cost_spot}")
    cell.fill = calc_fill
    cell.number_format = '$#,##0'
    cell.border = thin_border
    
    ws3.cell(row=10, column=1, value="Brand Focus (0-100)").border = thin_border
    brand_focus_val = marketing_template['brand_focus']
    if decision_overrides and 'brand_focus' in decision_overrides:
         brand_focus_val = decision_overrides['brand_focus']

    cell = ws3.cell(row=10, column=2, value=brand_focus_val)
    cell.border = thin_border
    cell.fill = input_fill
    ws3['C10'] = "0=Awareness focus, 100=Attributes focus"
    ws3['C10'].font = Font(italic=True, color="666666")
    
    # Section B: Zonal Allocations
    ws3['A13'] = "SECTION B: ZONAL ALLOCATIONS"
    ws3['A13'].font = section_font
    
    zonal_headers = ['Zone', 'Last Sales', 'Stockout?', 'Target Demand', 'Radio Spots (Qty)',
                     'Headcount', 'Price', 'Avg Comp Price', 'Payment', 'Est. Revenue', 'Mkt Cost', 'Contribution']
    
    for col, header in enumerate(zonal_headers, start=1):
        cell = ws3.cell(row=15, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row = 16
    for zone in ZONES:
        zone_sales = sales_data['by_zone'].get(zone, {})
        last_sales = zone_sales.get('units', 1000)
        is_stockout = inventory_data['is_stockout']
        
        ws3.cell(row=row, column=1, value=zone).border = thin_border
        
        # Reference data (gray)
        cell = ws3.cell(row=row, column=2, value=last_sales)
        cell.border = thin_border
        cell.fill = ref_fill
        cell.number_format = '#,##0'
        
        cell = ws3.cell(row=row, column=3, value="TRUE DEMAND HIGHER" if is_stockout else "OK")
        cell.border = thin_border
        if is_stockout:
            cell.fill = red_fill
            cell.font = Font(bold=True, color="9C0006")
        else:
            cell.fill = ref_fill
        
        # Target Demand
        dem_val = marketing_template['demand'].get(zone, 0)
        if decision_overrides and 'zones' in decision_overrides and zone in decision_overrides['zones']:
             dem_val = decision_overrides['zones'][zone].get('target_demand', dem_val)
        
        cell = ws3.cell(row=row, column=4, value=dem_val)
        cell.border = thin_border
        cell.fill = input_fill
        
        # Radio Spots (Qty)
        radio_bud = marketing_template['radio_budgets'].get(zone, 0)
        radio_spots_init = int(radio_bud / radio_cost_spot) if radio_cost_spot else 0
        
        if decision_overrides and 'zones' in decision_overrides and zone in decision_overrides['zones']:
             radio_spots_init = decision_overrides['zones'][zone].get('radio', radio_spots_init)
        
        cell = ws3.cell(row=row, column=5, value=radio_spots_init)
        cell.border = thin_border
        cell.fill = input_fill
        
        # Salespeople (Headcount)
        hc_val = marketing_template['salespeople'].get(zone, 0)
        if decision_overrides and 'zones' in decision_overrides and zone in decision_overrides['zones']:
             hc_val = decision_overrides['zones'][zone].get('salespeople', hc_val)
             
        cell = ws3.cell(row=row, column=6, value=hc_val)
        cell.border = thin_border
        cell.fill = input_fill
        
        # Price
        price_val = marketing_template['prices'].get(zone, 0)
        if decision_overrides and 'zones' in decision_overrides and zone in decision_overrides['zones']:
             price_val = decision_overrides['zones'][zone].get('price', price_val)
             
        cell = ws3.cell(row=row, column=7, value=price_val)
        cell.border = thin_border
        cell.fill = input_fill
        cell.number_format = '$#,##0'
        
        # Avg Comp Price (Reference)
        comp_price = comp_pricing.get(zone, 68.0)
        cell = ws3.cell(row=row, column=8, value=comp_price)
        cell.border = thin_border
        cell.fill = ref_fill
        cell.number_format = '$#,##0.00'
        
        # Payment Terms
        cell = ws3.cell(row=row, column=9, value=marketing_template['payment_terms'].get(zone, ''))
        cell.border = thin_border
        cell.fill = input_fill
        
        # Est Revenue
        cell = ws3.cell(row=row, column=10, value=f"=D{row}*G{row}")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
        
        # Mkt Cost (Simplified formula to avoid cross-sheet issues)
        # Components: TV (split evenly), Radio, Salaries, Hiring, Innovation
        # TV Cost = B9 * tv_cost_spot / 5 (split across 5 zones)
        # Radio Cost = E{row} * radio_cost_spot
        # Salary Cost = F{row} * salary_per_person
        # Hiring Cost = MAX(0, F{row} - prev_hc) * hiring_cost
        # Innovation = Total Innovation Cost / 5 (use hardcoded value from INNOVATION_LAB total)
        # Note: Simplified to avoid #VALUE! errors from cross-sheet refs that may not resolve
        prev_hc = 5  # Default previous headcount assumption
        cell = ws3.cell(row=row, column=11, 
            value=f"=(C9/5) + (E{row}*{radio_cost_spot}) + (F{row}*{salary_per_person}) + (MAX(0, F{row}-{prev_hc})*{hiring_cost})")
        cell.border = thin_border
        cell.fill = calc_fill
        cell.number_format = '$#,##0'
        
        # Contribution
        cell = ws3.cell(row=row, column=12, value=f"=J{row}-K{row}") 
        cell.border = thin_border
        cell.fill = output_fill
        cell.number_format = '$#,##0'
        
        # Conditional Formatting: Price Gouging
        ws3.conditional_formatting.add(
            f'G{row}',
            FormulaRule(formula=[f'G{row}>(H{row}*1.15)'], 
                        fill=red_fill, font=Font(color="FFFFFF", bold=True))
        )
        
        row += 1
    
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 14
    ws3.column_dimensions['G'].width = 12
    ws3.column_dimensions['H'].width = 14
    ws3.column_dimensions['I'].width = 14
    ws3.column_dimensions['J'].width = 14
    ws3.column_dimensions['K'].width = 16
    ws3.column_dimensions['L'].width = 16
    
    # =========================================================================
    # TAB 4: UPLOAD_READY_MARKETING
    # =========================================================================
    ws4 = wb.create_sheet("UPLOAD_READY_MARKETING")
    
    ws4['A1'] = "MARKETING DECISIONS - ExSim Upload Format"
    ws4['A1'].font = title_font
    ws4['A2'] = "Copy these values to ExSim Marketing upload"
    ws4['A2'].font = Font(italic=True, color="666666")
    
    # Recreate the side-by-side layout
    # Marketing Campaigns (cols A-E)
    ws4['A4'] = "Marketing Campaigns"
    ws4['A4'].font = section_font
    
    camp_headers = ['Brand', 'Zone', 'Channel', 'Amount', 'Brand Focus']
    for col, h in enumerate(camp_headers, start=1):
        cell = ws4.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # TV row
    ws4.cell(row=6, column=1, value='A').border = thin_border
    ws4.cell(row=6, column=2, value='All').border = thin_border
    ws4.cell(row=6, column=3, value='TV').border = thin_border
    ws4.cell(row=6, column=4, value='=STRATEGY_COCKPIT!C9').border = thin_border # C9 = Calculated Cost
    ws4.cell(row=6, column=5, value='=STRATEGY_COCKPIT!B10').border = thin_border
    
    # Radio rows
    row = 7
    # Cheat Sheet: Radio Cost/Spot is at B3 now (Row 2 is header)
    radio_cost_cell = "$B$3"
    
    for zone_idx, zone in enumerate(ZONES):
        ws4.cell(row=row, column=1, value='A').border = thin_border
        ws4.cell(row=row, column=2, value=zone).border = thin_border
        ws4.cell(row=row, column=3, value='Radio').border = thin_border
        # Radio Amount = Spots (E) * CostPerSpot (CheatSheet B2)
        # Zone rows start at 16
        source_row = 16 + zone_idx
        ws4.cell(row=row, column=4, value=f'=STRATEGY_COCKPIT!E{source_row}*STRATEGY_COCKPIT!{radio_cost_cell}').border = thin_border
        ws4.cell(row=row, column=5, value='=STRATEGY_COCKPIT!B10').border = thin_border
        row += 1
    
    # Demand section (cols G-H)
    ws4['G4'] = "Demand"
    ws4['G4'].font = section_font
    
    ws4.cell(row=5, column=7, value='Zone').font = header_font
    ws4.cell(row=5, column=7).fill = header_fill
    ws4.cell(row=5, column=8, value='Demand').font = header_font
    ws4.cell(row=5, column=8).fill = header_fill
    
    for zone_idx, zone in enumerate(ZONES):
        source_row = 16 + zone_idx
        ws4.cell(row=6+zone_idx, column=7, value=zone).border = thin_border
        ws4.cell(row=6+zone_idx, column=8, value=f'=STRATEGY_COCKPIT!D{source_row}').border = thin_border
    
    # Pricing section (cols J-L)
    ws4['J4'] = "Pricing Strategy"
    ws4['J4'].font = section_font
    
    ws4.cell(row=5, column=10, value='Zone').font = header_font
    ws4.cell(row=5, column=10).fill = header_fill
    ws4.cell(row=5, column=11, value='Brand').font = header_font
    ws4.cell(row=5, column=11).fill = header_fill
    ws4.cell(row=5, column=12, value='Price').font = header_font
    ws4.cell(row=5, column=12).fill = header_fill
    
    for zone_idx, zone in enumerate(ZONES):
        source_row = 16 + zone_idx
        ws4.cell(row=6+zone_idx, column=10, value=zone).border = thin_border
        ws4.cell(row=6+zone_idx, column=11, value='A').border = thin_border
        ws4.cell(row=6+zone_idx, column=12, value=f'=STRATEGY_COCKPIT!G{source_row}').border = thin_border
    
    # Channels section (cols N-P)
    ws4['N4'] = "Channels"
    ws4['N4'].font = section_font
    
    ws4.cell(row=5, column=14, value='Zone').font = header_font
    ws4.cell(row=5, column=14).fill = header_fill
    ws4.cell(row=5, column=15, value='Payment').font = header_font
    ws4.cell(row=5, column=15).fill = header_fill
    ws4.cell(row=5, column=16, value='Salespeople').font = header_font
    ws4.cell(row=5, column=16).fill = header_fill
    
    for zone_idx, zone in enumerate(ZONES):
        source_row = 16 + zone_idx
        ws4.cell(row=6+zone_idx, column=14, value=zone).border = thin_border
        # Payment is col 9 (I)
        ws4.cell(row=6+zone_idx, column=15, value=f'=STRATEGY_COCKPIT!I{source_row}').border = thin_border
        # Salespeople is col 6 (F)
        ws4.cell(row=6+zone_idx, column=16, value=f'=STRATEGY_COCKPIT!F{source_row}').border = thin_border
    
    # =========================================================================
    # TAB 5: UPLOAD_READY_INNOVATION
    # =========================================================================
    ws5 = wb.create_sheet("UPLOAD_READY_INNOVATION")
    
    ws5['A1'] = "INNOVATION DECISIONS - ExSim Upload Format"
    ws5['A1'].font = title_font
    ws5['A2'] = "Copy these values to ExSim Innovation upload"
    ws5['A2'].font = Font(italic=True, color="666666")
    
    # Headers
    ws5.cell(row=4, column=1, value='Brand').font = header_font
    ws5.cell(row=4, column=1).fill = header_fill
    ws5.cell(row=4, column=2, value='Improvement').font = header_font
    ws5.cell(row=4, column=2).fill = header_fill
    ws5.cell(row=4, column=3, value='Value').font = header_font
    ws5.cell(row=4, column=3).fill = header_fill
    
    for i, feature in enumerate(innovation_features):
        ws5.cell(row=5+i, column=1, value='A').border = thin_border
        ws5.cell(row=5+i, column=2, value=feature).border = thin_border
        # Link to INNOVATION_LAB decision
        ws5.cell(row=5+i, column=3, value=f'=INNOVATION_LAB!B{5+i}').border = thin_border
    
    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 35
    ws5.column_dimensions['C'].width = 10
    
    ws5.column_dimensions['A'].width = 10
    ws5.column_dimensions['B'].width = 35
    ws5.column_dimensions['C'].width = 10
    
    # =========================================================================
    # TAB 6: CROSS_REFERENCE (Upstream Data)
    # =========================================================================
    ws6 = wb.create_sheet("CROSS_REFERENCE")
    
    ws6['A1'] = "CROSS-REFERENCE SUMMARY - Upstream Support"
    ws6['A1'].font = title_font
    ws6['A2'] = "Key metrics from Production and Finance."
    ws6['A2'].font = Font(italic=True, color="666666")
    
    # Load shared data
    try:
        from shared_outputs import import_dashboard_data
        prod_data = import_dashboard_data('Production') or {}
        cfo_data = import_dashboard_data('CFO') or {}
    except ImportError:
        prod_data = {}
        cfo_data = {}
    
    row = 4
    
    # Production Section
    ws6.cell(row=row, column=1, value="Production (Capacity)").font = section_font
    ws6.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
    ws6.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    prod_metrics = [
        ("Production Plan (Total Units)", f"{sum([d.get('Target',0) for d in prod_data.get('production_plan', {}).values()]) if prod_data and 'production_plan' in prod_data else 'N/A'}"),
        ("Avg Capacity Utilization", f"{prod_data.get('capacity_utilization', {}).get('mean', 0)*100:.1f}%" if prod_data else "N/A"),
    ]
    
    for label, value in prod_metrics:
        ws6.cell(row=row, column=1, value=label).border = thin_border
        ws6.cell(row=row, column=2, value=value).border = thin_border
        row += 1
        
    row += 2
    
    # CFO Section
    ws6.cell(row=row, column=1, value="Finance (Budget)").font = section_font
    ws6.cell(row=row, column=1).fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Green
    ws6.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    row += 1
    
    cfo_metrics = [
        ("Budget Status", "Check Finance Dashboard"),
        ("Liquidity Status", cfo_data.get('liquidity_status', 'Unknown') if cfo_data else "Unknown"),
    ]
    
    for label, value in cfo_metrics:
        ws6.cell(row=row, column=1, value=label).border = thin_border
        ws6.cell(row=row, column=2, value=value).border = thin_border
        row += 1

    # Formatting
    for col in ['A', 'B']:
        ws6.column_dimensions[col].width = 30

    # Save
    # Save to buffer or file
    if output_buffer is not None:
        wb.save(output_buffer)
        output_buffer.seek(0)
        print("[SUCCESS] Created dashboard in BytesIO buffer")
    else:
        wb.save(OUTPUT_FILE)
        print(f"[SUCCESS] Created '{OUTPUT_FILE}'")
    print(f"[SUCCESS] Created '{OUTPUT_FILE}'")


def main():
    """Main function."""
    print("ExSim CMO Complete Dashboard Generator")
    print("=" * 50)
    
    print("\n[*] Loading data files...")
    from config import REPORTS_DIR, DATA_DIR
    print(f"    Primary source: {REPORTS_DIR}")
    print(f"    Fallback source: {DATA_DIR}")
    
    # Market Report
    market_path = get_data_path("market-report.xlsx")
    if market_path:
        market_data = load_market_report(market_path)
        print(f"  [OK] Loaded market report from {market_path.parent.name}/")
    else:
        market_data = load_market_report(None)
        print("  [!] Using default market data")
    
    # Innovation Template
    innov_path = get_data_path("Marketing Innovation Decisions.xlsx")
    if innov_path:
        innovation_features = load_innovation_features(innov_path)
        print(f"  [OK] Loaded {len(innovation_features)} innovation features")
    else:
        innovation_features = load_innovation_features(None)
        print("  [!] Using default innovation features")
    
    # Marketing Template
    mkt_path = get_data_path("Marketing Decisions.xlsx")
    if mkt_path:
        marketing_template = load_marketing_template(mkt_path)
        print(f"  [OK] Loaded marketing template")
    else:
        marketing_template = load_marketing_template(None)
        print("  [!] Using default marketing template")
    
    # Sales Data
    sales_path = get_data_path("sales_admin_expenses.xlsx")
    if sales_path:
        sales_data = load_sales_admin_expenses(sales_path)
        print(f"  [OK] Loaded sales data")
    else:
        sales_data = load_sales_admin_expenses(None)
        print("  [!] Using default sales data")
    
    # Inventory
    inv_path = get_data_path("finished_goods_inventory.xlsx")
    if inv_path:
        inventory_data = load_finished_goods_inventory(inv_path)
        stockout_status = "STOCKOUT DETECTED" if inventory_data['is_stockout'] else "OK"
        print(f"  [OK] Loaded inventory: {stockout_status}")
    else:
        inventory_data = load_finished_goods_inventory(None)
        print("  [!] Using default inventory data")
    
    # NEW: Marketing Intelligence (Unit Economics)
    print("  [*] Parsing Marketing Intelligence (Unit Economics)...")
    marketing_intelligence = load_marketing_intelligence(sales_path, market_path)
    
    print("\n[*] Generating CMO Dashboard...")
    
    create_complete_dashboard(market_data, innovation_features, marketing_template,
                              sales_data, inventory_data, marketing_intelligence)
    
    print("\nSheets created:")
    print("  * SEGMENT_PULSE (High/Low Segment Analysis)")
    print("  * INNOVATION_LAB (Feature Selection)")
    print("  * STRATEGY_COCKPIT (4 Ps Decisions + ROI)")
    print("  * UPLOAD_READY_MARKETING (ExSim Format)")
    print("  * UPLOAD_READY_INNOVATION (ExSim Format)")
    
    # Export key metrics for downstream dashboards
    if export_dashboard_data:
        # Safely calculate innovation costs
        try:
            innov_costs = sum(f.get('cost', 0) for f in innovation_features if isinstance(f, dict) and f.get('selected', False))
        except:
            innov_costs = 0
        
        export_dashboard_data('CMO', {
            'demand_forecast': {zone: market_data.get('zones', {}).get(zone, {}).get('demand', 0) for zone in ZONES},
            'marketing_spend': sum(sales_data.get(fn, {}).get('advertising', 0) for fn in range(1, 9)),
            'pricing': {zone: market_data.get('zones', {}).get(zone, {}).get('my_price', 0) for zone in ZONES},
            'innovation_costs': innov_costs
        })


if __name__ == "__main__":
    main()
