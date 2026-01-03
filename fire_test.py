"""
ExSim Fire Test - End-to-End Dashboard Testing

This script performs a comprehensive "fire test" of all dashboard generators:
1. Generates mock data in test_data/mock_reports/
2. Temporarily redirects dashboard generators to use mock data
3. Runs all 7 dashboard generators
4. Validates outputs exist and have correct structure
5. Reports success/failure

Usage:
    python fire_test.py [--seed 42] [--keep-output]
"""

import os
import sys
import subprocess
import shutil
from pathlib import Path
from openpyxl import load_workbook
import argparse

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = Path(__file__).parent
MOCK_DATA_DIR = BASE_DIR / "test_data" / "mock_reports"
MOCK_OUTPUT_DIR = BASE_DIR / "test_data" / "mock_outputs"

DASHBOARDS = {
    "CFO": {
        "dir": "CFO Dashboard",
        "script": "generate_finance_dashboard_final.py",
        "output": "Finance_Dashboard_Final.xlsx",
        "expected_sheets": ["LIQUIDITY_MONITOR", "PROFIT_CONTROL", "UPLOAD_READY_FINANCE"]
    },
    "CLO": {
        "dir": "CLO Dashboard",
        "script": "generate_logistics_dashboard.py",
        "output": "Logistics_Dashboard.xlsx",
        "expected_sheets": ["ROUTE_CONFIG", "INVENTORY_TETRIS", "UPLOAD_READY_LOGISTICS"]
    },
    "CPO": {
        "dir": "CPO Dashboard",
        "script": "generate_cpo_dashboard.py",
        "output": "CPO_Dashboard.xlsx",
        "expected_sheets": ["WORKFORCE_PLANNING", "COMPENSATION_STRATEGY", "UPLOAD_READY_PEOPLE"]
    },
    "CMO": {
        "dir": "CMO Dashboard",
        "script": "generate_cmo_dashboard_complete.py",
        "output": "CMO_Dashboard_Complete.xlsx",
        "expected_sheets": ["STRATEGY_COCKPIT", "INNOVATION_LAB", "UPLOAD_READY_MARKETING"]
    },
    "Purchasing": {
        "dir": "Purchasing Role",
        "script": "generate_purchasing_dashboard_v2.py",
        "output": "Purchasing_Dashboard.xlsx",
        "expected_sheets": ["MRP_ENGINE", "SUPPLIER_CONFIG", "UPLOAD_READY_PROCUREMENT"]
    },
    "ESG": {
        "dir": "ESG Dashboard",
        "script": "generate_esg_dashboard.py",
        "output": "ESG_Dashboard.xlsx",
        "expected_sheets": ["STRATEGY_SELECTOR", "UPLOAD_READY_ESG"]
    },
    "Production": {
        "dir": "Production Manager Dashboard",
        "script": "generate_production_dashboard_zones.py",
        "output": "Production_Dashboard_Zones.xlsx",
        "expected_sheets": ["ZONE_CALCULATORS", "UPLOAD_READY_PRODUCTION"]
    }
}


# =============================================================================
# TEST FUNCTIONS
# =============================================================================

def generate_mock_data(seed):
    """Step 1: Generate mock data."""
    print("\n" + "=" * 60)
    print("STEP 1: Generating Mock Data")
    print("=" * 60)
    
    from test_data.generate_mock_data import generate_all_mock_data
    generate_all_mock_data(seed=seed, output_dir=MOCK_DATA_DIR)
    
    # Verify files exist
    expected_files = [
        "workers_balance_overtime.xlsx",
        "sales_admin_expenses.xlsx",
        "finished_goods_inventory.xlsx",
        "initial_cash_flow.xlsx",
        "production.xlsx",
        "raw_materials.xlsx",
        "machine_spaces.xlsx",
        "accounts_receivable_payable.xlsx",
        "logistics.xlsx",
        "results_and_balance_statements.xlsx",
        "market-report.xlsx",
        "ESG.xlsx",
        "subperiod_cash_flow.xlsx",
    ]
    
    missing = [f for f in expected_files if not (MOCK_DATA_DIR / f).exists()]
    if missing:
        print(f"  [FAIL] Missing mock files: {missing}")
        return False
    
    print(f"  [OK] All {len(expected_files)} mock files generated")
    return True


def setup_mock_data_path():
    """Step 2: Set up environment to use mock data."""
    print("\n" + "=" * 60)
    print("STEP 2: Configuring Mock Data Path")
    print("=" * 60)
    
    # Set environment variable that dashboard generators can check
    os.environ["EXSIM_REPORTS_PATH"] = str(MOCK_DATA_DIR)
    print(f"  [OK] EXSIM_REPORTS_PATH = {MOCK_DATA_DIR}")
    
    # Create output directory
    MOCK_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"  [OK] Output directory: {MOCK_OUTPUT_DIR}")
    
    return True


def run_dashboard_generator(name, info):
    """Run a single dashboard generator."""
    script_path = BASE_DIR / info["dir"] / info["script"]
    working_dir = BASE_DIR / info["dir"]
    
    if not script_path.exists():
        return False, f"Script not found: {script_path}"
    
    try:
        # Run with environment variable set
        env = os.environ.copy()
        env["EXSIM_REPORTS_PATH"] = str(MOCK_DATA_DIR)
        
        result = subprocess.run(
            [sys.executable, str(script_path)],
            cwd=str(working_dir),
            capture_output=True,
            text=True,
            timeout=120,
            env=env
        )
        
        if result.returncode == 0:
            return True, "Generated successfully"
        else:
            return False, f"Error: {result.stderr[:500]}"
    except subprocess.TimeoutExpired:
        return False, "Generation timed out"
    except Exception as e:
        return False, str(e)


def run_all_generators():
    """Step 3: Run all dashboard generators."""
    print("\n" + "=" * 60)
    print("STEP 3: Running Dashboard Generators")
    print("=" * 60)
    
    results = {}
    
    for name, info in DASHBOARDS.items():
        print(f"\n  [{name}]", end=" ")
        success, message = run_dashboard_generator(name, info)
        results[name] = {"success": success, "message": message}
        
        if success:
            print(f"[OK] {message}")
        else:
            print(f"[FAIL] {message}")
    
    return results


def validate_outputs(generator_results):
    """Step 4: Validate generated dashboards."""
    print("\n" + "=" * 60)
    print("STEP 4: Validating Dashboard Outputs")
    print("=" * 60)
    
    validation_results = {}
    
    for name, info in DASHBOARDS.items():
        output_path = BASE_DIR / info["dir"] / info["output"]
        
        print(f"\n  [{name}]", end=" ")
        
        # Skip if generation failed
        if not generator_results.get(name, {}).get("success", False):
            print("[SKIP] Generation failed")
            validation_results[name] = {"valid": False, "reason": "Generation failed"}
            continue
        
        # Check file exists
        if not output_path.exists():
            print(f"[FAIL] Output file not found: {output_path}")
            validation_results[name] = {"valid": False, "reason": "File not found"}
            continue
        
        # Check sheets exist
        try:
            wb = load_workbook(output_path, read_only=True)
            missing_sheets = [s for s in info["expected_sheets"] if s not in wb.sheetnames]
            wb.close()
            
            if missing_sheets:
                print(f"[FAIL] Missing sheets: {missing_sheets}")
                validation_results[name] = {"valid": False, "reason": f"Missing sheets: {missing_sheets}"}
            else:
                print(f"[OK] All {len(info['expected_sheets'])} expected sheets present")
                validation_results[name] = {"valid": True, "reason": "All checks passed"}
        except Exception as e:
            print(f"[FAIL] Error opening file: {e}")
            validation_results[name] = {"valid": False, "reason": str(e)}
    
    return validation_results


def print_summary(generator_results, validation_results):
    """Print final test summary."""
    print("\n" + "=" * 60)
    print("FIRE TEST SUMMARY")
    print("=" * 60)
    
    gen_passed = sum(1 for r in generator_results.values() if r.get("success"))
    gen_failed = len(generator_results) - gen_passed
    
    val_passed = sum(1 for r in validation_results.values() if r.get("valid"))
    val_failed = len(validation_results) - val_passed
    
    print(f"\n  Generation: {gen_passed}/{len(generator_results)} passed")
    print(f"  Validation: {val_passed}/{len(validation_results)} passed")
    
    if gen_failed > 0 or val_failed > 0:
        print("\n  FAILURES:")
        for name in DASHBOARDS:
            if not generator_results.get(name, {}).get("success"):
                print(f"    - {name}: {generator_results.get(name, {}).get('message', 'Unknown error')}")
            elif not validation_results.get(name, {}).get("valid"):
                print(f"    - {name}: {validation_results.get(name, {}).get('reason', 'Unknown error')}")
    
    all_passed = gen_passed == len(generator_results) and val_passed == len(validation_results)
    
    print("\n" + "=" * 60)
    if all_passed:
        print("[SUCCESS] All fire tests passed!")
    else:
        print("[FAILURE] Some tests failed - see details above")
    print("=" * 60)
    
    return all_passed


def cleanup(keep_output):
    """Clean up generated files."""
    if not keep_output:
        print("\n  Cleaning up mock data...")
        if MOCK_DATA_DIR.exists():
            shutil.rmtree(MOCK_DATA_DIR)
        if MOCK_OUTPUT_DIR.exists():
            shutil.rmtree(MOCK_OUTPUT_DIR)
        print("  [OK] Cleaned up")
    else:
        print(f"\n  Mock data kept at: {MOCK_DATA_DIR}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="ExSim Fire Test - End-to-End Dashboard Testing")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for mock data")
    parser.add_argument("--keep-output", action="store_true", help="Keep generated mock data after test")
    args = parser.parse_args()
    
    print("=" * 60)
    print("ExSim FIRE TEST")
    print("=" * 60)
    print(f"Seed: {args.seed}")
    print(f"Testing {len(DASHBOARDS)} dashboards")
    
    # Step 1: Generate mock data
    if not generate_mock_data(args.seed):
        print("\n[ABORT] Mock data generation failed")
        return 1
    
    # Step 2: Set up mock data path
    if not setup_mock_data_path():
        print("\n[ABORT] Failed to set up mock data path")
        return 1
    
    # Step 3: Run all generators
    generator_results = run_all_generators()
    
    # Step 4: Validate outputs
    validation_results = validate_outputs(generator_results)
    
    # Step 5: Print summary
    all_passed = print_summary(generator_results, validation_results)
    
    # Step 6: Cleanup
    cleanup(args.keep_output)
    
    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
