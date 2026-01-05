import os
import sys
import openpyxl
from pathlib import Path

# Add project root and subdirectories to path
ROOT_DIR = Path(os.getcwd())
sys.path.append(str(ROOT_DIR))
sys.path.append(str(ROOT_DIR / "Purchasing Role"))
sys.path.append(str(ROOT_DIR / "CLO Dashboard"))
sys.path.append(str(ROOT_DIR / "CMO Dashboard"))
sys.path.append(str(ROOT_DIR / "CPO Dashboard"))
sys.path.append(str(ROOT_DIR / "ESG Dashboard"))
sys.path.append(str(ROOT_DIR / "CFO Dashboard"))

# Import Generators
# Using simple imports after adding to path

# 1. Purchasing
try:
    from generate_purchasing_dashboard_v2 import create_purchasing_dashboard
except ImportError as e:
    print(f"Error importing Purchasing generator: {e}")

# 2. Logistics (CLO)
try:
    from generate_logistics_dashboard import create_logistics_dashboard
except ImportError as e:
    print(f"Error importing Logistics generator: {e}")

# 3. CMO
try:
    from generate_cmo_dashboard_complete import create_complete_dashboard
    from generate_cmo_dashboard_complete import load_market_report, load_innovation_template, load_marketing_template, load_sales_data, load_inventory_data, load_marketing_intelligence
except ImportError as e:
    print(f"Error importing CMO generator: {e}")

# 4. CPO
try:
    from generate_cpo_dashboard import create_cpo_dashboard
    from generate_cpo_dashboard import load_workers_balance, load_sales_admin, load_labor_costs, load_absenteeism_data
except ImportError as e:
    print(f"Error importing CPO generator: {e}")

# 5. ESG
try:
    from generate_esg_dashboard import create_esg_dashboard
    from generate_esg_dashboard import load_esg_report, load_production_data
except ImportError as e:
    print(f"Error importing ESG generator: {e}")

# 6. CFO
try:
    from generate_finance_dashboard_final import create_finance_dashboard
    from generate_finance_dashboard_final import load_initial_cash_flow, load_balance_statements, load_sales_admin_expenses, load_receivables_payables, load_finance_template
except ImportError as e:
    print(f"Error importing CFO generator: {e}")

OUTPUT_DIR = Path("test_outputs_cross_ref")
OUTPUT_DIR.mkdir(exist_ok=True)

def verify_file(filename, expected_tab="CROSS_REFERENCE"):
    path = OUTPUT_DIR / filename
    if not path.exists():
        print(f"[FAIL] {filename} was not created.")
        return False
    
    try:
        wb = openpyxl.load_workbook(path)
        if expected_tab in wb.sheetnames:
            print(f"[SUCCESS] {filename} contains tab '{expected_tab}'")
            return True
        else:
            print(f"[FAIL] {filename} missing tab '{expected_tab}'. Found: {wb.sheetnames}")
            return False
    except Exception as e:
        print(f"[ERROR] reading {filename}: {e}")
        return False

def run_tests():
    print("Running Cross Reference Verification...")
    print("=" * 50)
    
    # --- Purchasing ---
    print("\n[Purchasing] Generating...")
    try:
        mock_rm = {'inventory': {}, 'lead_times': {}}
        mock_cost = {'consumption_cost': 100000}
        
        if 'create_purchasing_dashboard' in globals():
            create_purchasing_dashboard(mock_rm, mock_cost)
            
            if Path("Purchasing_Dashboard_v2_Standalone.xlsx").exists():
                 os.replace("Purchasing_Dashboard_v2_Standalone.xlsx", OUTPUT_DIR / "Purchasing.xlsx")
                 verify_file("Purchasing.xlsx")
            else:
                 print("[FAIL] Failed to generate Purchasing file")
        else:
             print("[SKIP] Purchasing generator not loaded")
             
    except Exception as e:
        print(f"[CRASH] Purchasing: {e}")

    # --- Logistics ---
    print("\n[Logistics] Generating...")
    try:
        mock_fg = {'zones': {}, 'is_stockout': False}
        mock_log = {'benchmarks': {}, 'penalties': {}}
        
        if 'create_logistics_dashboard' in globals():
            create_logistics_dashboard(mock_fg, mock_log)
            
            if Path("Detailed_Logistics_Dashboard.xlsx").exists():
                 os.replace("Detailed_Logistics_Dashboard.xlsx", OUTPUT_DIR / "Logistics.xlsx")
                 verify_file("Logistics.xlsx")
            else:
                 print("[FAIL] Failed to generate Logistics file")
        else:
            print("[SKIP] Logistics generator not loaded")
    except Exception as e:
        print(f"[CRASH] Logistics: {e}")

    # --- CMO ---
    print("\n[CMO] Generating...")
    try:
        if 'create_complete_dashboard' in globals():
            mkt = load_market_report(None)
            inn = load_innovation_template(None)
            tmpl = load_marketing_template(None)
            sales = load_sales_data(None)
            inv = load_inventory_data(None)
            intel = load_marketing_intelligence(None, None)
            
            create_complete_dashboard(mkt, inn, tmpl, sales, inv, intel)
            
            if Path("CMO_Dashboard_Complete.xlsx").exists():
                 os.replace("CMO_Dashboard_Complete.xlsx", OUTPUT_DIR / "CMO.xlsx")
                 verify_file("CMO.xlsx")
            else:
                 print("[FAIL] Failed to generate CMO file")
        else:
            print("[SKIP] CMO generator not loaded")
    except Exception as e:
        print(f"[CRASH] CMO: {e}")

    # --- CPO ---
    print("\n[CPO] Generating...")
    try:
        if 'create_cpo_dashboard' in globals():
            work = load_workers_balance(None)
            sales = load_sales_admin(None)
            labor = load_labor_costs(None)
            absent = load_absenteeism_data(None)
            
            create_cpo_dashboard(work, sales, labor, absent)
            
            if Path("CPO_Dashboard.xlsx").exists():
                 os.replace("CPO_Dashboard.xlsx", OUTPUT_DIR / "CPO.xlsx")
                 verify_file("CPO.xlsx")
            else:
                 print("[FAIL] Failed to generate CPO file")
        else:
            print("[SKIP] CPO generator not loaded")
            
    except Exception as e:
        print(f"[CRASH] CPO: {e}")
        
    # --- ESG ---
    print("\n[ESG] Generating...")
    try:
        if 'create_esg_dashboard' in globals():
            esg = load_esg_report(None)
            prod = load_production_data(None)
            
            create_esg_dashboard(esg, prod)
            
            if Path("ESG_Dashboard.xlsx").exists():
                 os.replace("ESG_Dashboard.xlsx", OUTPUT_DIR / "ESG.xlsx")
                 verify_file("ESG.xlsx")
            else:
                 print("[FAIL] Failed to generate ESG file")
        else:
            print("[SKIP] ESG generator not loaded")
    except Exception as e:
        print(f"[CRASH] ESG: {e}")

    # --- CFO ---
    print("\n[CFO] Generating...")
    try:
        if 'create_finance_dashboard' in globals():
            # Load with defaults (None triggers internal defaults)
            cash = load_initial_cash_flow(None)
            bal = load_balance_statements(None)
            sa = load_sales_admin_expenses(None)
            ar_ap = load_receivables_payables(None)
            tmpl = load_finance_template(None)
            
            create_finance_dashboard(cash, bal, sa, ar_ap, tmpl)
            
            if Path("Finance_Dashboard_Final.xlsx").exists():
                 os.replace("Finance_Dashboard_Final.xlsx", OUTPUT_DIR / "CFO.xlsx")
                 verify_file("CFO.xlsx", expected_tab="CROSS_REFERENCE")
            else:
                 print("[FAIL] Failed to generate CFO file")
        else:
            print("[SKIP] CFO generator not loaded")
    except Exception as e:
        print(f"[CRASH] CFO: {e}")

if __name__ == "__main__":
    run_tests()
